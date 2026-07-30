"""
SAP RPA MMSC — UI Cyber Edition v3
Tema: Cyber Dark · VS Code Light
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import subprocess
import os
import re
import openpyxl

PYTHON      = r"C:\Users\USER\AppData\Local\Programs\Python\Python314\python.exe"
SCRIPT      = "sap_rpa_mmsc_extend.py"
MASTER_FILE = "master_data.xlsx"

# ── CYBER DARK v3 ─────────────────────────────────
DARK = {
    "bg":        "#060B12",   # void black
    "bg2":       "#0A1220",   # card navy
    "bg3":       "#0F1A2C",   # inner
    "bg4":       "#162438",   # hover
    "fg":        "#EAF2FF",   # ice white
    "fg2":       "#5C85AE",   # muted
    "fg3":       "#243650",   # disabled
    "accent":    "#00D8FF",   # neon cyan
    "accent2":   "#0099BB",   # cyan mid
    "gold":      "#FFC020",   # gold spark
    "gold2":     "#FFE070",   # gold bright
    "green":     "#00EE80",   # neon mint
    "red":       "#FF3D55",   # neon red
    "amber":     "#FFA020",   # amber
    "teal":      "#00D8FF",
    "border":    "#0C2040",
    "border2":   "#070F1E",
    "chk_sel":   "#081C38",
    "chk_bg":    "#0A1220",
    "btn_pri":   "#003E5C",
    "btn_stop":  "#5C0010",
    "hdr_bg":    "#070E1A",
    "log_bg":    "#050A14",
    "log_fg":    "#B8D8FF",
    "sb_bg":     "#0C1A2C",   # scrollbar track
    "sb_fg":     "#1E4060",   # scrollbar thumb
    "name":      "🌙  GELAP",
}

# ── VS CODE LIGHT v3 ───────────────────────────────
LIGHT = {
    "bg":        "#F5F5F5",   # VS Code sidebar
    "bg2":       "#FFFFFF",   # card putih
    "bg3":       "#ECECEC",   # input
    "bg4":       "#E0E0E0",   # hover
    "fg":        "#111111",   # teks sangat pekat
    "fg2":       "#424242",   # sekunder
    "fg3":       "#909090",   # muted
    "accent":    "#0068C0",   # VS Code blue
    "accent2":   "#004E99",
    "gold":      "#7A5800",
    "gold2":     "#5C4200",
    "green":     "#0C7A0C",
    "red":       "#C41A1A",
    "amber":     "#7A4A00",
    "teal":      "#0068C0",
    "border":    "#C8C8C8",
    "border2":   "#DEDEDE",
    "chk_sel":   "#C4DCFF",
    "chk_bg":    "#FFFFFF",
    "btn_pri":   "#0068C0",
    "btn_stop":  "#B81818",
    "hdr_bg":    "#E8E8E8",
    "log_bg":    "#1E1E1E",
    "log_fg":    "#D4D4D4",
    "sb_bg":     "#E8E8E8",
    "sb_fg":     "#B8B8B8",
    "name":      "☀  TERANG",
}

# ── MIDNIGHT BLUE — tema ke-3 baru ─────────────────
MIDNIGHT = {
    "bg":       "#0A0F1E",   # deep navy
    "bg2":      "#0F1830",   # card
    "bg3":      "#141E3C",   # inner
    "bg4":      "#1A2548",   # hover
    "fg":       "#C8D8F8",   # lavender white
    "fg2":      "#7080A8",   # muted
    "fg3":      "#2A3560",   # disabled
    "accent":   "#6088FF",   # electric blue/purple
    "accent2":  "#4060CC",
    "gold":     "#F0A030",
    "gold2":    "#F8C060",
    "green":    "#40E0A0",   # aqua green
    "red":      "#FF5060",
    "amber":    "#F08020",
    "teal":     "#6088FF",
    "border":   "#1A2850",
    "border2":  "#0C1428",
    "chk_sel":  "#102040",
    "chk_bg":   "#0F1830",
    "btn_pri":  "#203080",
    "btn_stop": "#601020",
    "hdr_bg":   "#080C18",
    "log_bg":   "#060A16",
    "log_fg":   "#C0D0F0",
    "name":     "☀  TERANG",
}

THEMES = [DARK, MIDNIGHT, LIGHT]
THEME_NAMES = ["🌙 Gelap", "🔵 Midnight", "☀ Terang"]


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("SAP RPA — MMSC Extend")
        self.root.geometry("960x980")
        self.root.minsize(800, 800)
        self.root.resizable(True, True)

        self.dry_var    = tk.BooleanVar(value=False)
        self.status_var = tk.StringVar(value="⬡  Siap dijalankan")
        self.is_dark    = True
        self._theme_idx = 0
        self.T          = THEMES[0]
        self.is_dark    = True
        self.process    = None
        self.master_path = MASTER_FILE
        self.mat_vars   = []
        self.plant_vars = []
        self._all_widgets = []
        self._hdr_left  = None
        self._splits    = []
        self._cards     = []
        self._secs      = []

        # animasi topbar gradient
        self._tb_pos  = 0
        self._tb_anim = True

        # tracking proses
        self._start_time    = None
        self._final_elapsed = 0
        self._ok_count      = 0
        self._err_count     = 0
        self._total_count   = 0
        self._counter_var   = tk.StringVar(value="")
        self._skeleton_anim = False

        self._build()
        self._apply_theme()
        self._load_master()
        self._topbar_anim()   # mulai animasi

    # ──────────────────────────────────────────
    # BUILD UI
    # ──────────────────────────────────────────
    def _build(self):
        T = self.T

        # ── TOP ANIMATED BAR ─────────────────
        self.topbar = tk.Canvas(self.root, height=5, highlightthickness=0)
        self.topbar.pack(fill="x")

        # ── HEADER ───────────────────────────
        self.hdr = tk.Frame(self.root, pady=12, padx=22)
        self.hdr.pack(fill="x")

        left = tk.Frame(self.hdr)
        left.pack(side="left", fill="x", expand=True)
        self._hdr_left = left

        # Badge SAP
        self.lbl_badge = tk.Label(left, text="  SAP  ",
                                   font=("Segoe UI", 11, "bold"),
                                   padx=6, pady=5)
        self.lbl_badge.pack(side="left", anchor="center")

        # Title kolom
        title_col = tk.Frame(left)
        title_col.pack(side="left", anchor="center", padx=(10,0))
        self._hdr_title_col = title_col

        self.lbl_title = tk.Label(title_col, text="RPA MMSC",
                                   font=("Segoe UI", 16, "bold"))
        self.lbl_title.pack(anchor="w")

        self.lbl_sub = tk.Label(title_col,
                                 text="Automated SAP Material Extension",
                                 font=("Segoe UI", 9))
        self.lbl_sub.pack(anchor="w")

        self.lbl_ver = tk.Label(left, text="@IT.Apps",
                                 font=("Segoe UI", 9))
        self.lbl_ver.pack(side="left", anchor="s", padx=(10,0), pady=(0,3))

        # ── KANAN: badge + theme btn ──────────
        right = tk.Frame(self.hdr)
        right.pack(side="right", anchor="center")
        self._hdr_right = right

        # Badge status dengan animasi pulse
        self.lbl_status_badge = tk.Label(right, text="⬡  READY",
                                          font=("Segoe UI", 9, "bold"),
                                          padx=12, pady=5)
        self.lbl_status_badge.pack(side="right", anchor="center", padx=(6,0))
        self._badge_pulse = False

        self.btn_theme = tk.Button(right,
                                    font=("Segoe UI", 10, "bold"),
                                    relief="flat", bd=0, cursor="hand2",
                                    padx=14, pady=8,
                                    command=self._toggle_theme)
        self.btn_theme.pack(side="right", anchor="center")
        self._bind_hover(self.btn_theme)
        self._add_tooltip(self.btn_theme, "Toggle Dark / Light mode")

        # ── DIVIDER 3px gradient feel ─────────
        self.div = tk.Frame(self.root, height=2)
        self.div.pack(fill="x")

        # ── BODY ─────────────────────────────
        self.body = tk.Frame(self.root, padx=16, pady=2)
        self.body.pack(fill="both", expand=True)
        self.body.columnconfigure(0, weight=1)
        self.body.rowconfigure(1, weight=0)   # master data — fixed
        self.body.rowconfigure(3, weight=1)   # material+plant — expand
        self.body.rowconfigure(5, weight=0)   # eksekusi — fixed
        self.body.rowconfigure(7, weight=1)   # log — expand

        row = 0

        # ── MASTER FILE ──────────────────────
        self._label_row(row, "⬡  MASTER DATA"); row += 1
        c0 = tk.Frame(self.body, padx=10, pady=2, highlightthickness=1)
        c0.grid(row=row, column=0, sticky="ew", pady=(0,1)); row += 1
        if not hasattr(self, "_cards"):
            self._cards = []
        self._cards.append(c0)

        r0 = tk.Frame(c0)
        r0.pack(fill="x")
        self.lbl_master = tk.Label(r0, font=("Segoe UI", 11), anchor="w")
        self.lbl_master.pack(side="left")
        self.lbl_info = tk.Label(r0, font=("Segoe UI", 11))
        self.lbl_info.pack(side="left", padx=(10,0))
        r0.bind("<Configure>", lambda e: self.lbl_master.configure(
            wraplength=max(100, e.width - 280)))
        self.btn_ganti = self._mk_btn(r0, "Ganti File",
                                       self._ganti_master, side="right",
                                       small=True)

        # ── PILIH MATERIAL & PLANT (side by side) ────────
        self._label_row(row, "⬡  PILIH MATERIAL  /  PILIH PLANT"); row += 1

        # wrapper frame span 2 panel
        split = tk.Frame(self.body)
        split.grid(row=row, column=0, sticky="nsew", pady=(0,2)); row += 1
        split.columnconfigure(0, weight=1)
        split.columnconfigure(1, weight=2)
        split.rowconfigure(0, weight=1)

        # ── Panel kiri: MATERIAL ─────────────
        c1 = tk.Frame(split, padx=14, pady=10, highlightthickness=1)
        c1.grid(row=0, column=0, sticky="nsew", padx=(0,4))
        if not hasattr(self, "_cards"):
            self._cards = []
        self._cards.append(c1)

        tb1 = tk.Frame(c1)
        tb1.pack(fill="x", pady=(0,4))
        self.btn_mat_all  = self._mk_btn(tb1, "✓ Semua",
                                          lambda: self._sel_all(self.mat_vars, self._upd_mat),
                                          side="left", small=True, gold=True)
        self.btn_mat_none = self._mk_btn(tb1, "✗ Kosong",
                                          lambda: self._sel_none(self.mat_vars, self._upd_mat),
                                          side="left", small=True, pl=4)
        self.lbl_mat_sel = tk.Label(tb1, font=("Segoe UI", 11))
        self.lbl_mat_sel.pack(side="right")

        # Search box material
        self._mat_search_var = tk.StringVar()
        self._mat_search_var.trace_add("write", lambda *a: self._filter_chk(
            self.mat_vars, self._mat_search_var.get()))
        sf1 = tk.Frame(c1)
        sf1.pack(fill="x", pady=(0,4))
        tk.Label(sf1, text="🔍", font=("Segoe UI", 9)).pack(side="left")
        self.ent_mat_search = tk.Entry(sf1, textvariable=self._mat_search_var,
                                        font=("Segoe UI", 10), relief="flat",
                                        bd=1)
        self.ent_mat_search.pack(side="left", fill="x", expand=True, padx=(4,0))
        self._add_tooltip(self.ent_mat_search, "Filter material by keyword")

        self.mat_canvas, self.mat_frame = self._mk_chk_area(c1, height=120)
        self.mat_canvas.bind("<Configure>", lambda e: self._reflow_chk(
            self.mat_canvas, self.mat_vars, self._upd_mat))

        # ── Panel kanan: PLANT ───────────────
        c2 = tk.Frame(split, padx=14, pady=10, highlightthickness=1)
        c2.grid(row=0, column=1, sticky="nsew", padx=(4,0))
        self._cards.append(c2)

        tb2 = tk.Frame(c2)
        tb2.pack(fill="x", pady=(0,4))
        self.btn_plt_all  = self._mk_btn(tb2, "✓ Semua",
                                          lambda: self._sel_all(self.plant_vars, self._upd_plt),
                                          side="left", small=True, gold=True)
        self.btn_plt_none = self._mk_btn(tb2, "✗ Kosong",
                                          lambda: self._sel_none(self.plant_vars, self._upd_plt),
                                          side="left", small=True, pl=4)
        self.lbl_plt_sel = tk.Label(tb2, font=("Segoe UI", 11))
        self.lbl_plt_sel.pack(side="right")

        # Search box plant
        self._plt_search_var = tk.StringVar()
        self._plt_search_var.trace_add("write", lambda *a: self._filter_chk(
            self.plant_vars, self._plt_search_var.get()))
        sf2 = tk.Frame(c2)
        sf2.pack(fill="x", pady=(0,4))
        tk.Label(sf2, text="🔍", font=("Segoe UI", 9)).pack(side="left")
        self.ent_plt_search = tk.Entry(sf2, textvariable=self._plt_search_var,
                                        font=("Segoe UI", 10), relief="flat",
                                        bd=1)
        self.ent_plt_search.pack(side="left", fill="x", expand=True, padx=(4,0))
        self._add_tooltip(self.ent_plt_search, "Filter plant by keyword")

        self.plt_canvas, self.plt_frame = self._mk_chk_area(c2, height=120)
        self.plt_canvas.bind("<Configure>", lambda e: self._reflow_chk(
            self.plt_canvas, self.plant_vars, self._upd_plt))

        # simpan split frame dan search frames untuk theming
        if not hasattr(self, "_splits"):
            self._splits = []
        self._splits.append(split)
        if not hasattr(self, "_search_frames"):
            self._search_frames = []
        self._search_frames.extend([sf1, sf2])

        # ── EKSEKUSI (compact, tanpa label row terpisah) ──
        self._label_row(row, "⬡  EKSEKUSI"); row += 1
        c3 = tk.Frame(self.body, padx=10, pady=2, highlightthickness=1)
        c3.grid(row=row, column=0, sticky="ew", pady=(0,1)); row += 1
        self._cards.append(c3)

        # Semua dalam 1 baris: dry run · status · WH02
        r3 = tk.Frame(c3)
        r3.pack(fill="x")
        self.chk_dry = tk.Checkbutton(r3, variable=self.dry_var,
                                       text="Simulasi",
                                       font=("Segoe UI", 11),
                                       relief="flat", cursor="hand2")
        self.chk_dry.pack(side="left")
        self.lbl_status = tk.Label(r3, textvariable=self.status_var,
                                    font=("Segoe UI", 11))
        self.lbl_status.pack(side="left", padx=(10,0))
        self.lbl_sloc = tk.Label(r3, text="WH02 & WT01",
                                  font=("Segoe UI", 11))
        self.lbl_sloc.pack(side="right")

        # Counter detail real-time
        self._detail_var = tk.StringVar(value="")
        r3b = tk.Frame(c3)
        r3b.pack(fill="x", pady=(1,0))
        self.lbl_detail = tk.Label(r3b, textvariable=self._detail_var,
                                    font=("Segoe UI", 9),
                                    anchor="w")
        self.lbl_detail.pack(side="left")
        self.lbl_detail_right = tk.Label(r3b, text="",
                                          font=("Segoe UI", 9, "bold"),
                                          anchor="e")
        self.lbl_detail_right.pack(side="right")

        # Progress bar 2px
        self.pb_frame = tk.Frame(c3, height=2)
        self.pb_frame.pack(fill="x", pady=(2,2))
        self.pb_fill = tk.Frame(self.pb_frame, height=2, width=0)
        self.pb_fill.place(x=0, y=0, relheight=1)
        self._pb_anim = False

        # Tombol + warn satu baris
        br = tk.Frame(c3)
        br.pack(fill="x")
        self.btn_start = tk.Button(br, text="▶  MULAI RPA",
                                    font=("Segoe UI", 11, "bold"),
                                    relief="flat", cursor="hand2",
                                    padx=14, pady=3,
                                    command=self._start)
        self.btn_start.pack(side="left")
        self._bind_hover(self.btn_start)
        self._add_tooltip(self.btn_start, "Mulai proses RPA otomatis")

        self.btn_stop = tk.Button(br, text="■  STOP",
                                   font=("Segoe UI", 11, "bold"),
                                   relief="flat", cursor="hand2",
                                   state="disabled",
                                   padx=14, pady=3,
                                   command=self._stop)
        self.btn_stop.pack(side="left", padx=(5,0))
        self._bind_hover(self.btn_stop)
        self._add_tooltip(self.btn_stop, "Hentikan proses RPA")

        self.lbl_warn = tk.Label(br,
                                  text="⚠  Jangan sentuh SAP saat berjalan",
                                  font=("Segoe UI", 11))
        self.lbl_warn.pack(side="left", padx=(12,0))

        self.btn_out = tk.Button(br, text="📂  Output",
                                  font=("Segoe UI", 11),
                                  relief="flat", cursor="hand2",
                                  padx=8, pady=4,
                                  command=self._open_output)
        self.btn_out.pack(side="right")
        self._bind_hover(self.btn_out)
        self._add_tooltip(self.btn_out, "Buka folder output_rpa")

        # ── LOG ──────────────────────────────
        self._label_row(row, "⬡  LOG PROSES"); row += 1
        c4 = self._card(row, expand=True); row += 1

        self.log = scrolledtext.ScrolledText(c4, height=6,
                                              font=("Consolas", 10),
                                              relief="flat",
                                              state="disabled",
                                              wrap="word")
        self.log.pack(fill="both", expand=True)

        r_log = tk.Frame(c4)
        r_log.pack(fill="x", pady=(4,0))
        self.btn_clr = tk.Button(r_log, text="🗑  Bersihkan Log",
                                  font=("Segoe UI", 11),
                                  relief="flat", cursor="hand2",
                                  padx=6, pady=2,
                                  command=self._clear_log)
        self.btn_clr.pack(side="right")
        self._add_tooltip(self.btn_clr, "Hapus semua isi log")

        self.btn_export_log = tk.Button(r_log, text="💾  Export Log",
                                         font=("Segoe UI", 11),
                                         relief="flat", cursor="hand2",
                                         padx=6, pady=2,
                                         command=self._export_log)
        self.btn_export_log.pack(side="right", padx=(0,4))
        self._add_tooltip(self.btn_export_log, "Simpan log ke file .txt")

        tk.Frame(self.body, height=6).grid(row=row, column=0)

        # ── STATUS BAR BAWAH ─────────────────
        self.statusbar = tk.Frame(self.root, height=28)
        self.statusbar.pack(fill="x", side="bottom")

        self.lbl_counter = tk.Label(self.statusbar,
                                     textvariable=self._counter_var,
                                     font=("Segoe UI", 9, "bold"),
                                     anchor="w", padx=12)
        self.lbl_counter.pack(side="left", fill="y")

        self.lbl_elapsed = tk.Label(self.statusbar, text="",
                                     font=("Segoe UI", 9),
                                     anchor="e", padx=12)
        self.lbl_elapsed.pack(side="right", fill="y")

    # ──────────────────────────────────────────
    # WIDGET HELPERS
    # ──────────────────────────────────────────
    def _label_row(self, row, text):
        f = tk.Frame(self.body)
        f.grid(row=row, column=0, sticky="ew", pady=(4,1))
        # garis accent kiri
        bar = tk.Frame(f, width=3)
        bar.pack(side="left", fill="y", padx=(0,8))
        lbl = tk.Label(f, text=text, font=("Segoe UI", 11, "bold"))
        lbl.pack(side="left", anchor="w")
        if not hasattr(self, "_secs"):
            self._secs = []
        self._secs.append((f, lbl, bar))

    def _card(self, row, expand=False):
        f = tk.Frame(self.body, padx=14, pady=10, highlightthickness=1)
        if expand:
            f.grid(row=row, column=0, sticky="nsew", pady=(0,3))
        else:
            f.grid(row=row, column=0, sticky="ew", pady=(0,3))
        if not hasattr(self, "_cards"):
            self._cards = []
        self._cards.append(f)
        return f

    def _mk_btn(self, parent, text, cmd, side=None, small=False,
                gold=False, pl=0):
        font = ("Segoe UI", 8 if small else 10, "bold")
        padx = 12 if small else 20
        pady = 5 if small else 9
        b = tk.Button(parent, text=text, font=font,
                       relief="flat", cursor="hand2",
                       padx=padx, pady=pady, command=cmd)
        if side:
            b.pack(side=side, padx=(pl,0))
        self._bind_hover(b)
        return b

    def _bind_hover(self, btn, enter_bg=None, leave_bg=None):
        """Efek hover ringan pada tombol."""
        def on_enter(e):
            T = self.T
            bg = enter_bg or T["bg4"]
            try:
                btn.configure(bg=bg)
            except Exception:
                pass
        def on_leave(e):
            try:
                btn.configure(bg=btn._leave_bg if hasattr(btn, "_leave_bg") else leave_bg or self.T["bg3"])
            except Exception:
                pass
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)

    def _mk_chk_area(self, parent, height=100):
        outer = tk.Frame(parent, height=height)
        outer.pack(fill="both", expand=True)
        outer.pack_propagate(False)
        canvas = tk.Canvas(outer, highlightthickness=0)
        sb = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=sb.set)
        sb.pack(side="right", fill="y")
        canvas.pack(side="left", fill="both", expand=True)
        frame = tk.Frame(canvas)
        win = canvas.create_window((0,0), window=frame, anchor="nw")
        frame.bind("<Configure>", lambda e:
            canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e:
            canvas.itemconfig(win, width=e.width))
        canvas.bind("<MouseWheel>", lambda e:
            canvas.yview_scroll(-1*(e.delta//120), "units"))
        return canvas, frame

    # ──────────────────────────────────────────
    # TOOLTIP SYSTEM
    # ──────────────────────────────────────────
    def _add_tooltip(self, widget, text):
        """Tooltip muncul saat hover."""
        tip = None

        def show(e):
            nonlocal tip
            T = self.T
            tip = tk.Toplevel(widget)
            tip.wm_overrideredirect(True)
            tip.wm_geometry(f"+{e.x_root+12}+{e.y_root+18}")
            lbl = tk.Label(tip, text=text,
                           font=("Segoe UI", 8),
                           bg=T["bg4"], fg=T["fg"],
                           padx=8, pady=4,
                           relief="flat",
                           bd=1)
            lbl.pack()
            tip.configure(bg=T["border"])

        def hide(e):
            nonlocal tip
            if tip:
                try:
                    tip.destroy()
                except Exception:
                    pass
                tip = None

        widget.bind("<Enter>", show, add="+")
        widget.bind("<Leave>", hide, add="+")
        widget.bind("<ButtonPress>", hide, add="+")

    # ──────────────────────────────────────────
    # BADGE PULSE ANIMASI
    # ──────────────────────────────────────────
    def _pulse_badge(self, step=0):
        """Pulse animasi pada badge saat RUNNING."""
        if not self._badge_pulse:
            return
        T = self.T
        # Alternasi antara bright dan dim
        if step % 2 == 0:
            self.lbl_status_badge.configure(fg=T["accent"])
        else:
            self.lbl_status_badge.configure(fg=T["accent2"])
        self.root.after(600, self._pulse_badge, step + 1)

    # ──────────────────────────────────────────
    # PROGRESS BAR — gradient warna berubah
    # ──────────────────────────────────────────
    def _pb_start(self):
        self._pb_anim = True
        self._pb_pos  = 0
        self._pb_tick()

    def _pb_tick(self):
        if not self._pb_anim:
            return
        w = self.pb_frame.winfo_width()
        if w < 2:
            self.root.after(100, self._pb_tick)
            return
        bar_w = max(80, w // 4)
        self._pb_pos = (self._pb_pos + 5) % (w + bar_w)
        x = self._pb_pos - bar_w
        self.pb_fill.place(x=x, y=0, width=bar_w, relheight=1)
        # Warna berubah-ubah saat animasi
        colors = [self.T["accent"], self.T["gold"], self.T["green"],
                  self.T["accent2"], self.T["amber"]]
        step = (self._pb_pos // 40) % len(colors)
        self.pb_fill.configure(bg=colors[step])
        self.root.after(28, self._pb_tick)

    def _pb_stop(self):
        self._pb_anim = False
        self.pb_fill.place(x=0, y=0, width=0, relheight=1)

    # ──────────────────────────────────────────
    # TOPBAR GRADIENT ANIMASI — smooth interpolasi
    # ──────────────────────────────────────────
    def _topbar_anim(self):
        if not self._tb_anim:
            return
        try:
            w = self.topbar.winfo_width()
            if w < 2:
                self.root.after(50, self._topbar_anim)
                return
            self.topbar.delete("all")
            T = self.T
            # Palet gradient sesuai tema
            if self.is_dark:
                colors = [T["accent"], T["gold"], T["green"],
                          T["accent2"], T["amber"], T["accent"]]
            else:
                colors = [T["accent"], T["accent2"], T["accent"],
                          "#4CA8FF", T["accent2"], T["accent"]]

            # Gambar gradient smooth dengan banyak segment
            segs = 120
            seg_w = (w * 2) / segs
            pos = self._tb_pos % w

            def hex_lerp(c1, c2, t):
                r1,g1,b1 = int(c1[1:3],16),int(c1[3:5],16),int(c1[5:7],16)
                r2,g2,b2 = int(c2[1:3],16),int(c2[3:5],16),int(c2[5:7],16)
                r = int(r1 + (r2-r1)*t)
                g = int(g1 + (g2-g1)*t)
                b = int(b1 + (b2-b1)*t)
                return f"#{r:02X}{g:02X}{b:02X}"

            n = len(colors) - 1
            for i in range(segs):
                t_global = i / segs
                idx = int(t_global * n)
                idx = min(idx, n - 1)
                t_local = (t_global * n) - idx
                color = hex_lerp(colors[idx], colors[idx+1], t_local)
                x1 = i * seg_w - pos
                x2 = x1 + seg_w + 1
                self.topbar.create_rectangle(x1, 0, x2, 6,
                                              fill=color, outline="")
                self.topbar.create_rectangle(x1+w, 0, x2+w, 6,
                                              fill=color, outline="")
            self._tb_pos = (self._tb_pos + 2) % w
        except Exception:
            pass
        self.root.after(35, self._topbar_anim)

    # ──────────────────────────────────────────
    # TOGGLE TEMA — transisi smooth
    # ──────────────────────────────────────────
    def _toggle_theme(self):
        self.is_dark = not self.is_dark
        self.T = DARK if self.is_dark else LIGHT
        # Apply bertahap untuk kesan transisi
        self._apply_theme()
        self.root.after(20, self._apply_theme)
        self.root.after(60, self._apply_theme)


    def _load_master(self, path=None):
        if path is None:
            path = self.master_path
        if not os.path.exists(path):
            self.lbl_info.config(
                text=f"⚠  {os.path.basename(path)} tidak ditemukan — klik Ganti File",
                fg=self.T["amber"])
            return

        # ── Skeleton loading animasi ──
        self.lbl_master.config(text="⏳  Memuat file...")
        self.lbl_info.config(text="— — — — —", fg=self.T["fg3"])
        self._set_badge("LOADING", "amber")
        self.root.update_idletasks()
        self._skeleton_anim = True
        self._skeleton_tick(0)

        # Load di thread agar UI tidak freeze
        threading.Thread(target=self._do_load_master,
                         args=(path,), daemon=True).start()

    def _skeleton_tick(self, step):
        """Animasi shimmer teks saat loading."""
        if not self._skeleton_anim:
            return
        dots = ["⣾","⣽","⣻","⢿","⡿","⣟","⣯","⣷"]
        self.lbl_master.config(text=f"{dots[step % len(dots)]}  Memuat file...")
        self.root.after(100, self._skeleton_tick, step + 1)

    def _do_load_master(self, path):
        """Load master file di background thread."""
        try:
            wb = openpyxl.load_workbook(path, data_only=True)

            ws_mat = wb["MATERIAL"] if "MATERIAL" in wb.sheetnames else wb.active
            mats = [str(r[0]).strip()
                    for r in ws_mat.iter_rows(min_row=2, values_only=True)
                    if r[0]]

            ws_plt = wb["PLANT"] if "PLANT" in wb.sheetnames else wb.active
            plts = [str(r[0]).strip().upper()
                    for r in ws_plt.iter_rows(min_row=2, values_only=True)
                    if r[0]]

            self.root.after(0, self._finish_load_master, path, mats, plts)
        except Exception as e:
            self.root.after(0, self._fail_load_master, str(e))

    def _finish_load_master(self, path, mats, plts):
        """Update UI setelah load selesai."""
        self._skeleton_anim = False
        self._build_chk(self.mat_frame, self.mat_vars, mats,
                         cols=4, on_change=self._upd_mat)
        self._build_chk(self.plt_frame, self.plant_vars, plts,
                         cols=5, on_change=self._upd_plt)
        self.lbl_master.config(text=f"📄  {os.path.basename(path)}")
        self.lbl_info.config(
            text=f"✓  {len(mats)} material  ·  {len(plts)} plant",
            fg=self.T["green"])
        self._upd_mat()
        self._upd_plt()
        self._apply_chk_theme()
        self._set_badge("READY", "green")
        self._log(f"[INFO] Master dimuat — {len(mats)} material, {len(plts)} plant")
        self._load_selection()

    def _fail_load_master(self, err):
        self._skeleton_anim = False
        self.lbl_master.config(text="❌  Gagal memuat")
        self.lbl_info.config(text=f"Error: {err}", fg=self.T["red"])
        self._set_badge("ERROR", "red")

    def _ganti_master(self):
        path = filedialog.askopenfilename(
            filetypes=[("Excel", "*.xlsx *.xls"), ("All", "*.*")])
        if path:
            self.master_path = path
            self._load_master(path)

    def _build_chk(self, frame, var_list, items, cols, on_change):
        for w in frame.winfo_children():
            w.destroy()
        var_list.clear()
        T = self.T
        for i, item in enumerate(items):
            var = tk.BooleanVar(value=False)
            chk = tk.Checkbutton(frame, text=item,
                                  variable=var,
                                  font=("Segoe UI", 11),
                                  relief="flat", cursor="hand2",
                                  command=on_change)
            chk.grid(row=i//cols, column=i%cols,
                     sticky="w", padx=6, pady=2)
            var_list.append({"val": item, "var": var, "chk": chk})

    def _reflow_chk(self, canvas, var_list, on_change, chk_width=90):
        """Rebuild grid kolom sesuai lebar canvas saat resize."""
        w = canvas.winfo_width()
        if w < 2:
            return
        cols = max(1, w // chk_width)
        for i, x in enumerate(var_list):
            x["chk"].grid_forget()
            x["chk"].grid(row=i//cols, column=i%cols,
                          sticky="w", padx=6, pady=2)

    def _filter_chk(self, var_list, keyword):
        """Tampilkan hanya checkbox yang cocok dengan keyword."""
        kw = keyword.strip().lower()
        for x in var_list:
            orig = x.get("orig", x["val"])
            if kw == "" or kw in orig.lower():
                x["chk"].grid()
            else:
                x["chk"].grid_remove()
        # Trigger reflow
        if var_list is self.mat_vars:
            self._reflow_chk(self.mat_canvas, self.mat_vars, self._upd_mat)
        else:
            self._reflow_chk(self.plt_canvas, self.plant_vars, self._upd_plt)

    # ──────────────────────────────────────────
    # AUTO-SAVE PILIHAN
    # ──────────────────────────────────────────
    def _save_selection(self):
        """Simpan pilihan material & plant ke file .cfg lokal."""
        try:
            sel_mats = self._get_sel(self.mat_vars)
            sel_plts = self._get_sel(self.plant_vars)
            cfg_path = os.path.join(os.getcwd(), ".rpa_selection.cfg")
            with open(cfg_path, "w", encoding="utf-8") as f:
                f.write("MATERIAL:" + ",".join(sel_mats) + "\n")
                f.write("PLANT:" + ",".join(sel_plts) + "\n")
        except Exception:
            pass

    def _load_selection(self):
        """Restore pilihan material & plant dari file .cfg."""
        try:
            cfg_path = os.path.join(os.getcwd(), ".rpa_selection.cfg")
            if not os.path.exists(cfg_path):
                return
            saved_mats, saved_plts = set(), set()
            with open(cfg_path, "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("MATERIAL:"):
                        saved_mats = set(line.strip().replace("MATERIAL:", "").split(","))
                    elif line.startswith("PLANT:"):
                        saved_plts = set(line.strip().replace("PLANT:", "").split(","))
            for x in self.mat_vars:
                if x["val"] in saved_mats:
                    x["var"].set(True)
            for x in self.plant_vars:
                if x["val"] in saved_plts:
                    x["var"].set(True)
            self._upd_mat()
            self._upd_plt()
            if saved_mats or saved_plts:
                self._log(f"[INFO] Pilihan terakhir di-restore: "
                          f"{len(saved_mats)} material, {len(saved_plts)} plant")
        except Exception:
            pass
    def _sel_all(self, var_list, upd):
        for x in var_list:
            x["var"].set(True)
        upd()

    def _sel_none(self, var_list, upd):
        for x in var_list:
            x["var"].set(False)
        upd()

    def _upd_mat(self):
        n = sum(1 for x in self.mat_vars if x["var"].get())
        self.lbl_mat_sel.config(text=f"{n} dipilih")

    def _upd_plt(self):
        n = sum(1 for x in self.plant_vars if x["var"].get())
        self.lbl_plt_sel.config(text=f"{n} dipilih")

    def _get_sel(self, var_list):
        return [x["val"] for x in var_list if x["var"].get()]

    # ──────────────────────────────────────────
    # THEME
    # ──────────────────────────────────────────
    def _apply_theme(self):
        T = self.T
        # Font constants — Segoe UI tajam di Windows, Consolas untuk monospace
        F_HDR    = ("Segoe UI", 20, "bold")
        F_BADGE  = ("Segoe UI", 12, "bold")
        F_SUB    = ("Segoe UI", 10)
        F_VER    = ("Segoe UI", 10)
        F_THEME  = ("Segoe UI", 11, "bold")
        F_SEC    = ("Segoe UI", 11, "bold")
        F_LABEL  = ("Segoe UI", 11)
        F_LABEL_B= ("Segoe UI", 11, "bold")
        F_BTN    = ("Segoe UI", 11, "bold")
        F_BTN_SM = ("Segoe UI", 10, "bold")
        F_BTN_XS = ("Segoe UI", 10)
        F_CHK    = ("Segoe UI", 11)
        F_WARN   = ("Segoe UI", 10)
        F_LOG    = ("Consolas", 10)
        F_BTN_MAIN = ("Segoe UI", 12, "bold")

        self.root.configure(bg=T["bg"])
        self.hdr.configure(bg=T["hdr_bg"])
        if self._hdr_left:
            self._hdr_left.configure(bg=T["hdr_bg"])
            if hasattr(self, "_hdr_title_col"):
                self._hdr_title_col.configure(bg=T["hdr_bg"])

        self.div.configure(bg=T["accent"])
        self.body.configure(bg=T["bg"])

        self.lbl_badge.configure(bg=T["accent"], fg="#FFFFFF", font=F_BADGE)
        # Title: di gelap pakai gold, di terang pakai fg gelap
        title_fg = T["gold2"] if self.is_dark else T["fg"]
        self.lbl_title.configure(bg=T["hdr_bg"], fg=title_fg, font=F_HDR)
        self.lbl_sub.configure(bg=T["hdr_bg"], fg=T["accent2"], font=F_SUB)
        self.lbl_ver.configure(bg=T["hdr_bg"], fg=T["fg3"], font=F_VER)
        self.btn_theme.configure(
            text=T["name"],
            bg=T["bg3"], fg=T["accent"],
            activebackground=T["bg4"], activeforeground=T["accent"],
            font=F_THEME)
        self.btn_theme._leave_bg = T["bg3"]

        for item in getattr(self, "_secs", []):
            f, lbl, bar = item
            f.configure(bg=T["bg"])
            lbl.configure(bg=T["bg"], fg=T["accent"], font=F_SEC)
            bar.configure(bg=T["accent"])

        for s in getattr(self, "_splits", []):
            s.configure(bg=T["bg"])

        # Search boxes
        for attr in ["ent_mat_search", "ent_plt_search"]:
            if hasattr(self, attr):
                getattr(self, attr).configure(
                    bg=T["bg3"], fg=T["fg"],
                    insertbackground=T["accent"],
                    highlightbackground=T["border"],
                    highlightcolor=T["accent"],
                    highlightthickness=1)
        for sf in getattr(self, "_search_frames", []):
            sf.configure(bg=T["bg2"])
            for w in sf.winfo_children():
                if w.winfo_class() == "Label":
                    w.configure(bg=T["bg2"], fg=T["fg3"])

        for f in getattr(self, "_cards", []):
            f.configure(bg=T["bg2"],
                        highlightbackground=T["accent2"],
                        highlightthickness=1)
            self._rec(f, T["bg2"])

        # Master
        self.lbl_master.configure(bg=T["bg2"], fg=T["fg"], font=F_LABEL_B)
        self.lbl_info.configure(bg=T["bg2"], fg=T["green"], font=F_LABEL_B)
        self.btn_ganti.configure(bg=T["bg3"], fg=T["accent"],
                                  activebackground=T["bg4"], font=F_BTN_SM)
        self.btn_ganti._leave_bg = T["bg3"]

        # Material toolbar
        btn_gold = T["gold2"] if self.is_dark else T["accent"]
        self.btn_mat_all.configure(bg=T["bg3"], fg=btn_gold,
                                    activebackground=T["bg4"], font=F_BTN_SM)
        self.btn_mat_all._leave_bg = T["bg3"]
        self.btn_mat_none.configure(bg=T["bg3"], fg=T["fg2"],
                                     activebackground=T["bg4"], font=F_BTN_XS)
        self.btn_mat_none._leave_bg = T["bg3"]
        self.lbl_mat_sel.configure(bg=T["bg2"], fg=T["accent"], font=F_LABEL_B)

        # Plant toolbar
        self.btn_plt_all.configure(bg=T["bg3"], fg=btn_gold,
                                    activebackground=T["bg4"], font=F_BTN_SM)
        self.btn_plt_all._leave_bg = T["bg3"]
        self.btn_plt_none.configure(bg=T["bg3"], fg=T["fg2"],
                                     activebackground=T["bg4"], font=F_BTN_XS)
        self.btn_plt_none._leave_bg = T["bg3"]
        self.lbl_plt_sel.configure(bg=T["bg2"], fg=T["accent"], font=F_LABEL_B)

        # Eksekusi
        self.chk_dry.configure(bg=T["bg2"], fg=T["fg"],
                                selectcolor=T["bg2"],
                                activebackground=T["bg3"], font=F_CHK)
        self.lbl_sloc.configure(bg=T["bg2"], fg=T["accent2"], font=F_WARN)
        self.lbl_status.configure(bg=T["bg2"], fg=T["green"], font=F_LABEL_B)
        self.lbl_warn.configure(bg=T["bg2"], fg=T["amber"], font=F_WARN)
        self.pb_frame.configure(bg=T["bg3"])
        self.pb_fill.configure(bg=T["accent"])

        self.btn_start.configure(bg=T["btn_pri"], fg="#FFFFFF",
                                  activebackground=T["accent2"],
                                  activeforeground="#FFFFFF",
                                  font=F_BTN_MAIN)
        self.btn_start._leave_bg = T["btn_pri"]

        self.btn_stop.configure(bg=T["btn_stop"], fg="#FFFFFF",
                                 activebackground=T["red"],
                                 activeforeground="#FFFFFF",
                                 font=F_BTN_MAIN)
        self.btn_stop._leave_bg = T["btn_stop"]

        self.btn_out.configure(bg=T["bg3"], fg=T["fg2"],
                                activebackground=T["bg4"], font=F_BTN_XS)
        self.btn_out._leave_bg = T["bg3"]

        # Log — selalu dark agar terminal terbaca
        self.log.configure(bg=T["log_bg"], fg=T["log_fg"],
                            insertbackground=T["accent"],
                            font=F_LOG,
                            selectbackground=T["accent2"],
                            selectforeground=T["log_bg"])
        self.btn_clr.configure(bg=T["log_bg"], fg=T["fg3"],
                                activebackground="#2A2A2A" if not self.is_dark else T["bg3"],
                                font=("Segoe UI", 10))
        self.btn_clr._leave_bg = T["log_bg"]

        if hasattr(self, "btn_export_log"):
            self.btn_export_log.configure(
                bg=T["log_bg"], fg=T["accent"],
                activebackground="#2A2A2A" if not self.is_dark else T["bg3"],
                font=("Segoe UI", 10))
            self.btn_export_log._leave_bg = T["log_bg"]

        self._apply_chk_theme()

        # Scrollbar — pakai sb_bg/sb_fg per tema
        sty = ttk.Style()
        sty.theme_use("clam")
        sty.configure("Vertical.TScrollbar",
                       background=T["sb_fg"],
                       troughcolor=T["sb_bg"],
                       bordercolor=T["sb_bg"],
                       arrowcolor=T["sb_bg"],
                       relief="flat",
                       width=7)
        sty.map("Vertical.TScrollbar",
                background=[("active", T["accent"]),
                             ("!active", T["sb_fg"])])

        # Status bar
        if hasattr(self, "_hdr_right"):
            self._hdr_right.configure(bg=T["hdr_bg"])
        if hasattr(self, "lbl_status_badge"):
            self._set_badge("READY", "green")
        if hasattr(self, "lbl_detail"):
            self.lbl_detail.configure(bg=T["bg2"], fg=T["amber"])
        if hasattr(self, "lbl_detail_right"):
            self.lbl_detail_right.configure(bg=T["bg2"], fg=T["fg3"])
        if hasattr(self, "_detail_var"):
            pass  # StringVar tidak perlu theme
            self.statusbar.configure(bg=T["bg3"])
            self.lbl_counter.configure(bg=T["bg3"], fg=T["accent"])
            self.lbl_elapsed.configure(bg=T["bg3"], fg=T["fg3"])

    def _apply_chk_theme(self):
        T = self.T
        for canvas, frame in [(self.mat_canvas, self.mat_frame),
                               (self.plt_canvas, self.plt_frame)]:
            canvas.configure(bg=T["bg2"])
            frame.configure(bg=T["bg2"])
        for var_list in [self.mat_vars, self.plant_vars]:
            for x in var_list:
                x["chk"].configure(
                    bg=T["bg2"], fg=T["fg"],
                    selectcolor=T["bg2"] if self.is_dark else T["chk_sel"],
                    activebackground=T["bg3"],
                    activeforeground=T["accent"],
                    font=("Segoe UI", 11),
                    cursor="hand2")

    def _rec(self, w, bg):
        for c in w.winfo_children():
            cls = c.winfo_class()
            if cls == "Frame":
                c.configure(bg=bg)
                self._rec(c, bg)

    # ──────────────────────────────────────────
    # LOG
    # ──────────────────────────────────────────
    def _log(self, msg, color=None):
        # Strip ANSI escape codes
        msg = re.sub(r'\x1b\[[0-9;]*[mGKHF]', '', msg)
        msg = re.sub(r'\x1b\[[0-9]*m', '', msg)
        T = self.T

        # Warna otomatis berdasarkan konten jika tidak disediakan
        if color is None:
            if any(k in msg for k in ["SUCCESS", "SELESAI", "✓", "OK"]):
                color = T["green"]
            elif any(k in msg for k in ["ERROR", "❌", "FAIL", "Exception"]):
                color = T["red"]
            elif any(k in msg for k in ["WARN", "WARNING"]):
                color = T["amber"]
            elif any(k in msg for k in ["INFO", "START", "DEBUG"]):
                color = T["accent"]
            elif msg.startswith("["): 
                color = T["fg2"]
            else:
                color = T["log_fg"]

        self.log.configure(state="normal")
        tag = f"t{id(msg)}{len(msg)}"
        self.log.insert("end", msg + "\n", tag)
        if color:
            self.log.tag_configure(tag, foreground=color)
        self.log.configure(state="disabled")
        self.log.see("end")

    def _clear_log(self):
        self.log.configure(state="normal")
        self.log.delete("1.0", "end")
        self.log.configure(state="disabled")

    def _set_badge(self, text, color_key):
        """Update status badge di header."""
        T = self.T
        colors = {
            "green":  T["green"],
            "amber":  T["amber"],
            "red":    T["red"],
            "accent": T["accent"],
            "cyan":   T["accent"],
        }
        fg = colors.get(color_key, T["fg2"])
        icons = {"READY":"⬡", "RUNNING":"⚙", "LOADING":"⣾",
                 "DONE":"✅", "ERROR":"❌", "STOPPED":"■"}
        icon = icons.get(text, "⬡")
        if hasattr(self, "lbl_status_badge"):
            self.lbl_status_badge.configure(
                text=f"{icon}  {text}",
                fg=fg, bg=T["hdr_bg"],
                font=("Segoe UI", 9, "bold"))

        # Pulse animasi saat RUNNING
        if text == "RUNNING":
            self._badge_pulse = True
            self._pulse_badge()
        else:
            self._badge_pulse = False

    def _update_detail_counter(self, mat, plt_idx, plt_total, plt_name):
        """Update counter detail real-time di eksekusi."""
        T = self.T
        mat_idx = 0
        mat_total = 0
        sel_mats = [x["val"] for x in self.mat_vars if x["var"].get()]
        mat_total = len(sel_mats)
        if mat in sel_mats:
            mat_idx = sel_mats.index(mat) + 1
        self._detail_var.set(
            f"Material {mat_idx}/{mat_total} : {mat}  →  Plant {plt_idx}/{plt_total} : {plt_name}")
        if hasattr(self, "lbl_detail"):
            self.lbl_detail.configure(fg=T["amber"])

    def _clear_detail_counter(self):
        self._detail_var.set("")
        if hasattr(self, "lbl_detail_right"):
            self.lbl_detail_right.configure(text="")
        """Update timer elapsed di status bar selama proses berjalan."""
        if not self._start_time:
            return
        if self.btn_start.cget("state") == "normal":
            return  # proses sudah selesai
        import time
        elapsed = int(time.time() - self._start_time)
        m, s = divmod(elapsed, 60)
        if m > 0:
            self.lbl_elapsed.config(text=f"⏱  {m:02d}:{s:02d}")
        else:
            self.lbl_elapsed.config(text=f"⏱  {elapsed}d")
        self.root.after(1000, self._tick_elapsed)

    def _confirm_start(self, mats, plts):
        """Popup konfirmasi dengan preview + countdown + cek SAP aktif."""
        T = self.T
        total = len(mats) * len(plts)
        mode_txt = " [DRY RUN]" if self.dry_var.get() else ""

        # ── Cek apakah SAP window aktif ──────────────
        sap_found = self._check_sap_window()

        win = tk.Toplevel(self.root)
        win.title("Konfirmasi")
        win.resizable(False, False)
        win.configure(bg=T["bg2"])
        win.grab_set()

        pw, ph = 440, 440
        rx = self.root.winfo_x() + (self.root.winfo_width() - pw) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - ph) // 2
        win.geometry(f"{pw}x{ph}+{rx}+{ry}")

        # Header
        hdr_color = T["accent"] if sap_found else T["amber"]
        tk.Frame(win, bg=hdr_color, height=4).pack(fill="x")
        tk.Label(win, text=f"⬡  Konfirmasi Proses{mode_txt}",
                 font=("Segoe UI", 13, "bold"),
                 bg=T["bg2"], fg=hdr_color, pady=10).pack(fill="x")

        # SAP status warning
        sap_frame = tk.Frame(win, bg=T["bg3"], padx=16, pady=6)
        sap_frame.pack(fill="x", padx=20, pady=(0,6))
        if sap_found:
            sap_icon, sap_txt, sap_color = "✅", f"SAP terdeteksi: {sap_found}", T["green"]
        else:
            sap_icon, sap_txt, sap_color = "⚠", "SAP tidak terdeteksi — pastikan SAP sudah terbuka!", T["amber"]
        tk.Label(sap_frame, text=f"{sap_icon}  {sap_txt}",
                 font=("Segoe UI", 9, "bold"),
                 bg=T["bg3"], fg=sap_color, anchor="w").pack(fill="x")

        tk.Frame(win, bg=T["border"], height=1).pack(fill="x", padx=20)

        # Summary info
        sf = tk.Frame(win, bg=T["bg2"], padx=24, pady=6)
        sf.pack(fill="x")

        def info_row(label, value, color=None):
            r = tk.Frame(sf, bg=T["bg2"])
            r.pack(fill="x", pady=2)
            tk.Label(r, text=label, font=("Segoe UI", 10),
                     bg=T["bg2"], fg=T["fg2"], width=16, anchor="w").pack(side="left")
            tk.Label(r, text=value, font=("Segoe UI", 10, "bold"),
                     bg=T["bg2"], fg=color or T["fg"]).pack(side="left")

        info_row("Material dipilih", f"{len(mats)} item", T["accent"])
        info_row("Plant dipilih",    f"{len(plts)} item", T["accent"])
        info_row("Total kombinasi",  f"{total} proses", T["gold"])
        if self.dry_var.get():
            info_row("Mode", "DRY RUN — tidak ada perubahan", T["amber"])

        tk.Frame(win, bg=T["border"], height=1).pack(fill="x", padx=20, pady=(4,0))

        # Preview kombinasi
        prev_frame = tk.Frame(win, bg=T["bg3"], padx=16, pady=6)
        prev_frame.pack(fill="x", padx=20, pady=6)
        tk.Label(prev_frame, text="Preview kombinasi:",
                 font=("Segoe UI", 9, "bold"),
                 bg=T["bg3"], fg=T["fg2"]).pack(anchor="w")

        preview_txt = ""
        count = 0
        for m in mats:
            for p in plts:
                if count < 5:
                    preview_txt += f"  {m}  ×  {p}\n"
                count += 1
        if count > 5:
            preview_txt += f"  ... dan {count-5} kombinasi lainnya"

        tk.Label(prev_frame, text=preview_txt.rstrip(),
                 font=("Consolas", 9), bg=T["bg3"], fg=T["fg"],
                 justify="left").pack(anchor="w", pady=(2,0))

        tk.Frame(win, bg=T["border"], height=1).pack(fill="x", padx=20, pady=(4,0))

        # Tombol + countdown
        confirmed = tk.BooleanVar(value=False)
        countdown_val = [5]   # countdown detik

        btn_f = tk.Frame(win, bg=T["bg2"], pady=10)
        btn_f.pack()

        btn_ok = tk.Button(btn_f, text=f"▶  Mulai ({countdown_val[0]}s)",
                           font=("Segoe UI", 10, "bold"),
                           bg=T["btn_pri"], fg="white",
                           relief="flat", cursor="hand2",
                           padx=18, pady=7)
        btn_ok.pack(side="left", padx=6)
        self._bind_hover(btn_ok)

        btn_cancel = tk.Button(btn_f, text="  Batal  ",
                               font=("Segoe UI", 10),
                               bg=T["bg3"], fg=T["fg2"],
                               relief="flat", cursor="hand2",
                               padx=18, pady=7)
        btn_cancel.pack(side="left", padx=6)
        self._bind_hover(btn_cancel)

        # Warning jangan sentuh
        tk.Label(win, text="⚠  Jangan sentuh mouse/keyboard setelah klik Mulai",
                 font=("Segoe UI", 8), bg=T["bg2"], fg=T["amber"]).pack(pady=(0,6))

        def do_start():
            confirmed.set(True)
            win.destroy()

        def do_cancel():
            win.destroy()

        def tick():
            if not win.winfo_exists():
                return
            countdown_val[0] -= 1
            if countdown_val[0] <= 0:
                # Auto mulai setelah countdown habis
                do_start()
                return
            btn_ok.configure(text=f"▶  Mulai ({countdown_val[0]}s)")
            win.after(1000, tick)

        btn_ok.configure(command=do_start)
        btn_cancel.configure(command=do_cancel)

        # Mulai countdown setelah 500ms
        win.after(500, tick)
        win.wait_window()
        return confirmed.get()

    def _check_sap_window(self):
        """Cek apakah window SAP sedang aktif di Windows."""
        try:
            import subprocess
            result = subprocess.run(
                ["powershell", "-Command",
                 "Get-Process | Where-Object {$_.MainWindowTitle -like '*SAP*'} "
                 "| Select-Object -First 1 MainWindowTitle | ForEach-Object {$_.MainWindowTitle}"],
                capture_output=True, text=True, timeout=3)
            title = result.stdout.strip()
            if title:
                # Ambil max 40 karakter
                return title[:40] + ("..." if len(title) > 40 else "")
            return None
        except Exception:
            return None

    def _export_log(self):
        """Export isi log ke file .txt."""
        try:
            content = self.log.get("1.0", "end").strip()
            if not content:
                return
            import datetime
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            out_dir = os.path.join(os.getcwd(), "output_rpa")
            os.makedirs(out_dir, exist_ok=True)
            path = os.path.join(out_dir, f"log_{ts}.txt")
            with open(path, "w", encoding="utf-8") as f:
                f.write(content)
            self._log(f"[INFO] Log disimpan: output_rpa\\log_{ts}.txt",
                      self.T["green"])
        except Exception as e:
            self._log(f"[ERROR] Gagal export log: {e}", self.T["red"])
        """Update timer elapsed di status bar selama proses berjalan."""
        if not self._start_time:
            return
        if self.btn_start.cget("state") == "normal":
            return
        import time
        elapsed = int(time.time() - self._start_time)
        m, s = divmod(elapsed, 60)
        if m > 0:
            self.lbl_elapsed.config(text=f"⏱  {m:02d}:{s:02d}")
        else:
            self.lbl_elapsed.config(text=f"⏱  {elapsed}d")
        self.root.after(1000, self._tick_elapsed)

    def _tick_elapsed(self):
        """Update timer elapsed di status bar selama proses berjalan."""
        if not self._start_time:
            return
        if self.btn_start.cget("state") == "normal":
            return
        import time
        elapsed = int(time.time() - self._start_time)
        m, s = divmod(elapsed, 60)
        if m > 0:
            self.lbl_elapsed.config(text=f"⏱  {m:02d}:{s:02d}")
        else:
            self.lbl_elapsed.config(text=f"⏱  {elapsed}d")
        self.root.after(1000, self._tick_elapsed)

    def _update_counter(self):
        """Update live counter di status bar."""
        done = self._ok_count + self._err_count
        total = max(1, self._total_count)
        self._counter_var.set(
            f"▶  {done} / {total}  ·  ✅ {self._ok_count}  ·  ❌ {self._err_count}")

    def _play_sound(self, success=True):
        """Notifikasi suara saat selesai."""
        try:
            import winsound
            if success:
                # 3x ding pendek — sukses
                for freq, dur in [(800, 150), (1000, 150), (1200, 200)]:
                    winsound.Beep(freq, dur)
            else:
                # 2x nada rendah — ada error
                for freq, dur in [(400, 300), (300, 400)]:
                    winsound.Beep(freq, dur)
        except Exception:
            pass

    def _show_summary(self, rc):
        """Popup summary dengan mini bar chart."""
        T = self.T
        elapsed = getattr(self, "_final_elapsed", 0)
        m, s = divmod(elapsed, 60)
        if m > 0:
            durasi_str = f"{m:02d}:{s:02d} menit"
        else:
            durasi_str = f"{elapsed} detik"

        ok    = self._ok_count
        err   = self._err_count
        total = self._total_count
        done  = ok + err

        win = tk.Toplevel(self.root)
        win.title("Ringkasan Proses")
        win.resizable(False, False)
        win.configure(bg=T["bg2"])
        win.grab_set()

        win.update_idletasks()
        pw, ph = 380, 340
        rx = self.root.winfo_x() + (self.root.winfo_width() - pw) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - ph) // 2
        win.geometry(f"{pw}x{ph}+{rx}+{ry}")

        # Header
        status_color = T["green"] if err == 0 else T["amber"] if ok > 0 else T["red"]
        status_icon  = "✅" if err == 0 else "⚠" if ok > 0 else "❌"
        status_text  = "SELESAI" if err == 0 else "SELESAI (ada error)"
        tk.Frame(win, bg=status_color, height=4).pack(fill="x")
        tk.Label(win, text=f"{status_icon}  {status_text}",
                 font=("Segoe UI", 14, "bold"),
                 bg=T["bg2"], fg=status_color,
                 pady=14).pack(fill="x")

        # ── Mini bar chart ────────────────────
        chart_frame = tk.Frame(win, bg=T["bg2"], padx=24)
        chart_frame.pack(fill="x", pady=(0,8))

        chart_h = 36
        chart = tk.Canvas(chart_frame, height=chart_h,
                           bg=T["bg3"], highlightthickness=0)
        chart.pack(fill="x")

        def draw_chart():
            chart.delete("all")
            w = chart.winfo_width()
            if w < 10 or total == 0:
                return
            # Bar OK
            ok_w = int(w * ok / total)
            if ok_w > 0:
                chart.create_rectangle(0, 4, ok_w, chart_h-4,
                                        fill=T["green"], outline="")
                if ok_w > 30:
                    chart.create_text(ok_w//2, chart_h//2,
                                       text=f"✓ {ok}", fill=T["bg"],
                                       font=("Segoe UI", 9, "bold"))
            # Bar ERROR
            if err > 0:
                err_w = int(w * err / total)
                x1 = ok_w
                chart.create_rectangle(x1, 4, x1+err_w, chart_h-4,
                                        fill=T["red"], outline="")
                if err_w > 30:
                    chart.create_text(x1 + err_w//2, chart_h//2,
                                       text=f"✗ {err}", fill="white",
                                       font=("Segoe UI", 9, "bold"))

        chart.bind("<Configure>", lambda e: draw_chart())
        win.after(50, draw_chart)

        # Stats
        tk.Frame(win, height=1, bg=T["border"]).pack(fill="x", padx=20, pady=(4,0))
        stats = tk.Frame(win, bg=T["bg2"], pady=10)
        stats.pack(fill="x", padx=30)

        def stat_row(label, value, color):
            row = tk.Frame(stats, bg=T["bg2"])
            row.pack(fill="x", pady=3)
            tk.Label(row, text=label, font=("Segoe UI", 10),
                     bg=T["bg2"], fg=T["fg2"], width=16, anchor="w"
                     ).pack(side="left")
            tk.Label(row, text=value, font=("Segoe UI", 11, "bold"),
                     bg=T["bg2"], fg=color).pack(side="left")

        stat_row("Total Proses",  f"{done} / {total}", T["fg"])
        stat_row("✅  Berhasil",  f"{ok}", T["green"])
        stat_row("❌  Error",     f"{err}", T["red"] if err > 0 else T["fg3"])
        stat_row("⏱  Durasi",    durasi_str, T["accent"])

        tk.Frame(win, height=1, bg=T["border"]).pack(fill="x", padx=20, pady=(4,0))

        btn_frame = tk.Frame(win, bg=T["bg2"], pady=12)
        btn_frame.pack()

        def open_output():
            win.destroy()
            self._open_output()

        b1 = tk.Button(btn_frame, text="📂  Buka Output",
                       font=("Segoe UI", 10, "bold"),
                       bg=T["btn_pri"], fg="white",
                       relief="flat", cursor="hand2",
                       padx=16, pady=6, command=open_output)
        b1.pack(side="left", padx=6)
        self._bind_hover(b1)

        b2 = tk.Button(btn_frame, text="  Tutup  ",
                       font=("Segoe UI", 10),
                       bg=T["bg3"], fg=T["fg2"],
                       relief="flat", cursor="hand2",
                       padx=16, pady=6, command=win.destroy)
        b2.pack(side="left", padx=6)
        self._bind_hover(b2)

    # ──────────────────────────────────────────
    # STATUS CHECKBOX — tanda done/error langsung di list
    # ──────────────────────────────────────────
    def _reset_chk_status(self):
        """Reset semua checkbox ke tampilan normal sebelum proses baru."""
        T = self.T
        for var_list in [self.mat_vars, self.plant_vars]:
            for x in var_list:
                orig = re.sub(r'^[✅❌⚙]\s*', '', x["chk"].cget("text"))
                x["orig"] = orig
                x["chk"].config(text=orig, fg=T["fg"],
                                 font=("Segoe UI", 11))
        # tracking: {material: {"done": 0, "error": 0, "total": N}}
        self._mat_progress = {}
        self._current_mat  = None
        self._current_plt  = None

    def _plant_done(self, mat, result):
        """Dipanggil tiap kali 1 plant selesai untuk suatu material."""
        if mat not in self._mat_progress:
            return

        prog = self._mat_progress[mat]
        finished_before = prog["done"] + prog["error"]

        if result == "ok":
            prog["done"] += 1
            self._ok_count += 1
        elif result == "skip":
            prog["error"] += 1   # skip dihitung sebagai tidak berhasil
            self._err_count += 1
        else:
            prog["error"] += 1
            self._err_count += 1

        finished_after = prog["done"] + prog["error"]

        # Guard: jangan proses lebih dari total plant
        if finished_after > prog["total"]:
            # Rollback — sudah dihitung sebelumnya
            if result == "ok":
                prog["done"] -= 1
                self._ok_count -= 1
            else:
                prog["error"] -= 1
                self._err_count -= 1
            return

        # Update live counter status bar
        self._update_counter()

        # Jika semua plant sudah selesai untuk material ini → update material
        if finished_after >= prog["total"]:
            if prog["error"] == 0:
                self._set_chk_status(mat, "ok", self.mat_vars)
            else:
                self._set_chk_status(mat, "error", self.mat_vars)

    def _reset_plant_status(self):
        """Reset semua plant ke tampilan normal saat material baru mulai."""
        T = self.T
        for x in self.plant_vars:
            orig = x.get("orig", x["val"])
            x["chk"].config(text=orig, fg=T["fg"],
                             font=("Segoe UI", 11))

    def _set_chk_status(self, val, status, var_list):
        """Update tampilan 1 checkbox berdasarkan nilai & status."""
        for x in var_list:
            orig = x.get("orig", x["val"])
            if orig == val or x["val"] == val:
                if status == "running":
                    x["chk"].config(text=f"⚙  {orig}", fg="#FFC030",
                                     font=("Segoe UI", 11, "bold"))
                elif status == "ok":
                    x["chk"].config(text=f"✅  {orig}", fg="#00F080",
                                     font=("Segoe UI", 11, "bold"))
                elif status == "skip":
                    x["chk"].config(text=f"⏭  {orig}", fg="#AAAAAA",
                                     font=("Segoe UI", 11, "bold"))
                elif status == "error":
                    x["chk"].config(text=f"❌  {orig}", fg="#FF4455",
                                     font=("Segoe UI", 11, "bold"))
                break

    # ──────────────────────────────────────────
    # EKSEKUSI
    # ──────────────────────────────────────────
    def _start(self):
        T = self.T
        mats  = self._get_sel(self.mat_vars)
        plts  = self._get_sel(self.plant_vars)

        if not mats:
            messagebox.showwarning("Peringatan", "Pilih minimal 1 material!")
            return
        if not plts:
            messagebox.showwarning("Peringatan", "Pilih minimal 1 plant!")
            return
        if not os.path.exists(SCRIPT):
            messagebox.showerror("Error", f"{SCRIPT} tidak ditemukan!")
            return

        # Auto-save pilihan
        self._save_selection()

        # Popup konfirmasi dengan preview kombinasi
        if not self._confirm_start(mats, plts):
            return
        tmp = os.path.join(os.getcwd(), "_selected_temp.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["MATNR", "PLANT"])
        for m in mats:
            for p in plts:
                ws.append([m, p])
        wb.save(tmp)
        total = len(mats) * len(plts)

        # Update script — dry_run & input_file
        with open(SCRIPT, "r") as f:
            content = f.read()
        val = "True" if self.dry_var.get() else "False"
        content = re.sub(r'"dry_run"\s*:\s*(True|False)',
                          f'"dry_run": {val}', content)
        content = re.sub(r'"input_file"\s*:\s*"[^"]*"',
                          '"input_file": "_selected_temp.xlsx"', content)
        with open(SCRIPT, "w") as f:
            f.write(content)

        import time
        self._start_time  = time.time()
        self._ok_count    = 0
        self._err_count   = 0
        self._total_count = total
        self._counter_var.set(f"▶  0 / {total}  ·  ✅ 0  ·  ❌ 0")
        self._clear_detail_counter()
        self._set_badge("RUNNING", "cyan")
        self._clear_log()
        self._reset_chk_status()
        # Init tracker: tiap material punya counter plant
        total_plts = len(plts)
        for m in mats:
            self._mat_progress[m] = {"done": 0, "error": 0, "total": total_plts}
        self.btn_start.config(state="disabled")
        self.btn_stop.config(state="normal")
        self._pb_start()
        mode = " [DRY RUN]" if self.dry_var.get() else ""
        self.status_var.set(
            f"Memproses {total} kombinasi ({len(mats)} mat × {len(plts)} plant){mode}...")
        self.lbl_status.config(fg=T["amber"])
        self._log(f"[START] {len(mats)} material × {len(plts)} plant = {total} proses{mode}",
                  T["amber"])
        self._log(f"        Material : {', '.join(mats)}")
        self._log(f"        Plant    : {', '.join(plts)}")

        threading.Thread(target=self._run, daemon=True).start()
        self._tick_elapsed()

    def _run(self):
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            self.process = subprocess.Popen(
                [PYTHON, SCRIPT],
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True, bufsize=1,
                encoding="utf-8",
                errors="replace",
                env=env,
                cwd=os.getcwd())

            sel_mats = [x["val"] for x in self.mat_vars if x["var"].get()]
            sel_plts = [x["val"] for x in self.plant_vars if x["var"].get()]

            for line in self.process.stdout:
                line = line.rstrip()
                if not line:
                    continue
                c = (self.T["green"]  if ("SUCCESS" in line or "✓" in line)
                     else self.T["red"]    if ("ERROR" in line or "❌" in line)
                     else self.T["amber"]  if "WARN" in line
                     else self.T["accent"] if "INFO" in line
                     else self.T["fg"])
                self.root.after(0, self._log, line, c)

                found_mat = next((m for m in sel_mats if m in line), None)
                found_plt = next((p for p in sel_plts if p in line), None)

                # Update current tracker jika ditemukan
                if found_mat:
                    # Ganti material → reset plant
                    if self._current_mat and self._current_mat != found_mat:
                        self.root.after(0, self._reset_plant_status)
                        self._current_plt = None  # reset plant tracker
                    self._current_mat = found_mat

                if found_plt:
                    self._current_plt = found_plt

                # Gunakan current jika not found di baris ini
                eff_mat = found_mat or self._current_mat
                eff_plt = found_plt or self._current_plt

                # ── Material running ──
                if found_mat and ("Isi Material" in line or "Material:" in line):
                    self.root.after(0, self._set_chk_status,
                                    found_mat, "running", self.mat_vars)

                # ── Plant running + update detail counter ──
                if found_plt and ("Plant:" in line or "Isi Plant" in line
                                   or "Isi WH" in line or "SLoc" in line):
                    self.root.after(0, self._set_chk_status,
                                    found_plt, "running", self.plant_vars)
                    if eff_mat:
                        prog = self._mat_progress.get(eff_mat, {})
                        done_so_far = prog.get("done", 0) + prog.get("error", 0)
                        plt_idx = done_so_far + 1
                        plt_total = prog.get("total", len(sel_plts))
                        self.root.after(0, self._update_detail_counter,
                                        eff_mat, plt_idx, plt_total, found_plt)

                # ── Deteksi SKIP — perlakukan sebagai error/warning ──
                is_skip = "[SKIP]" in line or "SKIP" in line

                # ── Plant + Material berhasil (SUCCESS) — bukan SKIP ──
                if "SUCCESS" in line and not is_skip and eff_mat and eff_plt:
                    self.root.after(0, self._set_chk_status,
                                    eff_plt, "ok", self.plant_vars)
                    self.root.after(0, self._plant_done, eff_mat, "ok")

                # ── Plant + Material SKIP ──
                elif is_skip and eff_mat and eff_plt:
                    self.root.after(0, self._set_chk_status,
                                    eff_plt, "skip", self.plant_vars)
                    self.root.after(0, self._plant_done, eff_mat, "skip")

                # ── Plant + Material gagal (ERROR) ──
                elif ("ERROR" in line or "❌" in line) and "Total" not in line and not is_skip:
                    if eff_mat and eff_plt:
                        self.root.after(0, self._set_chk_status,
                                        eff_plt, "error", self.plant_vars)
                        self.root.after(0, self._plant_done, eff_mat, "error")

            self.process.wait()
            self.root.after(0, self._done, self.process.returncode)
        except Exception as e:
            self.root.after(0, self._error, str(e))

    def _done(self, rc):
        T = self.T
        self._pb_stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        tmp = os.path.join(os.getcwd(), "_selected_temp.xlsx")
        if os.path.exists(tmp):
            os.remove(tmp)

        if rc == 0:
            self.status_var.set("✅  Selesai! Laporan ada di folder output_rpa")
            self.lbl_status.config(fg=T["green"])
            self._log("[SELESAI] Laporan tersimpan di output_rpa\\", T["green"])
            self._counter_var.set(
                f"✅  Selesai  ·  ✅ {self._ok_count}  ·  ❌ {self._err_count}")
        else:
            self.status_var.set("❌  Terjadi error")
            self.lbl_status.config(fg=T["red"])
            self._log(f"[ERROR] Proses berakhir kode {rc}", T["red"])
            self._counter_var.set(
                f"❌  Error  ·  ✅ {self._ok_count}  ·  ❌ {self._err_count}")

        # Stop timer — simpan elapsed dulu
        import time
        self._final_elapsed = int(time.time() - self._start_time) if self._start_time else 0
        self._start_time = None
        self._clear_detail_counter()

        if rc == 0:
            self._set_badge("DONE", "green")
        else:
            self._set_badge("ERROR", "red")

        # Notifikasi suara
        threading.Thread(
            target=self._play_sound,
            args=(self._err_count == 0,),
            daemon=True).start()

        # Summary popup
        self.root.after(300, self._show_summary, rc)

    def _error(self, msg):
        T = self.T
        self._pb_stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self.status_var.set(f"❌  Error: {msg}")
        self.lbl_status.config(fg=T["red"])
        self._log(f"[ERROR] {msg}", T["red"])

    def _stop(self):
        T = self.T
        if self.process:
            self.process.terminate()
        self._pb_stop()
        self.btn_start.config(state="normal")
        self.btn_stop.config(state="disabled")
        self._set_badge("STOPPED", "amber")
        self._clear_detail_counter()
        self.status_var.set("■  Dihentikan manual")
        self.lbl_status.config(fg=T["amber"])
        self._log("[STOP] Robot dihentikan manual.", T["amber"])

    def _open_output(self):
        path = os.path.join(os.getcwd(), "output_rpa")
        os.makedirs(path, exist_ok=True)
        os.startfile(path)


if __name__ == "__main__":
    # ── DPI FIX untuk Windows — render tajam di 1080p ──
    try:
        from ctypes import windll
        windll.shcore.SetProcessDpiAwareness(1)  # System DPI aware
    except Exception:
        try:
            windll.user32.SetProcessDPIAware()
        except Exception:
            pass

    root = tk.Tk()

    # Scaling optimal untuk 1080p — nilai 1.33 setara ~96 DPI native Windows
    root.tk.call("tk", "scaling", 1.33)

    App(root)
    root.mainloop()
