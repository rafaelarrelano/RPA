"""
send_email_report.py
Buat draft email di Mozilla Thunderbird berisi laporan selisih stok per plant.
Alih-alih mengirim via SMTP, email disimpan sebagai file .eml yang bisa dibuka di Thunderbird.

Perubahan:
- Subject per plant: "Req. Adj. EOD Plant {code} {name} Tanggal {tanggal}"
- Body email per plant: format standar sesuai template (Posting Date, Material,
  Selisih2, UOM, Gudang, Plant, Adj, Plant Name)
- Buat SATU EMAIL DRAFT PER PLANT (bukan satu email untuk semua plant)
- TIDAK ada file Excel laporan dan TIDAK ada attachment — seluruh isi laporan
  ada di badan (body) email dalam bentuk tabel HTML
- Load Plant Name dari Excel plant_mapping (kolom B=Plant Code, C=Plant Name)
- Load UoM per material dari Excel Material_UoM (kolom A=Material, B=UoM) —
  file ini berisi daftar PENGECUALIAN material yang UoM-nya bukan CAR.
  Material yang tidak ditemukan di mapping tetap ditulis "CAR" (default)
- Draft email disimpan sebagai file .eml di folder report
- Bisa dibuka/diimpor ke Thunderbird untuk review sebelum mengirim
"""

import os
from datetime import datetime
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl

from config import Config
from logger import setup_logger
from email_config_ui import load_credentials

log = setup_logger()

# UoM default untuk material yang TIDAK ada di file mapping.
# File Material_UoM.xlsx hanya berisi daftar pengecualian — material yang
# UoM-nya BUKAN CAR. Material yang tidak ditemukan di mapping berarti
# memang UoM-nya CAR (default), jadi tetap ditulis "CAR".
DEFAULT_UOM = "CAR"


# ─────────────────────────────────────────────
# LOAD PLANT NAME MAPPING DARI EXCEL
# ─────────────────────────────────────────────

def load_plant_name_mapping(filepath: str = None) -> dict:
    """
    Baca mapping Plant Code → Plant Name dari Excel plant_mapping.
    Struktur Excel (sheet Plant_CostCenter):
      Kolom B = Plant Code, Kolom C = Plant Name
      Data mulai baris 5

    Return: { "4502": "PGD Surabaya", "4503": "PGD Jakarta GT", ... }
    """
    if filepath is None:
        filepath = Config.PLANT_MAPPING_FILE

    mapping = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb["Plant_CostCenter"]
        for row in ws.iter_rows(min_row=5, values_only=True):
            # row[1] = kolom B (Plant Code), row[2] = kolom C (Plant Name)
            if row[1] and row[2]:
                plant_code = str(row[1]).strip()
                plant_name = str(row[2]).strip()
                mapping[plant_code] = plant_name
        log.info(f"[MAPPING] {len(mapping)} plant name berhasil dibaca")
    except Exception as e:
        log.warning(f"[MAPPING] Gagal baca plant name mapping: {e} — akan pakai kode plant saja")

    return mapping


# ─────────────────────────────────────────────
# LOAD MATERIAL → UOM MAPPING DARI EXCEL
# ─────────────────────────────────────────────

def load_material_uom_mapping(filepath: str = None) -> dict:
    """
    Baca mapping Material → UoM dari Excel.
    Struktur Excel (sheet pertama / aktif):
      Kolom A = Material, Kolom B = UoM
      Header di baris 1, data mulai baris 2

    Material bisa berupa angka (contoh: 310049) atau teks — selalu
    dinormalisasi jadi string tanpa spasi supaya cocok dengan
    item.material (yang juga string).

    Return: { "310049": "PKT", "411028": "ZAK", ... }
    """
    if filepath is None:
        filepath = Config.MATERIAL_UOM_FILE

    mapping = {}
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
        ws = wb.active  # pakai sheet pertama/aktif
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None or row[1] is None:
                continue
            raw_material = row[0]
            # Normalisasi: 310049.0 / "310049" / 310049 → "310049"
            try:
                if isinstance(raw_material, float) and raw_material.is_integer():
                    material = str(int(raw_material))
                elif isinstance(raw_material, int):
                    material = str(raw_material)
                else:
                    material = str(raw_material).strip()
                    if material.endswith(".0"):
                        material = material[:-2]
            except Exception:
                material = str(raw_material).strip()

            uom = str(row[1]).strip()
            if material and uom:
                mapping[material] = uom

        log.info(f"[MAPPING] {len(mapping)} material UoM berhasil dibaca")
    except Exception as e:
        log.warning(
            f"[MAPPING] Gagal baca material UoM mapping: {e} — "
            f"semua material akan pakai default '{DEFAULT_UOM}'"
        )

    return mapping


# ─────────────────────────────────────────────
# BUILD HTML BODY EMAIL PER PLANT
# ─────────────────────────────────────────────

def _build_body_html_per_plant(plant: str, plant_name: str,
                                items: list, posting_date: str,
                                uom_map: dict = None) -> str:
    """
    Buat body HTML email untuk satu plant.
    Format tabel: Material | Selisih2 | UOM | Gudang | Plant | Adj | Plant Name
    Sesuai template standar dari gambar referensi.
    uom_map: { material: uom } daftar pengecualian material non-CAR —
             material yang tidak ditemukan di mapping tetap "CAR".
    """
    if uom_map is None:
        uom_map = {}

    sorted_items = sorted(items, key=lambda x: (x.mvt_type, x.sloc, x.material))
    sender_name = get_sender_display_name(
    load_credentials()["email_from"]
    )
    th_style = (
        "padding:7px 12px;background:#1F5C99;color:white;"
        "font-family:Calibri,Arial;font-size:12px;border:1px solid #CBD5E1;"
    )
    td_base = (
        "padding:6px 12px;font-family:Calibri,Arial;font-size:12px;"
        "border:1px solid #CBD5E1;text-align:center;"
    )

    header_row = (
        f"<tr>"
        f"<th style='{th_style}'>Material</th>"
        f"<th style='{th_style}'>Selisih2</th>"
        f"<th style='{th_style}'>UOM</th>"
        f"<th style='{th_style}'>Gudang</th>"
        f"<th style='{th_style}'>Plant</th>"
        f"<th style='{th_style}'>Adj</th>"
        f"<th style='{th_style}'>Plant Name</th>"
        f"</tr>"
    )

    data_rows = ""
    for i, item in enumerate(sorted_items):
        bg  = "#EBF3FB" if i % 2 == 0 else "#FFFFFF"
        td  = td_base + f"background:{bg};"
        # Format selisih 3 desimal pakai koma (standar Indonesia)
        selisih_fmt = f"{abs(item.diff):,.3f}".replace(",", "X").replace(".", ",").replace("X", ".")
        uom = uom_map.get(str(item.material).strip(), DEFAULT_UOM)
        data_rows += (
            f"<tr>"
            f"<td style='{td}'>{item.material}</td>"
            f"<td style='{td}'>{selisih_fmt}</td>"
            f"<td style='{td}'>{uom}</td>"
            f"<td style='{td}'>{item.sloc}</td>"
            f"<td style='{td}'>{item.plant}</td>"
            f"<td style='{td}'>{item.mvt_type}</td>"
            f"<td style='{td}'>{plant_name}</td>"
            f"</tr>"
        )

    total_917 = len([x for x in sorted_items if x.mvt_type == "917"])
    total_918 = len([x for x in sorted_items if x.mvt_type == "918"])

    body_html = f"""<html>
<body style="font-family:Calibri,Arial;font-size:13px;color:#1E293B;margin:0;padding:0">

<p style="margin:0 0 8px 0">Dear Team Accounting,</p>

<p style="margin:0 0 16px 0">
  Mohon dibantu untuk dilakukan Adjustment atas Selisih Endstock EOD:
</p>

<table border="0" cellpadding="0" cellspacing="0"
       style="border-collapse:collapse;margin-bottom:4px">
  {header_row}
  {data_rows}
</table>

<p style="margin:12px 0 4px 0;font-size:12px;color:#475569">
  Total item: <b>{len(sorted_items)}</b> &nbsp;|&nbsp;
  Mvt 917 (Kurangi SAP): <b>{total_917}</b> &nbsp;|&nbsp;
  Mvt 918 (Tambah SAP): <b>{total_918}</b>
</p>

<p style="margin:16px 0 0 0;font-size:13px;color:#1E293B">
    Terima Kasih,<br>
    <b>{sender_name}</b>
</p>

</body>
</html>"""
    return body_html

#BUAT SIGNATURE EMAIL DARI EMAIL PENGIRIMMaa
def get_sender_display_name(email_address: str) -> str:
    """
    Convert email menjadi nama display.

    Contoh:
    rafael.arrelano@mayora.co.id
    -> Rafael Arrelano
    """

    username = email_address.split("@")[0]

    # ganti titik & underscore jadi spasi
    username = username.replace(".", " ").replace("_", " ")

    # kapital tiap kata
    return username.title()


# ─────────────────────────────────────────────
# HELPER: FIND THUNDERBIRD DRAFTS FOLDER
# ─────────────────────────────────────────────

def _find_thunderbird_profile_dir() -> str:
    """
    Cari folder profil Thunderbird aktif via profiles.ini.

    Prioritas:
    1. Block [Install...] → key Default= (satu baris, tanpa DOTALL)
    2. Profil bernama "default-release"  ← Thunderbird modern
    3. [Profile...] dengan IsDefault=1
    4. Profil pertama yang ditemukan (fallback)
    """
    import re as _re
    appdata = os.environ.get("APPDATA", "")
    if not appdata:
        return ""
    tb_base      = os.path.join(appdata, "Thunderbird")
    profiles_ini = os.path.join(tb_base, "profiles.ini")
    if not os.path.exists(profiles_ini):
        return ""
    with open(profiles_ini, "r", encoding="utf-8", errors="ignore") as f:
        ini_content = f.read()

    # ── Prioritas 1: block [Install...] → Default= ──────────
    # MULTILINE tanpa DOTALL: [^\n\r]+ hanya ambil satu baris
    install_m = _re.search(
        r'^\[Install[^\]]+\][^\[]*?^Default=([^\n\r]+)',
        ini_content, _re.MULTILINE
    )
    if install_m:
        raw  = install_m.group(1).strip()
        path = os.path.join(tb_base, raw.replace("/", os.sep)) \
               if not os.path.isabs(raw) else raw
        if os.path.isdir(path):
            log.debug(f"[THUNDERBIRD] Profil dari Install block: {path}")
            return path

    # ── Kumpulkan semua profil yang valid ─────────────────────
    all_profiles = []   # list of (path, is_default, name, raw_path)
    blocks = _re.split(r'\[Profile\d+\]', ini_content)
    for block in blocks:
        path_m = _re.search(r'^Path=([^\n\r]+)',  block, _re.MULTILINE)
        def_m  = _re.search(r'^IsDefault=1',       block, _re.MULTILINE)
        rel_m  = _re.search(r'^IsRelative=1',      block, _re.MULTILINE)
        name_m = _re.search(r'^Name=([^\n\r]+)',  block, _re.MULTILINE)
        if not path_m:
            continue
        raw   = path_m.group(1).strip()
        name  = name_m.group(1).strip() if name_m else ""
        path  = os.path.join(tb_base, raw.replace("/", os.sep)) if rel_m else raw
        if not os.path.isdir(path):
            continue
        all_profiles.append((path, bool(def_m), name, raw))

    # ── Prioritas 2: profil "default-release" ────────────────
    for path, _, name, raw in all_profiles:
        if "default-release" in raw or "default-release" in name:
            log.debug(f"[THUNDERBIRD] Profil default-release: {path}")
            return path

    # ── Prioritas 3: IsDefault=1 ──────────────────────────────
    for path, is_default, _, _ in all_profiles:
        if is_default:
            log.debug(f"[THUNDERBIRD] Profil IsDefault=1: {path}")
            return path

    # ── Prioritas 4: profil pertama yang ada ──────────────────
    if all_profiles:
        log.debug(f"[THUNDERBIRD] Profil fallback: {all_profiles[0][0]}")
        return all_profiles[0][0]

    return ""
def _find_thunderbird_drafts_mbox(profile_dir: str, email_from: str = "") -> str:
    """
    Cari path file mbox 'Drafts' Thunderbird.

    Thunderbird menyimpan draft sebagai file mbox bernama 'Drafts'
    (tanpa ekstensi). Untuk POP3, lokasinya bisa di direktori custom
    yang di-set user, bukan di dalam folder profil Thunderbird.

    Strategi pencarian:
      1. Baca prefs.js → ambil semua nilai mail.server.serverN.directory
         Ini menangani POP3 dengan direktori custom (misal D:\\Email Rafael)
      2. Cek subfolder Mail/ dan ImapMail/ di dalam profile_dir (default)

    Jika direktori ditemukan tapi file Drafts belum ada, path-nya tetap
    dikembalikan agar bisa dibuat otomatis (mode append akan create file baru).

    Return: path file mbox Drafts (ada atau akan dibuat), atau "" jika tidak ada.
    """
    import re as _re

    domain     = email_from.split("@")[-1] if "@" in email_from else ""
    username   = email_from.split("@")[0].lower() if "@" in email_from else ""

    # Kumpulkan semua direktori kandidat yang akan dicek
    dir_candidates = []

    # ── Sumber 1: baca prefs.js ───────────────────────────────
    if profile_dir and os.path.isdir(profile_dir):
        prefs_file = os.path.join(profile_dir, "prefs.js")
        if os.path.exists(prefs_file):
            try:
                with open(prefs_file, "r", encoding="utf-8", errors="ignore") as f:
                    prefs_content = f.read()

                # Cari: user_pref("mail.server.server4.directory", "D:\\Email Rafael");
                # Thunderbird tulis backslash sebagai \\  di prefs.js
                # Saat Python baca file teks, \\ tetap \\ (dua karakter)
                # → perlu decode: ganti \\ → \  dan / → \  untuk Windows path
                matches = _re.findall(
                    r'user_pref\("mail\.server\.server\d+\.directory",\s*"([^"]+)"\)',
                    prefs_content
                )
                for raw in matches:
                    # Decode escape: \\ → \  (Thunderbird escape style)
                    decoded = raw.replace("\\\\", "\\").replace("/", os.sep)
                    log.debug(f"[THUNDERBIRD] prefs.js directory: {raw!r} → {decoded!r}")
                    if os.path.isdir(decoded):
                        # Prioritaskan direktori yang namanya mengandung email/domain
                        if domain and domain.lower() in decoded.lower():
                            dir_candidates.insert(0, decoded)
                        elif username and username in decoded.lower():
                            dir_candidates.insert(0, decoded)
                        else:
                            dir_candidates.append(decoded)
            except Exception as e:
                log.debug(f"[THUNDERBIRD] Gagal baca prefs.js: {e}")

    # ── Sumber 2: subfolder Mail/ImapMail di dalam profil ────
    if profile_dir and os.path.isdir(profile_dir):
        for subfolder in ("Mail", "ImapMail"):
            base = os.path.join(profile_dir, subfolder)
            if not os.path.isdir(base):
                continue
            for acct in os.listdir(base):
                acct_path = os.path.join(base, acct)
                if not os.path.isdir(acct_path):
                    continue
                if domain and domain in acct:
                    dir_candidates.insert(0, acct_path)
                else:
                    dir_candidates.append(acct_path)

    # ── Cari / siapkan file Drafts dari semua direktori kandidat ─
    fallback_create = ""   # direktori ada tapi Drafts belum ada

    for d in dir_candidates:
        drafts_path = os.path.join(d, "Drafts")
        if os.path.isfile(drafts_path):
            log.debug(f"[THUNDERBIRD] Drafts ditemukan: {drafts_path}")
            return drafts_path
        elif not fallback_create:
            # Simpan sebagai fallback: direktori valid, Drafts belum ada
            fallback_create = drafts_path

    # File Drafts belum ada, tapi direktori valid → kembalikan path untuk dibuat
    if fallback_create:
        log.info(
            f"[THUNDERBIRD] Drafts belum ada, akan dibuat: {fallback_create}"
        )
        return fallback_create

    return ""


# ─────────────────────────────────────────────
# BUAT DRAFT EMAIL DI THUNDERBIRD (.eml)
# ─────────────────────────────────────────────

def _create_thunderbird_draft(cred: dict, subject: str, body_html: str,
                              to: str, cc: str,
                              draft_folder: str = None) -> str:
    """
    Simpan draft email sebagai file .eml di folder report.
    User membuka sendiri file .eml di Thunderbird.
    Tidak ada attachment — seluruh isi laporan ada di body HTML.

    Return: path file .eml yang dibuat.
    """
    import email.utils

    # ── Buat objek email ──────────────────────────────────────
    msg = MIMEMultipart("mixed")
    msg["Subject"] = subject
    msg["From"]    = cred["email_from"]
    msg["To"]      = to
    if cc:
        msg["Cc"]  = cc
    msg["Date"]    = email.utils.formatdate(localtime=True)
    msg["X-Mozilla-Draft-Info"] = (
        "internal/draft; vcard=0; receipt=0; DSN=0; uuencode=0"
    )
    msg["X-Mailer"] = "RPA Stock Recon"
    msg.attach(MIMEText(body_html, "html", "utf-8"))

    # ── Simpan sebagai .eml di folder report ─────────────────
    if draft_folder is None:
        draft_folder = Config.FOLDER_REPORT
    os.makedirs(draft_folder, exist_ok=True)

    ts        = datetime.now().strftime("%Y%m%d_%H%M%S_%f")[:-3]
    clean_sub = "".join(
        c for c in subject if c.isalnum() or c in ("-", "_", " ")
    ).strip()[:50]
    eml_path  = os.path.join(draft_folder, f"Draft_{ts}_{clean_sub}.eml")

    with open(eml_path, "w", encoding="utf-8") as f:
        f.write(msg.as_string())

    log.info(f"[DRAFT] ✓ Draft dibuat: {os.path.basename(eml_path)}")
    log.info(f"[DRAFT] Path: {eml_path}")
    return eml_path



# ─────────────────────────────────────────────
# FUNGSI UTAMA — KIRIM SATU EMAIL PER PLANT
# ─────────────────────────────────────────────

def send_stock_diff_report(
    items_per_plant: dict,
    override_to: str = None,
    override_cc: str = None,
) -> int:
    """
    Buat SATU DRAFT EMAIL PER PLANT di Mozilla Thunderbird (.eml files).
    Tidak ada file Excel yang dibuat dan tidak ada attachment di email —
    isi laporan sepenuhnya ada di badan (body) email dalam bentuk tabel HTML.

    Hanya item FSTKGD yang masuk draft email.
    FSTKVN tetap diproses untuk compare & U2C, tapi tidak masuk laporan email.

    Subject per plant:
        Req. Adj. EOD Plant {code} {name} Tanggal {tanggal posting}

    Draft email disimpan sebagai file .eml di folder report.
    File bisa dibuka langsung dengan Thunderbird atau aplikasi email lainnya.

    override_to / override_cc: override nilai dari kredensial tersimpan
    Return: jumlah draft email yang berhasil dibuat (0 jika tidak ada selisih
            FSTKGD atau semua draft gagal dibuat)
    """
    # Filter: hanya FSTKGD yang masuk email
    items_email = {
        plant: [i for i in items if i.param == "FSTKGD"]
        for plant, items in items_per_plant.items()
    }
    items_email = {p: v for p, v in items_email.items() if v}   # hapus plant kosong

    if not items_email:
        log.info("[REPORT] Tidak ada selisih FSTKGD — draft email tidak dibuat")
        return 0

    cred = load_credentials()

    if override_to:
        cred["email_to"] = override_to
    if override_cc is not None:
        cred["email_cc"] = override_cc

    # Load mapping plant name & material UoM dari Excel
    plant_name_map = load_plant_name_mapping()
    uom_map        = load_material_uom_mapping()

    os.makedirs(Config.FOLDER_REPORT, exist_ok=True)

    # ── Buat satu draft email per plant ──────────────────────
    draft_count = 0
    drafts_created = []
    
    for plant, items in sorted(items_email.items()):
        if not items:
            continue

        plant_name   = plant_name_map.get(plant, "")
        display_name = f"{plant} {plant_name}".strip()

        # Posting date dari item pertama (format dd.mm.yyyy)
        posting_date = items[0].posting_date

        # Subject sesuai template standar
        subject = f"Req.Adj.EOD Plant {display_name} Tgl {posting_date}"

        # Body email per plant dengan tabel adjustment
        body_html = _build_body_html_per_plant(
            plant        = plant,
            plant_name   = plant_name,
            items        = items,
            posting_date = posting_date,
            uom_map      = uom_map,
        )

        log.info(f"[DRAFT] Buat draft plant {display_name} → {cred['email_to']}")
        log.info(f"[DRAFT] Subject: {subject}")

        try:
            draft_path = _create_thunderbird_draft(
                cred      = cred,
                subject   = subject,
                body_html = body_html,
                to        = cred["email_to"],
                cc        = cred.get("email_cc", ""),
            )
            draft_count += 1
            drafts_created.append(draft_path)
            log.info(f"[DRAFT] ✓ Plant {plant} draft dibuat")
        except Exception as e:
            log.error(f"[DRAFT] ✗ Plant {plant} gagal: {e}")
            continue  # Lanjut ke plant berikutnya meski ada error

    total_plant = len(items_email)
    total_item  = sum(len(v) for v in items_email.values())
    fstkvn_skip = sum(
        len([i for i in v if i.param == "FSTKVN"])
        for v in items_per_plant.values()
    )
    
    log.info(
        f"[REPORT] Selesai | {draft_count}/{total_plant} plant draft dibuat | "
        f"{total_item} item FSTKGD | {fstkvn_skip} item FSTKVN dilewati (tidak ada draft email)"
    )
    
    # Log informasi draft folder
    if drafts_created:
        log.info(f"[REPORT] Draft email tersimpan di: {Config.FOLDER_REPORT}")
        log.info(f"[REPORT] Buka folder dan double-click file .eml untuk membuka di Thunderbird")
        
    return draft_count