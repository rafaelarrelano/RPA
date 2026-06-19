"""
extend_material.py
Extend Material Module — RPA PT Mayora Indah Tbk

Launched from launcher.py — receives back_callback to return to menu.

This module wraps the existing standalone `sap_rpa_mmsc_extend.py` script
(unchanged — already tested and working) inside a launcher-compatible GUI
class. The automation logic (SAP keyboard/clipboard driving, MMSC screen
handling, storage location detection) is untouched; only the presentation
layer is adapted to embed inside the launcher's single window instead of
opening its own separate Tk window.

Flow:
  1. User loads master_data.xlsx (MATERIAL + PLANT sheets)
  2. User selects materials and plants via checklists
  3. Click Run → writes _selected_temp.xlsx → spawns sap_rpa_mmsc_extend.py
     as a subprocess and streams its stdout into the activity log
  4. Per-step status (running/ok/skip/error) is parsed from log lines and
     reflected in the checklists and a live counter
  5. Report saved to output_rpa/hasil_extend_<timestamp>.xlsx by the
     underlying script itself

No SAP COM / SAP GUI scripting used — sap_rpa_mmsc_extend.py drives SAP
purely via pyautogui keyboard/clipboard, same approach as Maintain Material.
"""

import os
import re
import sys
import time
import threading
import subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    pass

try:
    from theme_manager import get_theme_manager
except ImportError:
    get_theme_manager = None

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

FONT       = "Segoe UI"
FONT_TITLE = ("Segoe UI", 12, "bold")
FONT_SUB   = ("Segoe UI", 9)
FONT_SMALL = ("Segoe UI", 8)

SCRIPT_NAME      = "sap_rpa_mmsc_extend.py"
MASTER_FILE_NAME = "master_data.xlsx"
TEMP_INPUT_NAME  = "_selected_temp.xlsx"


def _python_executable() -> str:
    """Use the same interpreter currently running launcher.py."""
    return sys.executable


# ─────────────────────────────────────────────
# EXTEND MATERIAL GUI CLASS
# ─────────────────────────────────────────────

class ExtendMaterialGui:
    """
    Extend Material module — launched from launcher.py.
    parent         : tk.Frame provided by launcher
    back_callback  : callable to return to main menu
    theme_manager  : ThemeManager instance (optional)
    """

    def __init__(self, parent: tk.Frame, back_callback=None, theme_manager=None):
        self.parent        = parent
        self.back_callback = back_callback
        self.theme = theme_manager or (get_theme_manager() if get_theme_manager else None)

        self.dry_var      = tk.BooleanVar(value=False)
        self.status_var   = tk.StringVar(value="Ready")
        self.process       = None
        self.master_path   = os.path.join(os.getcwd(), MASTER_FILE_NAME)
        self.mat_vars      = []
        self.plant_vars    = []
        self._mat_search_var = tk.StringVar()
        self._plt_search_var = tk.StringVar()

        self._ok_count    = 0
        self._err_count   = 0
        self._total_count = 0
        self._start_time  = None
        self._final_elapsed = 0
        self._mat_progress  = {}
        self._current_mat   = None
        self._current_plt   = None

        self.parent.tk.call('tk', 'scaling', 2.0)

        self._refresh_colors()
        self._build_ui()
        self._load_master()

    def _refresh_colors(self):
        if self.theme:
            self.BG        = self.theme.get_color("bg")
            self.BG_CARD   = self.theme.get_color("bg_card")
            self.BG_INPUT  = self.theme.get_color("input_bg")
            self.TEXT      = self.theme.get_color("text_dark")
            self.TEXT2     = self.theme.get_color("text_mid")
            self.TEXT3     = self.theme.get_color("text_light")
            self.BORDER    = self.theme.get_color("border")
            self.SUCCESS   = self.theme.get_color("success")
            self.WARNING   = self.theme.get_color("warning")
            self.DANGER    = self.theme.get_color("danger")
            self.BG_DARK   = self.theme.get_color("bg_active")
        else:
            self.BG        = "#F5F5F5"
            self.BG_CARD   = "#FFFFFF"
            self.BG_INPUT  = "#FFFFFF"
            self.TEXT      = "#1A1A1A"
            self.TEXT2     = "#555555"
            self.TEXT3     = "#999999"
            self.BORDER    = "#E0E0E0"
            self.SUCCESS   = "#2E7D32"
            self.WARNING   = "#E65100"
            self.DANGER    = "#C62828"
            self.BG_DARK   = "#1A1A1A"

    # ── BUILD UI ─────────────────────────────────────────────

    def _build_ui(self):
        self.parent.configure(bg=self.BG)

        main = tk.Frame(self.parent, bg=self.BG)
        main.pack(fill="both", expand=True)

        # ── LEFT PANEL (scrollable) ───────────────────────────
        left_outer = tk.Frame(main, bg=self.BG, width=360)
        left_outer.pack(side="left", fill="both", padx=(32, 0), pady=24)
        left_outer.pack_propagate(False)

        left_canvas = tk.Canvas(left_outer, bg=self.BG, highlightthickness=0, width=340)
        left_vsb = tk.Scrollbar(left_outer, orient="vertical", command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_vsb.set)
        left_vsb.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)

        left = tk.Frame(left_canvas, bg=self.BG)
        _win = left_canvas.create_window((0, 0), window=left, anchor="nw")

        def _on_left_cfg(e):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))

        def _on_canvas_cfg(e):
            left_canvas.itemconfig(_win, width=e.width)

        left.bind("<Configure>", _on_left_cfg)
        left_canvas.bind("<Configure>", _on_canvas_cfg)

        def _scroll_left(e):
            left_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")

        def _on_mousewheel_global(event):
            try:
                x, y = event.x_root, event.y_root
                x1 = left_outer.winfo_rootx()
                y1 = left_outer.winfo_rooty()
                x2 = x1 + left_outer.winfo_width()
                y2 = y1 + left_outer.winfo_height()
                if x1 <= x <= x2 and y1 <= y <= y2:
                    _scroll_left(event)
            except Exception:
                pass

        left_canvas.bind("<MouseWheel>", _scroll_left)
        self.parent.bind_all("<MouseWheel>", _on_mousewheel_global)

        tk.Label(left, text="Extend Material", font=("Segoe UI", 15, "bold"),
                 fg=self.TEXT, bg=self.BG).pack(anchor="w", pady=(0, 2))
        tk.Label(left, text="SAP MMSC — Extend Storage Locations (WH02 & WT01)",
                 font=(FONT, 8), fg=self.TEXT3, bg=self.BG).pack(anchor="w", pady=(0, 4))

        notice = tk.Frame(left, bg="#FFF8E1", highlightbackground="#FFB300", highlightthickness=1)
        notice.pack(fill="x", pady=(0, 12))
        tk.Label(notice,
                 text="⚠  Open SAP GUI and log in before clicking Run.\n"
                      "   Robot types like a human — no SAP scripting needed.",
                 font=(FONT, 8), fg="#7A5400", bg="#FFF8E1",
                 justify="left", padx=8, pady=6).pack(anchor="w")

        # ── SECTION: Master File ──────────────────────────────
        self._section_label(left, "Master Data")
        f_card = self._card(left)

        self._master_lbl = tk.Label(
            f_card, text="No file loaded", font=(FONT, 9), fg=self.TEXT2,
            bg=self.BG_CARD, anchor="w", wraplength=260, justify="left")
        self._master_lbl.pack(anchor="w")

        self._master_info_lbl = tk.Label(
            f_card, text="", font=(FONT, 8), fg=self.TEXT3, bg=self.BG_CARD, anchor="w")
        self._master_info_lbl.pack(anchor="w", pady=(2, 6))

        tk.Button(f_card, text="📁  Change File", font=(FONT, 8),
                  fg=self.TEXT2, bg=self.BG_CARD, relief="flat", bd=0,
                  cursor="hand2", command=self._browse_master_file,
                  ).pack(anchor="w")

        # ── SECTION: Materials ────────────────────────────────
        self._section_label(left, "Select Materials")
        m_card = self._card(left)

        mh = tk.Frame(m_card, bg=self.BG_CARD)
        mh.pack(fill="x", pady=(0, 4))
        self._mat_count_lbl = tk.Label(mh, text="Load master file first",
                                        font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD)
        self._mat_count_lbl.pack(side="left")
        tk.Button(mh, text="All", font=(FONT, 7), fg=self.TEXT2, bg=self.BG_CARD,
                  relief="flat", bd=0, cursor="hand2",
                  command=lambda: self._select_all(self.mat_vars, self._update_mat_count),
                  ).pack(side="right", padx=(4, 0))
        tk.Button(mh, text="Clear", font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD,
                  relief="flat", bd=0, cursor="hand2",
                  command=lambda: self._clear_all(self.mat_vars, self._update_mat_count),
                  ).pack(side="right")

        msr = tk.Frame(m_card, bg=self.BG_CARD)
        msr.pack(fill="x", pady=(2, 4))
        tk.Label(msr, text="🔍", font=(FONT, 9), fg=self.TEXT3, bg=self.BG_CARD).pack(side="left", padx=(0, 4))
        self._mat_search_var.trace_add("write", lambda *a: self._filter_checklist(
            self.mat_vars, self._mat_search_var.get(), self.mat_grid_canvas))
        tk.Entry(msr, textvariable=self._mat_search_var, font=(FONT, 9),
                 bg=self.BG_INPUT, fg=self.TEXT, insertbackground=self.TEXT,
                 relief="flat", bd=0, highlightthickness=1,
                 highlightbackground=self.BORDER, highlightcolor=self.BG_DARK,
                 ).pack(side="left", fill="x", expand=True, ipady=4)

        self.mat_grid_canvas, self.mat_grid = self._make_checklist_area(m_card, height=140)

        # ── SECTION: Plants ───────────────────────────────────
        self._section_label(left, "Select Plants")
        p_card = self._card(left)

        ph = tk.Frame(p_card, bg=self.BG_CARD)
        ph.pack(fill="x", pady=(0, 4))
        self._plt_count_lbl = tk.Label(ph, text="Load master file first",
                                        font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD)
        self._plt_count_lbl.pack(side="left")
        tk.Button(ph, text="All", font=(FONT, 7), fg=self.TEXT2, bg=self.BG_CARD,
                  relief="flat", bd=0, cursor="hand2",
                  command=lambda: self._select_all(self.plant_vars, self._update_plt_count),
                  ).pack(side="right", padx=(4, 0))
        tk.Button(ph, text="Clear", font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD,
                  relief="flat", bd=0, cursor="hand2",
                  command=lambda: self._clear_all(self.plant_vars, self._update_plt_count),
                  ).pack(side="right")

        psr = tk.Frame(p_card, bg=self.BG_CARD)
        psr.pack(fill="x", pady=(2, 4))
        tk.Label(psr, text="🔍", font=(FONT, 9), fg=self.TEXT3, bg=self.BG_CARD).pack(side="left", padx=(0, 4))
        self._plt_search_var.trace_add("write", lambda *a: self._filter_checklist(
            self.plant_vars, self._plt_search_var.get(), self.plt_grid_canvas))
        tk.Entry(psr, textvariable=self._plt_search_var, font=(FONT, 9),
                 bg=self.BG_INPUT, fg=self.TEXT, insertbackground=self.TEXT,
                 relief="flat", bd=0, highlightthickness=1,
                 highlightbackground=self.BORDER, highlightcolor=self.BG_DARK,
                 ).pack(side="left", fill="x", expand=True, ipady=4)

        self.plt_grid_canvas, self.plt_grid = self._make_checklist_area(p_card, height=140)

        # ── SECTION: Options ──────────────────────────────────
        self._section_label(left, "Options")
        o_card = self._card(left)
        tk.Checkbutton(
            o_card, text="Dry run (simulate, no SAP save)",
            variable=self.dry_var, font=(FONT, 9), fg=self.TEXT2, bg=self.BG_CARD,
            selectcolor=self.BG_CARD, activebackground=self.BG_CARD,
            activeforeground=self.TEXT2, relief="flat",
        ).pack(anchor="w")
        tk.Label(o_card, text="Storage locations to extend: WH02 & WT01",
                 font=(FONT, 8), fg=self.TEXT3, bg=self.BG_CARD).pack(anchor="w", pady=(4, 0))

        # ── SECTION: Info ──────────────────────────────────────
        self._section_label(left, "Info")
        i_card = self._card(left)
        for dot_c, msg in [
            (self.SUCCESS, "OK = storage location added"),
            (self.WARNING, "SKIP = already exists or locked"),
            (self.DANGER,  "ERROR = unexpected SAP issue"),
            (self.WARNING, "SAP must be open & logged in"),
            (self.WARNING, "Do NOT touch mouse/keyboard while running"),
        ]:
            rf = tk.Frame(i_card, bg=self.BG_CARD)
            rf.pack(fill="x", pady=1)
            tk.Label(rf, text="●", fg=dot_c, bg=self.BG_CARD, font=(FONT, 8)).pack(side="left", padx=(0, 6))
            tk.Label(rf, text=msg, fg=self.TEXT2, bg=self.BG_CARD, font=(FONT, 8)).pack(side="left")

        # ── RUN / STOP buttons ─────────────────────────────────
        btn_frame = tk.Frame(left, bg=self.BG)
        btn_frame.pack(fill="x", pady=(16, 0))

        self.stop_btn = tk.Button(
            btn_frame, text="■  Stop", font=("Segoe UI", 10, "bold"),
            fg=self.BG_CARD, bg="#8B0000", activebackground="#C62828",
            activeforeground=self.BG_CARD, relief="flat", bd=0,
            padx=14, pady=10, cursor="hand2", state="disabled", command=self._on_stop,
        )
        self.stop_btn.pack(side="right", padx=(8, 0))

        self.run_btn = tk.Button(
            btn_frame, text="▶   Run MMSC", font=("Segoe UI", 11, "bold"),
            fg=self.BG_CARD, bg=self.BG_DARK, activebackground="#333333",
            activeforeground=self.BG_CARD, relief="flat", bd=0,
            padx=20, pady=10, cursor="hand2", command=self._on_run,
        )
        self.run_btn.pack(side="left", fill="x", expand=True)

        # ── Counter bar ──────────────────────────────────────
        self._counter_lbl = tk.Label(left, text="", font=(FONT, 8, "bold"),
                                      fg=self.TEXT2, bg=self.BG)
        self._counter_lbl.pack(anchor="w", pady=(8, 0))

        # ── RIGHT PANEL: log ───────────────────────────────────
        right = tk.Frame(main, bg=self.BG)
        right.pack(side="left", fill="both", expand=True, padx=24, pady=24)

        log_hdr = tk.Frame(right, bg=self.BG)
        log_hdr.pack(fill="x", pady=(0, 6))
        tk.Label(log_hdr, text="ACTIVITY LOG", font=(FONT, 8, "bold"),
                 fg=self.TEXT3, bg=self.BG).pack(side="left")
        tk.Button(log_hdr, text="Clear", font=(FONT, 7), fg=self.TEXT3, bg=self.BG,
                  relief="flat", bd=0, cursor="hand2", command=self._clear_log,
                  ).pack(side="right")
        tk.Button(log_hdr, text="📂 Output", font=(FONT, 7), fg=self.TEXT3, bg=self.BG,
                  relief="flat", bd=0, cursor="hand2", command=self._open_output,
                  ).pack(side="right", padx=(0, 8))

        log_outer = tk.Frame(right, bg=self.BORDER, bd=0)
        log_outer.pack(fill="both", expand=True)
        log_inner = tk.Frame(log_outer, bg=self.BG_CARD)
        log_inner.pack(fill="both", padx=1, pady=1)

        self.log_box = scrolledtext.ScrolledText(
            log_inner, bg=self.BG_CARD, fg=self.TEXT2, font=(FONT, 9),
            relief="flat", bd=0, state="disabled", wrap="word", padx=12, pady=10,
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("OK", foreground=self.SUCCESS)
        self.log_box.tag_config("ERROR", foreground=self.DANGER)
        self.log_box.tag_config("WARN", foreground=self.WARNING)
        self.log_box.tag_config("INFO", foreground=self.TEXT2)

        self._write_log("Extend Material ready.", "INFO")
        self._write_log("1. Confirm master data loaded", "INFO")
        self._write_log("2. Open SAP GUI and log in", "INFO")
        self._write_log("3. Select materials & plants → Run MMSC", "INFO")

    # ── UI HELPERS ───────────────────────────────────────────

    def _section_label(self, parent, text):
        f = tk.Frame(parent, bg=self.BG)
        f.pack(fill="x", pady=(12, 4))
        tk.Frame(f, bg=self.BG_DARK, width=3, height=14).pack(side="left", padx=(0, 6))
        tk.Label(f, text=text.upper(), font=(FONT, 8, "bold"),
                 fg=self.TEXT, bg=self.BG).pack(side="left")

    def _card(self, parent):
        outer = tk.Frame(parent, bg=self.BORDER)
        outer.pack(fill="x", pady=(0, 4))
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=14, pady=10)
        inner.pack(fill="x", padx=1, pady=1)
        return inner

    def _make_checklist_area(self, parent, height=140):
        canvas = tk.Canvas(parent, bg=self.BG_CARD, highlightthickness=0, height=height)
        vsb = tk.Scrollbar(parent, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side="right", fill="y")
        canvas.pack(fill="both", expand=True)

        grid = tk.Frame(canvas, bg=self.BG_CARD)
        win = canvas.create_window((0, 0), window=grid, anchor="nw")

        def _on_cfg(e):
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _on_canvas_cfg(e):
            canvas.itemconfig(win, width=e.width)

        grid.bind("<Configure>", _on_cfg)
        canvas.bind("<Configure>", _on_canvas_cfg)

        def _scroll(e):
            canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
            return "break"

        canvas.bind("<MouseWheel>", _scroll)
        grid.bind("<MouseWheel>", _scroll)

        return canvas, grid

    def _build_checklist(self, grid, canvas, var_list, items):
        for w in grid.winfo_children():
            w.destroy()
        var_list.clear()
        for i, item in enumerate(items):
            var = tk.BooleanVar(value=True)
            cb = tk.Checkbutton(
                grid, text=item, variable=var, font=(FONT, 8),
                fg=self.TEXT2, bg=self.BG_CARD, selectcolor=self.BG_CARD,
                activebackground=self.BG_CARD, activeforeground=self.TEXT2,
                relief="flat", bd=0,
            )
            cb.grid(row=i, column=0, sticky="w", padx=4, pady=1)
            cb.bind("<MouseWheel>", lambda e, c=canvas: c.yview_scroll(
                int(-1 * (e.delta / 120)), "units") or "break")
            var_list.append({"val": item, "var": var, "chk": cb, "orig": item})

    def _filter_checklist(self, var_list, keyword, canvas):
        kw = keyword.strip().lower()
        for x in var_list:
            if kw == "" or kw in x["orig"].lower():
                x["chk"].grid()
            else:
                x["chk"].grid_remove()

    def _select_all(self, var_list, upd_fn):
        for x in var_list:
            x["var"].set(True)
        upd_fn()

    def _clear_all(self, var_list, upd_fn):
        for x in var_list:
            x["var"].set(False)
        upd_fn()

    def _get_selected(self, var_list):
        return [x["val"] for x in var_list if x["var"].get()]

    def _update_mat_count(self):
        n = sum(1 for x in self.mat_vars if x["var"].get())
        self._mat_count_lbl.config(text=f"{n} of {len(self.mat_vars)} materials")

    def _update_plt_count(self):
        n = sum(1 for x in self.plant_vars if x["var"].get())
        self._plt_count_lbl.config(text=f"{n} of {len(self.plant_vars)} plants")

    # ── MASTER FILE ──────────────────────────────────────────

    def _browse_master_file(self):
        path = filedialog.askopenfilename(
            title="Select Master Data Excel File",
            filetypes=[("Excel File", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if not path:
            return
        self.master_path = path
        self._load_master()

    def _load_master(self):
        if not os.path.exists(self.master_path):
            self._master_lbl.config(text=f"⚠ {os.path.basename(self.master_path)} not found", fg=self.WARNING)
            return
        try:
            wb = openpyxl.load_workbook(self.master_path, data_only=True)
            ws_mat = wb["MATERIAL"] if "MATERIAL" in wb.sheetnames else wb.active
            mats = [str(r[0]).strip() for r in ws_mat.iter_rows(min_row=2, values_only=True) if r[0]]
            ws_plt = wb["PLANT"] if "PLANT" in wb.sheetnames else wb.active
            plts = [str(r[0]).strip().upper() for r in ws_plt.iter_rows(min_row=2, values_only=True) if r[0]]

            self._build_checklist(self.mat_grid, self.mat_grid_canvas, self.mat_vars, mats)
            self._build_checklist(self.plt_grid, self.plt_grid_canvas, self.plant_vars, plts)
            self._update_mat_count()
            self._update_plt_count()

            self._master_lbl.config(text=f"📄  {os.path.basename(self.master_path)}", fg=self.TEXT)
            self._master_info_lbl.config(text=f"✓  {len(mats)} materials  ·  {len(plts)} plants", fg=self.SUCCESS)
            self._write_log(f"Master data loaded: {len(mats)} materials, {len(plts)} plants", "OK")
            self._load_selection()
        except Exception as e:
            self._master_lbl.config(text="⚠ Failed to load", fg=self.DANGER)
            self._master_info_lbl.config(text=str(e), fg=self.DANGER)
            self._write_log(f"Master data load error: {e}", "ERROR")

    # ── SELECTION PERSISTENCE ─────────────────────────────────

    def _save_selection(self):
        try:
            sel_mats = self._get_selected(self.mat_vars)
            sel_plts = self._get_selected(self.plant_vars)
            cfg_path = os.path.join(os.getcwd(), ".rpa_selection.cfg")
            with open(cfg_path, "w", encoding="utf-8") as f:
                f.write("MATERIAL:" + ",".join(sel_mats) + "\n")
                f.write("PLANT:" + ",".join(sel_plts) + "\n")
        except Exception:
            pass

    def _load_selection(self):
        try:
            cfg_path = os.path.join(os.getcwd(), ".rpa_selection.cfg")
            if not os.path.exists(cfg_path):
                return
            saved_mats, saved_plts = set(), set()
            with open(cfg_path, "r", encoding="utf-8") as f:
                for line in f:
                    if line.startswith("MATERIAL:"):
                        saved_mats = set(line.strip().replace("MATERIAL:", "").split(","))
                    elif line.startswith("PLANT:"):
                        saved_plts = set(line.strip().replace("PLANT:", "").split(","))
            for x in self.mat_vars:
                x["var"].set(x["val"] in saved_mats)
            for x in self.plant_vars:
                x["var"].set(x["val"] in saved_plts)
            self._update_mat_count()
            self._update_plt_count()
        except Exception:
            pass

    # ── LOG ──────────────────────────────────────────────────

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    def _write_log(self, msg: str, level: str = "INFO"):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n", level)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _log_line(self, line: str):
        """Color a raw stdout line from the subprocess and write it."""
        level = "INFO"
        if "SUCCESS" in line or "✓" in line:
            level = "OK"
        elif "ERROR" in line or "❌" in line:
            level = "ERROR"
        elif "WARN" in line or "SKIP" in line:
            level = "WARN"
        self._write_log(line, level)

    def _open_output(self):
        path = os.path.join(os.getcwd(), "output_rpa")
        os.makedirs(path, exist_ok=True)
        try:
            os.startfile(path)
        except Exception:
            self._write_log(f"Output folder: {path}", "INFO")

    # ── STATUS TRACKING (checklist highlighting) ──────────────

    def _reset_status(self):
        for var_list in [self.mat_vars, self.plant_vars]:
            for x in var_list:
                x["chk"].config(text=x["orig"], fg=self.TEXT2, font=(FONT, 8))
        self._mat_progress = {}
        self._current_mat = None
        self._current_plt = None

    def _set_status(self, val, status, var_list):
        icons = {"running": "⚙", "ok": "✓", "skip": "↷", "error": "✗"}
        colors = {"running": self.WARNING, "ok": self.SUCCESS,
                  "skip": self.TEXT3, "error": self.DANGER}
        for x in var_list:
            if x["orig"] == val or x["val"] == val:
                icon = icons.get(status, "")
                x["chk"].config(text=f"{icon}  {x['orig']}", fg=colors.get(status, self.TEXT2),
                                font=(FONT, 8, "bold"))
                break

    def _reset_plant_status(self):
        for x in self.plant_vars:
            x["chk"].config(text=x["orig"], fg=self.TEXT2, font=(FONT, 8))

    def _plant_done(self, mat, result):
        if mat not in self._mat_progress:
            return
        prog = self._mat_progress[mat]
        if result == "ok":
            prog["done"] += 1
            self._ok_count += 1
        else:
            prog["error"] += 1
            self._err_count += 1
        self._counter_lbl.config(text=f"{self._ok_count + self._err_count} / {self._total_count}   "
                                       f"✓ {self._ok_count}   ✗ {self._err_count}")
        if prog["done"] + prog["error"] >= prog["total"]:
            self._set_status(mat, "ok" if prog["error"] == 0 else "error", self.mat_vars)

    # ── RUN / STOP ────────────────────────────────────────────

    def _on_run(self):
        mats = self._get_selected(self.mat_vars)
        plts = self._get_selected(self.plant_vars)

        if not mats:
            messagebox.showwarning("Missing Input", "Please select at least one material.")
            return
        if not plts:
            messagebox.showwarning("Missing Input", "Please select at least one plant.")
            return

        script_path = os.path.join(os.getcwd(), SCRIPT_NAME)
        if not os.path.exists(script_path):
            messagebox.showerror("Error", f"{SCRIPT_NAME} not found in {os.getcwd()}.")
            return

        if not messagebox.askyesno(
            "SAP Ready?",
            "Before running:\n\n"
            "  1. SAP GUI is open and you are logged in\n"
            "  2. You will NOT touch the mouse or keyboard\n\n"
            f"Process {len(mats)} material(s) × {len(plts)} plant(s) "
            f"= {len(mats) * len(plts)} combinations.\n\n"
            "Is SAP ready to go?\n\n"
        ):
            return

        self._save_selection()

        # Write the temp input file the underlying script reads
        tmp_path = os.path.join(os.getcwd(), TEMP_INPUT_NAME)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["MATNR", "PLANT"])
        for m in mats:
            for p in plts:
                ws.append([m, p])
        wb.save(tmp_path)

        # Patch dry_run + input_file in the script's CONFIG dict
        try:
            with open(script_path, "r", encoding="utf-8") as f:
                content = f.read()
            val = "True" if self.dry_var.get() else "False"
            content = re.sub(r'"dry_run"\s*:\s*(True|False)', f'"dry_run": {val}', content)
            content = re.sub(r'"input_file"\s*:\s*"[^"]*"',
                             f'"input_file": "{TEMP_INPUT_NAME}"', content)
            with open(script_path, "w", encoding="utf-8") as f:
                f.write(content)
        except Exception as e:
            self._write_log(f"Could not patch script config: {e}", "WARN")

        total = len(mats) * len(plts)
        self._ok_count = 0
        self._err_count = 0
        self._total_count = total
        self._start_time = time.time()
        self._reset_status()
        for m in mats:
            self._mat_progress[m] = {"done": 0, "error": 0, "total": len(plts)}

        self.run_btn.configure(state="disabled", text="⏳  Running...", bg="#555555")
        self.stop_btn.configure(state="normal")
        self._counter_lbl.config(text=f"0 / {total}   ✓ 0   ✗ 0")

        mode = " [DRY RUN]" if self.dry_var.get() else ""
        self._write_log("━" * 45, "INFO")
        self._write_log(f"Extend Material started{mode}", "OK")
        self._write_log(f"Materials : {', '.join(mats)}", "INFO")
        self._write_log(f"Plants    : {', '.join(plts)}", "INFO")
        self._write_log(f"Total     : {total} combinations", "INFO")
        self._write_log("━" * 45, "INFO")

        threading.Thread(target=self._run_subprocess, args=(mats, plts), daemon=True).start()

    def _run_subprocess(self, sel_mats, sel_plts):
        try:
            env = os.environ.copy()
            env["PYTHONIOENCODING"] = "utf-8"
            self.process = subprocess.Popen(
                [_python_executable(), SCRIPT_NAME],
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT,
                text=True, bufsize=1, encoding="utf-8", errors="replace",
                env=env, cwd=os.getcwd(),
            )

            for line in self.process.stdout:
                line = line.rstrip()
                if not line:
                    continue
                line = re.sub(r'\x1b\[[0-9;]*[mGKHF]', '', line)
                self.parent.after(0, self._log_line, line)

                found_mat = next((m for m in sel_mats if m in line), None)
                found_plt = next((p for p in sel_plts if p in line), None)

                if found_mat:
                    if self._current_mat and self._current_mat != found_mat:
                        self.parent.after(0, self._reset_plant_status)
                    self._current_mat = found_mat

                if found_plt:
                    self._current_plt = found_plt

                eff_mat = found_mat or self._current_mat
                eff_plt = found_plt or self._current_plt

                is_skip = "[SKIP]" in line or "SKIP" in line

                if "SUCCESS" in line and not is_skip and eff_mat and eff_plt:
                    self.parent.after(0, self._set_status, eff_plt, "ok", self.plant_vars)
                    self.parent.after(0, self._plant_done, eff_mat, "ok")
                elif is_skip and eff_mat and eff_plt:
                    self.parent.after(0, self._set_status, eff_plt, "skip", self.plant_vars)
                    self.parent.after(0, self._plant_done, eff_mat, "ok")  # skip counted as handled, not error
                elif ("ERROR" in line or "❌" in line) and "Total" not in line and not is_skip:
                    if eff_mat and eff_plt:
                        self.parent.after(0, self._set_status, eff_plt, "error", self.plant_vars)
                        self.parent.after(0, self._plant_done, eff_mat, "error")

            self.process.wait()
            self.parent.after(0, self._on_done, self.process.returncode)

        except Exception as e:
            self.parent.after(0, self._on_error, str(e))

    def _on_done(self, rc):
        self.run_btn.configure(state="normal", text="▶   Run MMSC", bg=self.BG_DARK)
        self.stop_btn.configure(state="disabled")

        tmp_path = os.path.join(os.getcwd(), TEMP_INPUT_NAME)
        if os.path.exists(tmp_path):
            try:
                os.remove(tmp_path)
            except Exception:
                pass

        self._write_log("━" * 45, "INFO")
        if rc == 0:
            self._write_log(f"Done — ✓ {self._ok_count}  ✗ {self._err_count}  "
                            f"(report saved to output_rpa/)", "OK")
        else:
            self._write_log(f"Process exited with code {rc}", "ERROR")
        self._write_log("━" * 45, "INFO")

    def _on_error(self, msg):
        self.run_btn.configure(state="normal", text="▶   Run MMSC", bg=self.BG_DARK)
        self.stop_btn.configure(state="disabled")
        self._write_log(f"FATAL: {msg}", "ERROR")

    def _on_stop(self):
        if self.process:
            try:
                self.process.terminate()
            except Exception:
                pass
        self.run_btn.configure(state="normal", text="▶   Run MMSC", bg=self.BG_DARK)
        self.stop_btn.configure(state="disabled")
        self._write_log("─" * 45, "WARN")
        self._write_log("Stopped by user.", "WARN")
        self._write_log("─" * 45, "WARN")


# ─────────────────────────────────────────────
# STANDALONE ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Extend Material — MMSC")
    root.geometry("1100x700")
    root.configure(bg="#F5F5F5")
    frame = tk.Frame(root, bg="#F5F5F5")
    frame.pack(fill="both", expand=True)
    ExtendMaterialGui(frame, back_callback=root.destroy)
    root.mainloop()
