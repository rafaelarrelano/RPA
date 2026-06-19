"""
SAP RPA - Auto Extend Material MMSC
Metode: Keyboard only — delay optimal
"""

import pyautogui
import pyperclip
import time
import os
import sys
from datetime import datetime
from loguru import logger
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

try:
    import win32gui
    import win32con
    HAS_WIN32 = True
except ImportError:
    HAS_WIN32 = False

try:
    import uiautomation as auto
    HAS_UIA = True
except ImportError:
    HAS_UIA = False

CONFIG = {
    "storage_locs": ["WH02", "WT01"],
    "input_file": "_selected_temp.xlsx",
    "output_dir":   "output_rpa",
    "dry_run": False,
}

os.makedirs(CONFIG["output_dir"], exist_ok=True)
RUN_TS   = datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_FILE = os.path.join(CONFIG["output_dir"], f"rpa_mmsc_{RUN_TS}.log")

logger.remove()
logger.add(sys.stdout,
           format="<green>{time:HH:mm:ss}</green> | <level>{level:<8}</level> | {message}",
           colorize=True)
logger.add(LOG_FILE,
           format="{time:YYYY-MM-DD HH:mm:ss} | {level:<8} | {message}",
           encoding="utf-8")

pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.05


def _copy_terverifikasi(nilai, max_coba=4):
    """
    Set clipboard ke `nilai`, lalu BACA ULANG untuk memastikan benar2
    ter-set sebelum dipakai (Ctrl+V). Ini mencegah race condition di mana
    Ctrl+V dieksekusi sebelum OS benar2 selesai update clipboard, yang
    bisa menyebabkan data lama (nyasar) ke-paste.
    """
    nilai = str(nilai)
    for _ in range(max_coba):
        pyperclip.copy(nilai)
        time.sleep(0.12)
        if pyperclip.paste() == nilai:
            return True
        time.sleep(0.1)
    logger.warning(f"  Clipboard tidak sinkron setelah {max_coba}x coba (target: '{nilai}', isi: '{pyperclip.paste()}')")
    return False


def hapus_ketik(nilai):
    pyautogui.press("home")
    time.sleep(0.2)
    pyautogui.hotkey("ctrl", "shift", "right")
    time.sleep(0.2)
    pyautogui.press("backspace")
    time.sleep(0.2)
    _copy_terverifikasi(nilai)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.3)

def tempel(nilai):
    _copy_terverifikasi(nilai)
    pyautogui.hotkey("ctrl", "v")
    time.sleep(0.3)

def baca_pesan_status():
    """
    Baca teks pesan/status SAP (mis. "Storage location already exists...")
    via Microsoft UI Automation (uiautomation) — lebih kuat untuk elemen
    custom-rendered seperti status bar SAP, dibanding GetWindowText biasa.
    Fallback ke win32gui child-window text kalau uiautomation tidak ada
    atau gagal. Tetap tanpa OCR/pytesseract.

    Kalau ini belum terpasang, install dulu di komputer yang menjalankan
    robot (BUKAN di sini):  pip install uiautomation
    """
    texts = []

    if HAS_UIA:
        try:
            win = auto.GetForegroundControl()

            def walk(ctrl, depth=0):
                if depth > 14:
                    return
                try:
                    name = ctrl.Name
                    if name:
                        ctype = ""
                        try:
                            ctype = ctrl.ControlTypeName
                        except Exception:
                            pass
                        texts.append(f"[{ctype}:Name]{name}")
                except Exception:
                    pass
                try:
                    legacy = ctrl.GetLegacyIAccessiblePattern()
                    if legacy:
                        val = legacy.Value
                        if val:
                            texts.append(f"[Legacy:Value]{val}")
                        desc = legacy.Description
                        if desc:
                            texts.append(f"[Legacy:Desc]{desc}")
                except Exception:
                    pass
                try:
                    children = ctrl.GetChildren()
                except Exception:
                    children = []
                for c in children:
                    walk(c, depth + 1)

            walk(win)
        except Exception as e:
            logger.debug(f"  Gagal baca status via UIA: {e}")
    else:
        logger.debug("  uiautomation belum terpasang -- pip install uiautomation untuk hasil lebih akurat")

    # Tambahan/fallback: cara lama via win32gui (kadang tetap menangkap sesuatu)
    if HAS_WIN32:
        try:
            hwnd = win32gui.GetForegroundWindow()

            def enum_child(child, _):
                try:
                    t = win32gui.GetWindowText(child)
                    if t:
                        texts.append(t)
                except Exception:
                    pass
                return True

            win32gui.EnumChildWindows(hwnd, enum_child, None)
        except Exception:
            pass

    return " | ".join(texts).upper()


def cek_sloc_sudah_ada(max_check=12):
    """
    Cek SLoc yang sudah ada di tabel Storage Location (Layar 2), via
    clipboard -- BUKAN OCR, tidak butuh pytesseract.

    PENTING -- perubahan dari versi sebelumnya: versi lama mulai dengan
    "Up sebanyak max_rows kali" untuk asumsi mentok ke baris 1 tabel.
    Ternyata itu TIDAK reliable (terbukti dari log: kadang cuma sampai
    baris 3, bukan baris 1 sungguhan).

    Jadi sekarang baca dimulai dari posisi kursor SAAT INI (yang pasti
    benar -- ini baris kosong tempat kursor otomatis mendarat begitu
    pindah dari Layar 1 -> Layar 2), lalu MUNDUR ke atas SATU PER SATU
    sambil dihitung presisi jumlah langkahnya (variabel `naik`). Berhenti
    kalau:
      a) ketemu baris kosong (berarti sudah lewat baris paling atas), atau
      b) baris yang terbaca SAMA dengan baris sebelumnya 2x berturut-turut
         (tanda kursor mentok di baris 1, "Up" sudah tidak bergerak lagi)

    Karena jumlah langkah naik dihitung presisi, kursor bisa dikembalikan
    ke posisi awal dengan jumlah "Down" yang SAMA PERSIS -- tidak perlu
    menebak/menghitung ulang berdasarkan isi data (itu yang kemarin bikin
    salah posisi).
    """
    naik = 0
    existing_reversed = []
    try:
        prev_val = None
        for i in range(max_check):
            pyautogui.press("up")
            time.sleep(0.1)
            naik += 1

            pyautogui.press("home")
            time.sleep(0.05)
            pyautogui.hotkey("shift", "end")
            time.sleep(0.05)
            pyperclip.copy("")
            pyautogui.hotkey("ctrl", "c")
            time.sleep(0.18)
            val = pyperclip.paste().strip().upper()
            logger.debug(f"  [DEBUG] Mundur ke-{i+1}: '{val}'")

            if not val:
                # sudah lewat baris paling atas -> berhenti, jangan dihitung
                break

            if val == prev_val:
                # baris sama dgn sebelumnya -> kursor mentok, "Up" tidak
                # bergerak lagi -> ini bacaan ulang baris yang sama, stop
                break

            existing_reversed.append(val)
            prev_val = val

        # Kembalikan kursor ke posisi awal (baris kosong tempat mulai):
        # turun sejumlah "naik" -- presisi, tidak ditebak.
        for _ in range(naik):
            pyautogui.press("down")
            time.sleep(0.06)

        existing = list(reversed(existing_reversed))
        logger.debug(f"  [DEBUG] SLoc yang terbaca: {existing}")
        return existing

    except Exception as e:
        logger.warning(f"  Gagal cek SLoc via clipboard: {e}")
        return []

def cek_di_layar2():
    """
    Cek apakah SAP sudah pindah ke Layar 2 (Storage Locations List).
    Caranya: cek title window SAP aktif.
    - Layar 2 = title mengandung "List"
    - Layar 1 = title masih "Initial Screen" (artinya ada error)
    Return True jika sudah di Layar 2, False jika masih di Layar 1 (error)
    """
    if not HAS_WIN32:
        return True  # asumsikan berhasil jika win32gui tidak ada

    try:
        hwnd = win32gui.GetForegroundWindow()
        title = win32gui.GetWindowText(hwnd)
        if "List" in title or "list" in title.lower():
            return True   # berhasil masuk Layar 2
        if "Initial Screen" in title or "initial" in title.lower():
            return False  # masih di Layar 1 = ada error
        return True  # title lain, asumsikan berhasil
    except:
        return True


def focus_sap():
    """
    Cari window SAP yang sudah login dan fokus ke sana.
    Menggunakan subprocess untuk hindari konflik thread dengan UI.
    """
    if not HAS_WIN32:
        logger.warning("  win32gui tidak tersedia")
        return False

    # Kata kunci judul window SAP yang sudah login
    sap_keywords = [
        "SAP Easy Access",
        "Enter Storage Locations",
        "User menu for",
        "MMSC",
        "CSPro",
    ]

    found_hwnd  = None
    found_title = ""

    def enum_cb(hwnd, _):
        nonlocal found_hwnd, found_title
        if not win32gui.IsWindowVisible(hwnd):
            return
        if not win32gui.IsWindowEnabled(hwnd):
            return
        title = win32gui.GetWindowText(hwnd)
        if not title:
            return
        for kw in sap_keywords:
            if kw.lower() in title.lower():
                found_hwnd  = hwnd
                found_title = title
                return  # stop di yang pertama ditemukan

    win32gui.EnumWindows(enum_cb, None)

    if not found_hwnd:
        logger.error("  Window SAP tidak ditemukan!")
        logger.error("  Pastikan SAP sudah dibuka dan login terlebih dahulu.")
        return False

    logger.info(f"  SAP ditemukan: '{found_title}'")

    try:
        # Step 1: Tampilkan window SAP — maximize (full screen)
        win32gui.ShowWindow(found_hwnd, win32con.SW_SHOWMAXIMIZED)
        time.sleep(0.5)

        # Step 2: Bawa ke depan
        win32gui.BringWindowToTop(found_hwnd)
        time.sleep(0.2)
        win32gui.SetForegroundWindow(found_hwnd)
        time.sleep(0.5)

        logger.info("  Fokus SAP berhasil ✓")
        return True

    except Exception as e:
        logger.warning(f"  Gagal fokus SAP: {e}")
        return False


def split_screen():
    """
    Split layar: SAP di kiri, UI RPA di kanan.
    Menggunakan win32gui untuk resize dan reposisi kedua window.
    """
    if not HAS_WIN32:
        return

    import win32process
    import ctypes

    # Dapatkan resolusi layar
    user32 = ctypes.windll.user32
    screen_w = user32.GetSystemMetrics(0)
    screen_h = user32.GetSystemMetrics(1)

    half_w = screen_w // 2

    # Cari window SAP
    sap_keywords = [
        "SAP Easy Access", "Enter Storage Locations",
        "User menu for", "MMSC", "CSPro",
    ]
    sap_hwnd = None

    def find_sap(hwnd, _):
        nonlocal sap_hwnd
        if not win32gui.IsWindowVisible(hwnd):
            return
        title = win32gui.GetWindowText(hwnd)
        for kw in sap_keywords:
            if kw.lower() in title.lower() and sap_hwnd is None:
                sap_hwnd = hwnd

    win32gui.EnumWindows(find_sap, None)

    # Cari window UI RPA (title mengandung "SAP RPA")
    rpa_hwnd = None

    def find_rpa(hwnd, _):
        nonlocal rpa_hwnd
        if not win32gui.IsWindowVisible(hwnd):
            return
        title = win32gui.GetWindowText(hwnd)
        if "SAP RPA" in title and rpa_hwnd is None:
            rpa_hwnd = hwnd

    win32gui.EnumWindows(find_rpa, None)

    try:
        if sap_hwnd:
            # SAP di kiri: x=0, y=0, w=half, h=screen_h
            win32gui.ShowWindow(sap_hwnd, win32con.SW_RESTORE)
            win32gui.MoveWindow(sap_hwnd, 0, 0, half_w, screen_h, True)
            logger.info(f"  SAP → kiri layar ({half_w}x{screen_h})")

        if rpa_hwnd:
            # UI RPA di kanan: x=half, y=0, w=half, h=screen_h
            win32gui.ShowWindow(rpa_hwnd, win32con.SW_RESTORE)
            win32gui.MoveWindow(rpa_hwnd, half_w, 0, half_w, screen_h, True)
            logger.info(f"  UI RPA → kanan layar ({half_w}x{screen_h})")

        time.sleep(0.5)

        # Fokus kembali ke SAP agar robot bisa ketik
        if sap_hwnd:
            win32gui.SetForegroundWindow(sap_hwnd)
            time.sleep(0.3)

    except Exception as e:
        logger.warning(f"  Split screen gagal: {e}")


def buka_mmsc():
    # Fokus ke window SAP terlebih dahulu
    focus_sap()
    time.sleep(0.3)
    # Split layar SAP kiri, UI RPA kanan
    split_screen()
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "/")
    time.sleep(0.3)
    tempel("/nMMSC")
    pyautogui.press("enter")
    time.sleep(2.5)
    logger.info("  MMSC terbuka")

def read_input(filepath):
    if not os.path.exists(filepath):
        logger.error(f"File tidak ditemukan: {filepath}")
        sys.exit(1)
    wb   = openpyxl.load_workbook(filepath, data_only=True)
    ws   = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        rows.append({
            "matnr": str(row[0]).strip(),
            "plant": str(row[1]).strip().upper()
        })
    logger.info(f"Total material: {len(rows)} baris")
    return rows

def proses_material(matnr, plant):
    try:
        # LAYAR 1
        logger.debug(f"  Isi Material: {matnr}")
        hapus_ketik(matnr)

        pyautogui.press("tab")
        time.sleep(0.5)

        logger.debug(f"  Isi Plant: {plant}")
        hapus_ketik(plant)

        pyautogui.press("enter")
        time.sleep(1.5)

        # Cek apakah berhasil masuk Layar 2
        # Jika masih di Layar 1 = ada error (locked / not exist) → skip
        if not cek_di_layar2():
            logger.warning(f"  [SKIP] [{plant}] {matnr} -- Error SAP (locked/not exist), lanjut plant berikutnya")
            pyautogui.press("escape")
            time.sleep(0.5)
            return "SKIP", "Error SAP: locked atau tidak ada di plant"

        # LAYAR 2
        existing_sloc = cek_sloc_sudah_ada()
        missing_sloc  = [s for s in CONFIG["storage_locs"] if s.upper() not in existing_sloc]

        if not missing_sloc:
            logger.warning(f"  [SKIP] [{plant}] {matnr} -- WH02 & WT01 sudah ada, skip (tanpa save), lanjut ke berikutnya")
            pyautogui.press("f3")
            time.sleep(1.5)
            return "SKIP", "WH02 & WT01 sudah ada - skip tanpa save"
        else:
            logger.info(f"  Isi SLoc yang belum ada: {', '.join(missing_sloc)}")

            for idx, sloc in enumerate(missing_sloc):
                tempel(sloc)
                logger.debug(f"  SLoc baru: {sloc}")
                if idx < len(missing_sloc) - 1:
                    pyautogui.press("down")
                    time.sleep(0.4)

            pyautogui.press("tab")
            time.sleep(0.4)

            # ── Pengaman tambahan: cek apakah SAP munculkan warning
            #    "Storage location already exists..." setelah kita input.
            #    Kalau pre-check di atas (existing_sloc) salah/lolos, ini jadi
            #    jaring pengaman terakhir sebelum sempat menyimpan data salah.
            pesan_status = baca_pesan_status()
            logger.debug(f"  [DEBUG] Jumlah elemen terbaca: {pesan_status.count('|') + 1}")
            logger.debug(f"  [DEBUG] Status text: {pesan_status[:1500]}")

            sudah_ada_warning = (
                "ALREADY EXISTS" in pesan_status
                or "SUDAH ADA" in pesan_status
                or "ALREADY EXIST" in pesan_status
            )

            if sudah_ada_warning:
                logger.warning(f"  [SKIP] [{plant}] {matnr} -- Warning 'storage location already exists', back tanpa save, lanjut plant berikutnya")
                pyautogui.press("escape")
                time.sleep(0.5)
                pyautogui.press("f3")
                time.sleep(1.5)
                return "SKIP", "Storage location already exists (warning) - skip tanpa save"

            if not CONFIG["dry_run"]:
                pyautogui.press("f11")
                time.sleep(2.0)
                logger.success(f"  ✓ [{plant}] {matnr} — {', '.join(missing_sloc)} DITAMBAH")
                return "OK", f"{', '.join(missing_sloc)} ditambah"
            else:
                pyautogui.press("f3")
                time.sleep(1.5)
                return "DRY", f"dry-run tambah {', '.join(missing_sloc)}"

    except Exception as e:
        logger.error(f"  ERROR [{plant}] {matnr}: {e}")
        try:
            pyautogui.press("f3")
            time.sleep(1.5)
        except:
            pass
        return "ERROR", str(e)

def save_report(results):
    path = os.path.join(CONFIG["output_dir"], f"hasil_extend_{RUN_TS}.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Hasil RPA MMSC"
    headers = ["No", "Material", "Plant", "Status", "Keterangan", "Waktu"]
    ws.append(headers)
    thin   = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_fill = PatternFill("solid", fgColor="0070C0")
    for c in ws[1]:
        c.font      = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
        c.fill      = h_fill
        c.alignment = Alignment(horizontal="center")
        c.border    = border
    colors = {"OK": "C6EFCE", "DRY": "DDEBF7", "ERROR": "FFC7CE", "SKIP": "FFEB9C"}
    for i, r in enumerate(results, 1):
        ws.append([i, r["matnr"], r["plant"], r["status"], r["msg"], r["time"]])
        fill = PatternFill("solid", fgColor=colors.get(r["status"], "FFFFFF"))
        for c in ws[ws.max_row]:
            c.fill = fill
            c.font = Font(name="Calibri", size=10)
            c.border = border
    for col, w in zip("ABCDEF", [5, 22, 10, 10, 40, 20]):
        ws.column_dimensions[col].width = w
    wb.save(path)
    logger.success(f"Laporan: {path}")

def main():
    logger.info("=" * 60)
    logger.info("  SAP RPA — Auto Extend Material MMSC")
    logger.info(f"  Storage Loc : {' & '.join(CONFIG['storage_locs'])}")
    logger.info(f"  Input       : {CONFIG['input_file']}")
    logger.info(f"  Dry Run     : {CONFIG['dry_run']}")
    logger.info("=" * 60)
    logger.info("")
    logger.info("  PERSIAPAN:")
    logger.info("  1. SAP sudah login")
    logger.info("  2. Klik jendela SAP supaya aktif")
    logger.info("  3. JANGAN sentuh mouse/keyboard saat robot berjalan!")
    logger.info("  4. DARURAT: gerak mouse ke pojok KIRI ATAS = STOP")
    logger.info("")

    materials = read_input(CONFIG["input_file"])
    if not materials:
        logger.error("Tidak ada data. Cek material_list.xlsx")
        sys.exit(1)

    for i in range(5, 0, -1):
        logger.info(f"  Mulai dalam {i} detik...")
        time.sleep(1)

    logger.info("  ROBOT MULAI!")
    logger.info("")

    buka_mmsc()

    results = []
    ok_n = err_n = skip_n = dry_n = 0

    for idx, mat in enumerate(materials, 1):
        matnr = mat["matnr"]
        plant = mat["plant"]
        logger.info(f"[{idx:>4}/{len(materials)}]  {matnr}  →  Plant: {plant}")

        status, msg = proses_material(matnr, plant)

        if status == "OK":
            ok_n += 1
        elif status == "SKIP":
            skip_n += 1
        elif status == "DRY":
            dry_n += 1
        else:
            err_n += 1

        results.append({
            "matnr":  matnr,
            "plant":  plant,
            "status": status,
            "msg":    msg,
            "time":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        })

    logger.info("")
    logger.info("=" * 60)
    logger.info(f"  SELESAI -- OK: {ok_n} | Dry: {dry_n} | Skip: {skip_n} | Error: {err_n} | Total: {len(materials)}")
    logger.info("=" * 60)
    save_report(results)
    logger.info(f"Log: {LOG_FILE}")

if __name__ == "__main__":
    main()
