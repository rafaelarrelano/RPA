"""
rpa_gui.py
GUI untuk RPA Stock Reconciliation — UI Modern v2
Dark mode / Light mode toggle, font adaptif, UI/UX lebih baik
Jalankan: python rpa_gui.py
"""
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox
import threading
import queue
import os
from datetime import datetime

# ─────────────────────────────────────────────
# QUEUE & EVENTS
# ─────────────────────────────────────────────
log_queue   = queue.Queue()
stop_event  = threading.Event()
login_event = threading.Event()
sap_event   = threading.Event()

# ─────────────────────────────────────────────
# THEME SYSTEM — Dark & Light
# ─────────────────────────────────────────────

THEMES = {
    "dark": {
        "bg":         "#0D1117",
        "surface":    "#161B22",
        "surface2":   "#21262D",
        "surface3":   "#2D333B",
        "border":     "#30363D",
        "accent":     "#2F81F7",
        "accent2":    "#1F6FEB",
        "accent_txt": "#FFFFFF",
        "success":    "#3FB950",
        "warning":    "#D29922",
        "danger":     "#F85149",
        "text":       "#E6EDF3",
        "text2":      "#8B949E",
        "text3":      "#484F58",
        "run_bg":     "#238636",
        "run_hov":    "#2EA043",
        "stop_bg":    "#6E3535",
        "stop_hov":   "#DA3633",
        "log_sel":    "#264F78",
        "scrollbar":  "#30363D",
        "pill_bg":    "#21262D",
        "tag_bg":     "#1F6FEB",
        "tag_txt":    "#FFFFFF",
        "sep":        "#30363D",
        "hdr_bg":     "#0D1117",
        "hdr_txt":    "#2F81F7",
        "input_bg":   "#0D1117",
        "input_hl":   "#2F81F7",
        "rb_sel":     "#21262D",
    },
    "light": {
        "bg":         "#F6F8FA",
        "surface":    "#FFFFFF",
        "surface2":   "#F0F2F5",
        "surface3":   "#E4E7EB",
        "border":     "#D0D7DE",
        "accent":     "#0969DA",
        "accent2":    "#0550AE",
        "accent_txt": "#FFFFFF",
        "success":    "#1A7F37",
        "warning":    "#9A6700",
        "danger":     "#CF222E",
        "text":       "#1F2328",
        "text2":      "#57606A",
        "text3":      "#8C959F",
        "run_bg":     "#1F883D",
        "run_hov":    "#2DA44E",
        "stop_bg":    "#A40E26",
        "stop_hov":   "#CF222E",
        "log_sel":    "#BDD6F5",
        "scrollbar":  "#D0D7DE",
        "pill_bg":    "#F0F2F5",
        "tag_bg":     "#0969DA",
        "tag_txt":    "#FFFFFF",
        "sep":        "#D0D7DE",
        "hdr_bg":     "#FFFFFF",
        "hdr_txt":    "#0969DA",
        "input_bg":   "#FFFFFF",
        "input_hl":   "#0969DA",
        "rb_sel":     "#F0F2F5",
    },
}

# Font sizes — adaptive
FS = {
    "xs":    9,
    "sm":    10,
    "base":  11,
    "md":    12,
    "lg":    13,
    "xl":    14,
    "xxl":  16,
    "head": 15,
}

FONT = "Segoe UI"  # modern, readable


def C(key):
    """Get current theme color."""
    return _current_theme[key]


# ─────────────────────────────────────────────
# THEME PREFERENCE — persisted to file
# ─────────────────────────────────────────────

import json as _json

_PREF_FILE = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "config", "ui_prefs.json"
)


def _load_theme_pref() -> str:
    """Read saved theme name from file. Default: dark."""
    try:
        with open(_PREF_FILE, "r", encoding="utf-8") as f:
            return _json.load(f).get("theme", "dark")
    except Exception:
        return "dark"


def _save_theme_pref(name: str):
    """Write current theme name to file."""
    try:
        os.makedirs(os.path.dirname(_PREF_FILE), exist_ok=True)
        data = {}
        try:
            with open(_PREF_FILE, "r", encoding="utf-8") as f:
                data = _json.load(f)
        except Exception:
            pass
        data["theme"] = name
        with open(_PREF_FILE, "w", encoding="utf-8") as f:
            _json.dump(data, f, indent=2)
    except Exception:
        pass


_theme_name    = _load_theme_pref()
_current_theme = THEMES.get(_theme_name, THEMES["dark"])


def toggle_theme(root, app):
    """Switch between dark and light mode, save preference, rebuild widgets."""
    global _current_theme, _theme_name
    _theme_name    = "light" if _theme_name == "dark" else "dark"
    _current_theme = THEMES[_theme_name]
    _save_theme_pref(_theme_name)
    app._rebuild_theme(root)


# ─────────────────────────────────────────────
# LOG
# ─────────────────────────────────────────────

def send_log(msg: str, level: str = "INFO"):
    timestamp = datetime.now().strftime("%H:%M:%S")
    log_queue.put((level, f"{timestamp}  {msg}"))


# ─────────────────────────────────────────────
# ROBOT LOGIC (unchanged)
# ─────────────────────────────────────────────

def wait_for_login(send_log):
    log_queue.put(("WAIT_LOGIN", ""))


def run_robot(plants, posting_date, email_to, email_cc, mode):
    try:
        import urllib.request
        cdp_ok = False
        for host in ["127.0.0.1", "localhost"]:
            try:
                urllib.request.urlopen(f"http://{host}:9222/json/version", timeout=2)
                cdp_ok = True
                break
            except Exception:
                continue

        if not cdp_ok:
            send_log("Membuka Chrome ke halaman login portal...", "INFO")
            import subprocess, os
            chrome_paths = [
                r"C:\Program Files\Google\Chrome\Application\chrome.exe",
                r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            ]
            chrome = None
            for cp in chrome_paths:
                if os.path.exists(cp):
                    chrome = cp
                    break
            if not chrome:
                send_log("Chrome tidak ditemukan!", "ERROR")
                log_queue.put(("DONE", ""))
                return

            import time

            # Chrome lama TIDAK ditutup — buka instance baru dengan profil
            # terpisah (--user-data-dir=C:\ChromeRPA) supaya bisa jalan
            # bersamaan tanpa konflik dengan Chrome yang sudah aktif.
            subprocess.Popen([
                chrome,
                "--remote-debugging-port=9222",
                "--remote-debugging-address=0.0.0.0",
                "--user-data-dir=C:\\ChromeRPA",
                "--no-first-run",
                "--no-default-browser-check",
                "https://portal.mayora.co.id/v2login",
            ])

            for _ in range(40):
                time.sleep(0.5)
                try:
                    urllib.request.urlopen("http://127.0.0.1:9222/json/version", timeout=1)
                    break
                except Exception:
                    continue

            send_log("Chrome terbuka — silakan LOGIN di browser!", "WARN")
            send_log("Setelah login, klik OK di dialog konfirmasi.", "WARN")

            log_queue.put(("WAIT_LOGIN", ""))

            login_event.wait()
            login_event.clear()
            send_log("Login dikonfirmasi — melanjutkan RPA...", "OK")

        from test_compare import run_full_pipeline
        send_log("━" * 50, "INFO")
        send_log(f"RPA dimulai  ·  mode={mode}  ·  {posting_date}", "OK")
        send_log(f"Email → {email_to}" + (f"  CC: {email_cc}" if email_cc else ""), "INFO")
        send_log("━" * 50, "INFO")
        send_log("FASE 1: Scan portal untuk plant dengan selisih...", "INFO")

        original_send_log = send_log
        def send_log_with_sap_warning(msg, level="INFO"):
            if "SAP akan dikontrol" in msg:
                log_queue.put(("SAP_WARNING", ""))
            elif level == "SAP_WAIT":
                log_queue.put(("SAP_WAIT", ""))
                sap_event.wait()
                sap_event.clear()
                return
            original_send_log(msg, level)

        items = run_full_pipeline(
            plants       = None if mode == "auto_scan" else plants,
            posting_date = posting_date,
            email_to     = email_to,
            email_cc     = email_cc,
            send_log     = send_log_with_sap_warning,
        )

        send_log("", "INFO")
        send_log("━" * 50, "INFO")
        send_log("SUMMARY", "INFO")
        if items:
            for plant, itms in sorted(items.items()):
                t917 = sum(1 for i in itms if i.mvt_type == "917")
                t918 = sum(1 for i in itms if i.mvt_type == "918")
                send_log(f"Plant {plant}  ·  {len(itms)} item  ·  917:{t917}  918:{t918}", "OK")
        else:
            send_log("Tidak ada selisih ditemukan.", "INFO")

        send_log("RPA selesai!", "OK")
        log_queue.put(("DONE", ""))
    except Exception as e:
        send_log(f"FATAL: {e}", "ERROR")
        log_queue.put(("DONE", ""))


# ─────────────────────────────────────────────
# HOVER BUTTON
# ─────────────────────────────────────────────

class HBtn(tk.Button):
    def __init__(self, master, n, h, **kw):
        self._n, self._h = n, h
        super().__init__(master, bg=n, activebackground=h, **kw)
        self.bind("<Enter>", lambda e: self.config(bg=self._h) if str(self["state"]) != "disabled" else None)
        self.bind("<Leave>", lambda e: self.config(bg=self._n) if str(self["state"]) != "disabled" else None)

    def recolor(self, n, h):
        self._n, self._h = n, h
        self.config(bg=n, activebackground=h)


# ─────────────────────────────────────────────
# SECTION LABEL HELPER
# ─────────────────────────────────────────────

def _make_section_label(parent, text):
    f = tk.Frame(parent, bg=C("bg"))
    f.pack(fill="x", pady=(14, 4))
    # Accent line
    tk.Frame(f, bg=C("accent"), width=3, height=16).pack(side="left", padx=(0, 8))
    tk.Label(
        f, text=text.upper(),
        fg=C("accent"), bg=C("bg"),
        font=(FONT, FS["xs"], "bold"),
    ).pack(side="left", anchor="w")
    return f


def _make_card(parent, padx=12, pady=10):
    outer = tk.Frame(parent, bg=C("border"), bd=0)
    outer.pack(fill="x", pady=(0, 6))
    inner = tk.Frame(outer, bg=C("surface"), padx=padx, pady=pady)
    inner.pack(fill="both", padx=1, pady=1)
    return inner


# ─────────────────────────────────────────────
# MAIN GUI CLASS
# ─────────────────────────────────────────────

class RpaGui:
    def __init__(self, root):
        self.root   = root
        self.root.title("RPA Stock Reconciliation")
        self.root.geometry("1120x800")
        self.root.minsize(900, 660)
        self.root.configure(bg=C("bg"))
        self._running = False
        self._tooltip = None
        self._plant_popup_visible = False
        self._alert_total = 0
        self._email_to_tags: list = []
        self._email_cc_tags: list = []
        self._plant_vars: dict = {}
        # Email lists live here (NOT in _build_left) so theme rebuild never resets them
        self._email_to_list: list = []
        self._email_cc_list: list = []
        self._build()
        self._load_email_defaults()
        self._poll()

    # ── THEME REBUILD ─────────────────────────────────────────

    def _rebuild_theme(self, root):
        """Destroy all widgets and rebuild with new theme colors.
        State (email lists, plant selections, form values) is preserved."""

        # ── Save state before destroy ─────────────────────────
        # Commit any pending email entry text before rebuild
        for chip_attr in ("_email_to_chips", "_email_cc_chips"):
            chips = getattr(self, chip_attr, None)
            if chips and chips[2]:          # chips[2] = commit_fn
                try:
                    chips[2]()
                except Exception:
                    pass

        # Email lists are instance-level (not recreated by _build), no need to save/restore
        saved_plant_sel  = {p: v.get() for p, v in
                            getattr(self, "_plant_vars", {}).items()}
        saved_date       = getattr(self, "date_var",
                                   tk.StringVar()).get()
        saved_plant_map  = getattr(self, "plant_map_var",
                                   tk.StringVar()).get()
        saved_u2c_path   = getattr(self, "u2c_path_var",
                                   tk.StringVar()).get()
        saved_portal     = getattr(self, "_portal_var",
                                   tk.StringVar()).get()
        saved_mode       = getattr(self, "mode_var",
                                   tk.StringVar()).get()

        # ── Destroy all widgets ───────────────────────────────
        for w in root.winfo_children():
            w.destroy()

        # ── Reset transient state (not persistent data) ───────
        self._running              = False
        self._tooltip              = None
        self._plant_popup_visible  = False
        self._alert_total          = 0
        self._email_to_tags        = []
        self._email_cc_tags        = []
        self._plant_vars           = {}
        # Clear stale UI references (will be recreated by _build)
        self._email_to_chips       = None
        self._email_cc_chips       = None
        self._rebuild_plant_grid   = None

        # ── Rebuild UI ────────────────────────────────────────
        root.configure(bg=C("bg"))
        self._build()

        # ── Restore saved state ───────────────────────────────
        # Form values
        if saved_date and hasattr(self, "date_var"):
            self.date_var.set(saved_date)
        if saved_plant_map and hasattr(self, "plant_map_var"):
            self.plant_map_var.set(saved_plant_map)
        if saved_u2c_path and hasattr(self, "u2c_path_var"):
            self.u2c_path_var.set(saved_u2c_path)
        if saved_portal and hasattr(self, "_portal_var"):
            self._portal_var.set(saved_portal)
            self._on_portal_change()
        if saved_mode and hasattr(self, "mode_var"):
            self.mode_var.set(saved_mode)

        # Plant selections — restore checked state
        for p, was_checked in saved_plant_sel.items():
            if p in self._plant_vars:
                self._plant_vars[p].set(was_checked)
        self._update_plant_summary()

        # Email chips — lists are already intact (owned by __init__)
        # Just need to re-render chip widgets after _build()
        for chip_attr in ("_email_to_chips", "_email_cc_chips"):
            chips = getattr(self, chip_attr, None)
            if chips and chips[3]:          # chips[3] = refresh_fn
                try:
                    chips[3]()
                except Exception:
                    pass

    # ── UI BUILD ─────────────────────────────────────────────

    def _build(self):
        from config import Config
        self._build_topbar()
        tk.Frame(self.root, bg=C("sep"), height=1).pack(fill="x")
        # Bottom bar MUST be packed before body so it is never pushed off-screen
        tk.Frame(self.root, bg=C("sep"), height=1).pack(fill="x", side="bottom")
        self._build_bottom()
        self._build_body(Config)

    def _build_topbar(self):
        top = tk.Frame(self.root, bg=C("hdr_bg"), height=56)
        top.pack(fill="x")
        top.pack_propagate(False)

        # Logo + title
        lf = tk.Frame(top, bg=C("hdr_bg"))
        lf.pack(side="left", padx=20, pady=10)

        tk.Label(
            lf, text="◈",
            font=(FONT, FS["xl"], "bold"),
            fg=C("accent"), bg=C("hdr_bg")
        ).pack(side="left", padx=(0, 8))

        tf = tk.Frame(lf, bg=C("hdr_bg"))
        tf.pack(side="left")
        tk.Label(
            tf, text="RPA Stock Reconciliation",
            font=(FONT, FS["head"], "bold"),
            fg=C("text"), bg=C("hdr_bg")
        ).pack(anchor="w")
        tk.Label(
            tf, text="PT Mayora Indah Tbk — Automated Stock Reconciliation",
            font=(FONT, FS["xs"]),
            fg=C("text3"), bg=C("hdr_bg")
        ).pack(anchor="w")

        # Right side: theme toggle + status pill
        rf = tk.Frame(top, bg=C("hdr_bg"))
        rf.pack(side="right", padx=16, pady=10)

        # Theme toggle button
        self._theme_icon = "☀" if _theme_name == "dark" else "🌙"
        self._theme_btn = HBtn(
            rf, C("surface2"), C("surface3"),
            text=self._theme_icon,
            font=(FONT, FS["lg"]),
            fg=C("text2"), relief="flat", bd=0,
            padx=10, pady=4, cursor="hand2",
            command=lambda: toggle_theme(self.root, self)
        )
        self._theme_btn.pack(side="right", padx=(8, 0))

        # Status pill
        pill = tk.Frame(rf, bg=C("pill_bg"),
                        highlightbackground=C("border"), highlightthickness=1,
                        padx=2, pady=2)
        pill.pack(side="right")
        self._dot = tk.Label(pill, text="●", fg=C("success"),
                             bg=C("pill_bg"), font=(FONT, FS["sm"]))
        self._dot.pack(side="left", padx=(10, 4))
        self._stat = tk.Label(pill, text="Ready",
                              fg=C("text2"), bg=C("pill_bg"),
                              font=(FONT, FS["sm"], "bold"))
        self._stat.pack(side="left", padx=(0, 12))

    def _build_body(self, Config):
        body = tk.Frame(self.root, bg=C("bg"))
        body.pack(fill="both", expand=True, padx=0, pady=0)

        # PanedWindow — panel kiri bisa digeser kanan-kiri dengan sash
        self._paned = tk.PanedWindow(
            body, orient="horizontal",
            bg=C("sep"), sashwidth=6, sashrelief="flat",
            handlesize=0, opaqueresize=True,
        )
        self._paned.pack(fill="both", expand=True)

        # LEFT panel (scrollable) — di dalam paned
        left_outer = tk.Frame(self._paned, bg=C("bg"))
        self._paned.add(left_outer, minsize=240, width=340, stretch="never")

        self._left_canvas = tk.Canvas(
            left_outer, bg=C("bg"), highlightthickness=0
        )
        left_sb = tk.Scrollbar(left_outer, orient="vertical",
                               command=self._left_canvas.yview,
                               width=6)
        self._left_canvas.configure(yscrollcommand=left_sb.set)
        left_sb.pack(side="right", fill="y")
        self._left_canvas.pack(side="left", fill="both", expand=True)

        self._left_frame = tk.Frame(self._left_canvas, bg=C("bg"))
        self._left_win   = self._left_canvas.create_window(
            (0, 0), window=self._left_frame, anchor="nw"
        )
        self._left_frame.bind("<Configure>", self._on_left_configure)
        self._left_canvas.bind("<Configure>", self._on_canvas_configure)
        self._left_canvas.bind("<Enter>",
            lambda e: self._left_canvas.bind_all("<MouseWheel>", self._on_mousewheel))
        self._left_canvas.bind("<Leave>",
            lambda e: self._left_canvas.unbind_all("<MouseWheel>"))

        inner_pad = tk.Frame(self._left_frame, bg=C("bg"))
        inner_pad.pack(fill="both", padx=14)
        self._build_left(inner_pad, Config)

        # RIGHT panel — di dalam paned
        right = tk.Frame(self._paned, bg=C("bg"))
        self._paned.add(right, minsize=400, stretch="always")
        self._build_right(right)

    def _on_left_configure(self, event):
        self._left_canvas.configure(scrollregion=self._left_canvas.bbox("all"))

    def _on_canvas_configure(self, event):
        self._left_canvas.itemconfig(self._left_win, width=event.width)

    def _on_mousewheel(self, event):
        self._left_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

    def _on_portal_change(self, event=None):
        from config import Config
        label = self._portal_var.get()
        url   = self._portal_options.get(label, "")
        try:
            from test_compare import set_portal_url
            set_portal_url(url)
        except Exception:
            pass
        tmap = Config.PORTAL_TCODE_MAP.get(label, {})
        Config.ACTIVE_TCODE_SAPSTK = tmap.get("sapstk", "ZPGD_SAPSTK")
        Config.ACTIVE_TCODE_U2C    = tmap.get("u2c",    "ZPGD_U2C")
        self._portal_url_lbl.config(text=f"  {url}")
        if hasattr(self, "_tcode_info_lbl"):
            self._tcode_info_lbl.config(text=self._tcode_info_text(label))

    def _tcode_info_text(self, portal_label: str = None) -> str:
        from config import Config
        if portal_label is None:
            portal_label = self._portal_var.get() if hasattr(self, "_portal_var") else "PGDMTX"
        tmap   = Config.PORTAL_TCODE_MAP.get(portal_label, {})
        sapstk = tmap.get("sapstk", Config.ACTIVE_TCODE_SAPSTK)
        u2c    = tmap.get("u2c",    Config.ACTIVE_TCODE_U2C)
        return f"/{sapstk}  +  /{u2c}"

    def _load_plants_from_excel(self, filepath: str = None) -> list:
        from config import Config as Cfg
        if filepath is None:
            filepath = self.plant_map_var.get() if hasattr(self, 'plant_map_var') else Cfg.PLANT_MAPPING_FILE
        Cfg.PLANT_MAPPING_FILE = filepath
        try:
            import openpyxl, os
            if not os.path.exists(filepath):
                raise FileNotFoundError(f"File tidak ditemukan: {filepath}")
            wb = openpyxl.load_workbook(filepath, data_only=True)
            if "Plant_CostCenter" not in wb.sheetnames:
                raise ValueError(f"Sheet 'Plant_CostCenter' tidak ada")
            ws = wb["Plant_CostCenter"]
            plants = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if row[1]:
                    code = str(row[1]).strip()
                    if code and code not in plants:
                        plants.append(code)
            if not plants:
                raise ValueError("Tidak ada data plant")
            return sorted(plants)
        except Exception as e:
            err_msg = f"[Plant Map] Gagal baca '{filepath}': {e}"
            print(err_msg)
            try:
                self.root.after(500, lambda: self._write_log(
                    f"⚠  {err_msg}\n   → Pakai hardcode plants sebagai fallback.",
                    "WARN"
                ))
            except Exception:
                pass
            return list(Cfg.PLANTS)

    def _rebuild_plant_checklist(self, grid_frame, plants: list):
        old_states = {p: v.get() for p, v in self._plant_vars.items()}
        self._plant_vars = {}
        for plant in plants:
            var = tk.BooleanVar(value=old_states.get(plant, True))
            self._plant_vars[plant] = var
        # Use search-aware rebuild if available (after popup is built)
        if hasattr(self, "_rebuild_plant_grid"):
            q = self._plant_search_var.get() if hasattr(self, "_plant_search_var") else ""
            self._rebuild_plant_grid(q)
        else:
            # Fallback: direct grid rebuild
            for w in grid_frame.winfo_children():
                w.destroy()
            for i, plant in enumerate(plants):
                cb = tk.Checkbutton(
                    grid_frame, text=plant,
                    variable=self._plant_vars[plant],
                    font=(FONT, FS["base"]),
                    fg=C("text"), bg=C("surface2"),
                    selectcolor=C("surface3"),
                    activebackground=C("surface2"),
                    activeforeground=C("text"),
                    relief="flat", bd=0,
                    command=self._update_plant_summary,
                )
                cb.grid(row=i // 3, column=i % 3, sticky="w", padx=6, pady=2)
        self._update_plant_summary()

    # ─────────────────────────────────────────
    # LEFT PANEL
    # ─────────────────────────────────────────

    def _build_left(self, p, Config):

        def entry_row(parent, label, var, width=22, show=""):
            f = tk.Frame(parent, bg=C("surface"))
            f.pack(fill="x", pady=3)
            tk.Label(f, text=label, fg=C("text2"), bg=C("surface"),
                     font=(FONT, FS["base"]), width=14, anchor="w"
                     ).pack(side="left", padx=(0, 8))
            e = tk.Entry(f, textvariable=var, show=show, width=width,
                         bg=C("input_bg"), fg=C("text"),
                         insertbackground=C("accent"),
                         font=(FONT, FS["base"]), relief="flat", bd=0,
                         highlightthickness=2,
                         highlightbackground=C("border"),
                         highlightcolor=C("input_hl"))
            e.pack(side="left", ipady=4, padx=2)
            return e

        # ── CONFIGURATION ──────────────────────────────────
        _make_section_label(p, "Configuration")
        c1 = _make_card(p)

        self.date_var      = tk.StringVar(value=datetime.now().strftime("%d.%m.%Y"))
        self.plant_map_var = tk.StringVar(value=Config.PLANT_MAPPING_FILE)

        _initial_plants = self._load_plants_from_excel(Config.PLANT_MAPPING_FILE)
        _from_excel = sorted(_initial_plants) != sorted(list(Config.PLANTS)) or \
                      len(_initial_plants) != len(Config.PLANTS)
        self._plant_map_loaded_ok = _from_excel
        self._plant_vars = {p2: tk.BooleanVar(value=True) for p2 in _initial_plants}

        entry_row(c1, "Posting Date", self.date_var, width=14)

        # Portal EOD — custom themed dropdown (no ttk gray problem)
        from test_compare import PORTAL_EOD_URLS
        self._portal_options = PORTAL_EOD_URLS
        self._portal_var = tk.StringVar(value=list(self._portal_options.keys())[0])
        self._portal_popup = None

        prow = tk.Frame(c1, bg=C("surface"))
        prow.pack(fill="x", pady=3)
        tk.Label(prow, text="Portal EOD", fg=C("text2"), bg=C("surface"),
                 font=(FONT, FS["base"]), width=14, anchor="w"
                 ).pack(side="left", padx=(0, 8))

        # Fully themed dropdown button
        portal_wrap = tk.Frame(prow, bg=C("border"), padx=1, pady=1)
        portal_wrap.pack(side="left")
        portal_inner = tk.Frame(portal_wrap, bg=C("input_bg"))
        portal_inner.pack()

        self._portal_display = tk.Label(
            portal_inner,
            textvariable=self._portal_var,
            bg=C("input_bg"), fg=C("text"),
            font=(FONT, FS["base"], "bold"),
            width=10, anchor="w", padx=10, pady=6,
            cursor="hand2"
        )
        self._portal_display.pack(side="left")
        _arr_lbl = tk.Label(portal_inner, text="▾",
                            bg=C("input_bg"), fg=C("accent"),
                            font=(FONT, FS["sm"]), padx=6, cursor="hand2")
        _arr_lbl.pack(side="left")

        def _open_portal_dropdown(event=None):
            if self._portal_popup and self._portal_popup.winfo_exists():
                self._portal_popup.destroy()
                self._portal_popup = None
                return
            popup = tk.Toplevel(self.root)
            popup.wm_overrideredirect(True)
            popup.configure(bg=C("border"))
            self._portal_popup = popup
            x = portal_wrap.winfo_rootx()
            y = portal_wrap.winfo_rooty() + portal_wrap.winfo_height() + 2
            popup.geometry(f"+{x}+{y}")
            inner_pop = tk.Frame(popup, bg=C("surface2"))
            inner_pop.pack(padx=1, pady=1)
            for key in self._portal_options.keys():
                is_sel = key == self._portal_var.get()
                def _pick(k=key):
                    self._portal_var.set(k)
                    self._on_portal_change()
                    popup.destroy()
                    self._portal_popup = None
                item_btn = tk.Button(
                    inner_pop, text=key,
                    bg=C("accent2") if is_sel else C("surface2"),
                    fg=C("accent_txt") if is_sel else C("text"),
                    font=(FONT, FS["base"], "bold"),
                    relief="flat", bd=0, padx=20, pady=8,
                    anchor="w", width=12, cursor="hand2",
                    activebackground=C("accent"),
                    activeforeground=C("accent_txt"),
                    command=_pick
                )
                item_btn.pack(fill="x")
            popup.bind("<FocusOut>", lambda e: (popup.destroy(),
                       setattr(self, "_portal_popup", None))
                       if popup.winfo_exists() else None)
            popup.focus_set()

        self._portal_display.bind("<Button-1>", _open_portal_dropdown)
        _arr_lbl.bind("<Button-1>", _open_portal_dropdown)
        portal_inner.bind("<Button-1>", _open_portal_dropdown)

        # URL label
        self._portal_url_lbl = tk.Label(
            c1, text=f"  {list(self._portal_options.values())[0]}",
            fg=C("text3"), bg=C("surface"),
            font=(FONT, FS["xs"]), anchor="w", wraplength=280
        )
        self._portal_url_lbl.pack(anchor="w", pady=(0, 2))

        # T-code label  
        _default_portal = list(self._portal_options.keys())[0]
        tcode_f = tk.Frame(c1, bg=C("surface2"), padx=8, pady=4)
        tcode_f.pack(fill="x", pady=(2, 4))
        tk.Label(tcode_f, text="SAP:", fg=C("text3"), bg=C("surface2"),
                 font=(FONT, FS["xs"])).pack(side="left")
        self._tcode_info_lbl = tk.Label(
            tcode_f, text=self._tcode_info_text(_default_portal),
            fg=C("accent"), bg=C("surface2"),
            font=(FONT, FS["sm"], "bold"), anchor="w",
        )
        self._tcode_info_lbl.pack(side="left", padx=(4, 0))

        # Email To / CC — vertical list + entry input di bawah
        # Setiap email tampil satu baris penuh dengan tombol × di kanan
        # Box punya scrollbar jika email banyak, entry selalu terlihat di bawah box
        # NOTE: _email_to_list / _email_cc_list are owned by __init__, NOT reset here
        for lbl_txt, list_attr, chip_attr in [
            ("Email To", "_email_to_list", "_email_to_chips"),
            ("Email CC", "_email_cc_list", "_email_cc_chips"),
        ]:
            email_list = getattr(self, list_attr)
            email_var  = tk.StringVar()

            # Section label row
            lrow = tk.Frame(c1, bg=C("surface"))
            lrow.pack(fill="x", pady=(8, 2))
            tk.Label(lrow, text=lbl_txt, fg=C("text2"), bg=C("surface"),
                     font=(FONT, FS["base"], "bold"), anchor="w"
                     ).pack(side="left")

            # ── Scrollable list box ───────────────────────────
            # Layout: [list_outer border]
            #           [list_vsb RIGHT] [list_canvas LEFT fill]
            # Scrollbar di-pack DULU (side=right) supaya canvas tidak overlap
            list_outer = tk.Frame(c1, bg=C("border"), padx=1, pady=1)
            list_outer.pack(fill="x", pady=(0, 2))

            # Scrollbar custom — strip 10px lebar, warna accent agar jelas
            list_vsb = tk.Scrollbar(
                list_outer, orient="vertical", width=10,
                bg=C("accent"), troughcolor=C("surface3"),
                activebackground=C("accent2"),
                highlightthickness=0, bd=0, relief="flat",
            )
            list_vsb.pack(side="right", fill="y")   # pack DULU

            list_canvas = tk.Canvas(
                list_outer, bg=C("input_bg"),
                highlightthickness=0, height=0
            )
            list_canvas.configure(yscrollcommand=list_vsb.set)
            list_vsb.configure(command=list_canvas.yview)
            list_canvas.pack(side="left", fill="both", expand=True)

            list_frame = tk.Frame(list_canvas, bg=C("input_bg"))
            list_win   = list_canvas.create_window((0, 0), window=list_frame, anchor="nw")

            def _on_list_canvas_cfg(event, lc=list_canvas, lw=list_win):
                lc.itemconfig(lw, width=event.width)
            list_canvas.bind("<Configure>", _on_list_canvas_cfg)

            # Mousewheel scroll pada canvas maupun konten dalamnya
            def _mw_scroll(event, lc=list_canvas):
                lc.yview_scroll(int(-1 * (event.delta / 120)), "units")
            list_canvas.bind("<MouseWheel>", _mw_scroll)
            list_frame.bind("<MouseWheel>", _mw_scroll)

            # ── Resize handle — geser vertikal untuk atur tinggi email box ──
            _rh = tk.Frame(c1, bg=C("border"), height=5,
                           cursor="sb_v_double_arrow")
            _rh.pack(fill="x", pady=(0, 0))
            _drag = {"y": 0, "h": 82}

            def _rh_press(event, _d=_drag, lc=list_canvas):
                _d["y"] = event.y_root
                _d["h"] = lc.winfo_height() or 82

            def _rh_drag(event, _d=_drag, lc=list_canvas):
                delta = event.y_root - _d["y"]
                new_h = max(27, min(_d["h"] + delta, 320))
                lc.configure(height=int(new_h))

            _rh.bind("<ButtonPress-1>", _rh_press)
            _rh.bind("<B1-Motion>",     _rh_drag)
            _rh.bind("<Enter>",  lambda e, f=_rh: f.config(bg=C("accent")))
            _rh.bind("<Leave>",  lambda e, f=_rh: f.config(bg=C("border")))

            def _refresh_list(el=email_list, lf=list_frame,
                              lc=list_canvas, ca=chip_attr, ev=email_var,
                              vsb=list_vsb):
                """Rebuild vertical email list."""
                for w in lf.winfo_children():
                    w.destroy()

                for em in list(el):
                    row = tk.Frame(lf, bg=C("input_bg"))
                    row.pack(fill="x", padx=4, pady=1)

                    # Bind mousewheel ke setiap row agar scroll tetap jalan
                    row.bind("<MouseWheel>", _mw_scroll)

                    # Email label — full width, truncate bila perlu
                    disp = em if len(em) <= 30 else em[:27] + "…"
                    lbl  = tk.Label(
                        row, text=disp,
                        bg=C("input_bg"), fg=C("text"),
                        font=(FONT, FS["sm"]), anchor="w",
                        cursor="hand2"
                    )
                    lbl.pack(side="left", fill="x", expand=True, ipady=3)
                    lbl.bind("<MouseWheel>", _mw_scroll)
                    lbl.bind("<Enter>", lambda e, full=em:
                             self._show_tooltip(e.widget, full))
                    lbl.bind("<Leave>", lambda e: self._hide_tooltip())

                    # Tombol hapus
                    def _del(e=em, _el=el):
                        if e in _el:
                            _el.remove(e)
                        _refresh_list(_el, lf, lc, ca, ev, vsb)
                    del_btn = tk.Button(
                        row, text="×",
                        bg=C("input_bg"), fg=C("text3"),
                        font=(FONT, FS["md"], "bold"), relief="flat", bd=0,
                        padx=6, pady=1, cursor="hand2",
                        activebackground=C("danger"),
                        activeforeground="#FFFFFF",
                        command=_del
                    )
                    del_btn.pack(side="right")
                    del_btn.bind("<MouseWheel>", _mw_scroll)

                    # Garis pemisah tipis antar email
                    sep = tk.Frame(lf, bg=C("border"), height=1)
                    sep.pack(fill="x", padx=4)
                    sep.bind("<MouseWheel>", _mw_scroll)

                # Sesuaikan tinggi canvas: max 3 baris (~27px/baris) sebelum scroll
                lf.update_idletasks()
                req_h = lf.winfo_reqheight()
                new_h = min(req_h, 82) if el else 0
                lc.configure(height=new_h,
                             scrollregion=lc.bbox("all") or (0, 0, 0, 0))

                # Update ref: (el, entry_widget, commit_fn, refresh_fn)
                chips = getattr(self, ca, None)
                entry_w  = chips[1] if chips else None
                commit_f = chips[2] if chips else None
                setattr(self, ca, (el, entry_w, commit_f,
                                   lambda: _refresh_list(el, lf, lc, ca, ev, vsb)))

            # ── Entry input + tombol Tambah ───────────────────
            entry_row2 = tk.Frame(c1, bg=C("border"), padx=1, pady=1)
            entry_row2.pack(fill="x", pady=(0, 2))
            entry_inner = tk.Frame(entry_row2, bg=C("input_bg"))
            entry_inner.pack(fill="x")

            chip_entry = tk.Entry(
                entry_inner, textvariable=email_var,
                bg=C("input_bg"), fg=C("text"),
                insertbackground=C("accent"),
                font=(FONT, FS["base"]), relief="flat", bd=0,
                highlightthickness=0,
            )
            chip_entry.pack(side="left", fill="x", expand=True,
                            padx=(8, 4), pady=5, ipady=3)

            def _commit(event=None, el=email_list, ev=email_var,
                        lf=list_frame, lc=list_canvas, ca=chip_attr):
                raw   = ev.get().strip()
                parts = [p2.strip().rstrip(",;") for p2 in
                         raw.replace(";", ",").split(",")]
                added = False
                for part in parts:
                    if part and part not in el:
                        el.append(part)
                        added = True
                if added:
                    ev.set("")
                    _refresh_list(el, lf, lc, ca, ev)
                return "break"

            def _on_key(event, el=email_list, ev=email_var,
                        lf=list_frame, lc=list_canvas, ca=chip_attr):
                if event.keysym == "BackSpace" and not ev.get() and el:
                    el.pop()
                    _refresh_list(el, lf, lc, ca, ev)
                    return "break"
                if event.keysym in ("comma", "semicolon", "Return", "Tab"):
                    _commit(el=el, ev=ev, lf=lf, lc=lc, ca=ca)
                    return "break"

            chip_entry.bind("<Return>",
                            lambda e, el=email_list, ev=email_var,
                            lf=list_frame, lc=list_canvas, ca=chip_attr:
                            _commit(e, el, ev, lf, lc, ca))
            chip_entry.bind("<Tab>",
                            lambda e, el=email_list, ev=email_var,
                            lf=list_frame, lc=list_canvas, ca=chip_attr:
                            _commit(e, el, ev, lf, lc, ca))
            chip_entry.bind("<KeyPress>",
                            lambda e, el=email_list, ev=email_var,
                            lf=list_frame, lc=list_canvas, ca=chip_attr:
                            _on_key(e, el, ev, lf, lc, ca))
            chip_entry.bind("<FocusOut>",
                            lambda e, el=email_list, ev=email_var,
                            lf=list_frame, lc=list_canvas, ca=chip_attr:
                            _commit(e, el, ev, lf, lc, ca) if ev.get().strip() else None)

            add_btn = tk.Button(
                entry_inner, text="+ Tambah",
                bg=C("accent2"), fg=C("accent_txt"),
                font=(FONT, FS["sm"], "bold"), relief="flat", bd=0,
                padx=10, pady=4, cursor="hand2",
                activebackground=C("accent"),
                activeforeground=C("accent_txt"),
                command=lambda el=email_list, ev=email_var,
                lf=list_frame, lc=list_canvas, ca=chip_attr:
                _commit(el=el, ev=ev, lf=lf, lc=lc, ca=ca)
            )
            add_btn.pack(side="right", padx=(0, 6), pady=4)

            # Store ref: (el, entry, commit, refresh)
            setattr(self, chip_attr, (email_list, chip_entry, _commit,
                                      lambda el=email_list, lf=list_frame,
                                      lc=list_canvas, ca=chip_attr, ev=email_var,
                                      vsb=list_vsb:
                                      _refresh_list(el, lf, lc, ca, ev, vsb)))

            # Initial render
            _refresh_list(email_list, list_frame, list_canvas, chip_attr, email_var, list_vsb)

        tk.Label(c1, text="  Ketik email → Enter / koma / tombol Tambah",
                 fg=C("text3"), bg=C("surface"),
                 font=(FONT, FS["xs"])).pack(anchor="w", pady=(2, 6))

        # Plants
        prow2 = tk.Frame(c1, bg=C("surface"))
        prow2.pack(fill="x", pady=(4, 0))
        tk.Label(prow2, text="Plants", fg=C("text2"), bg=C("surface"),
                 font=(FONT, FS["base"]), width=14, anchor="w").pack(side="left")

        self._plant_summary = tk.StringVar()
        self._update_plant_summary()

        plant_btn = tk.Button(
            prow2, textvariable=self._plant_summary,
            bg=C("input_bg"), fg=C("text"),
            font=(FONT, FS["base"]), relief="flat", bd=0,
            highlightthickness=2,
            highlightbackground=C("border"),
            highlightcolor=C("input_hl"),
            anchor="w", cursor="hand2", width=18,
            activebackground=C("surface2"),
            activeforeground=C("text"),
            command=lambda: self._toggle_plant_popup(plant_btn)
        )
        plant_btn.pack(side="left", padx=(0, 4), ipady=4)
        tk.Label(prow2, text="▾", fg=C("text3"), bg=C("surface"),
                 font=(FONT, FS["sm"])).pack(side="left")

        # Plant popup with search
        self._plant_popup = tk.Frame(
            c1, bg=C("surface2"),
            highlightbackground=C("border"), highlightthickness=1
        )
        self._plant_popup_visible = False

        # Header row: count label + Semua + Hapus buttons
        ph = tk.Frame(self._plant_popup, bg=C("surface2"))
        ph.pack(fill="x", padx=8, pady=(8, 4))
        self._plant_count_lbl = tk.Label(
            ph, text=f"{len(self._plant_vars)} plant tersedia",
            fg=C("text3"), bg=C("surface2"),
            font=(FONT, FS["sm"])
        )
        self._plant_count_lbl.pack(side="left")
        HBtn(ph, C("surface2"), C("surface3"),
             text="Semua", font=(FONT, FS["sm"]),
             fg=C("accent"), relief="flat", bd=0,
             cursor="hand2",
             command=self._plant_select_all).pack(side="right", padx=(4, 0))
        HBtn(ph, C("surface2"), C("surface3"),
             text="Hapus", font=(FONT, FS["sm"]),
             fg=C("text3"), relief="flat", bd=0,
             cursor="hand2",
             command=self._plant_clear_all).pack(side="right")

        # Search box
        search_row = tk.Frame(self._plant_popup, bg=C("surface2"))
        search_row.pack(fill="x", padx=8, pady=(2, 4))

        search_icon = tk.Label(search_row, text="🔍",
                               bg=C("surface2"), fg=C("text3"),
                               font=(FONT, FS["sm"]))
        search_icon.pack(side="left", padx=(0, 4))

        self._plant_search_var = tk.StringVar()
        search_entry = tk.Entry(
            search_row,
            textvariable=self._plant_search_var,
            bg=C("input_bg"), fg=C("text"),
            insertbackground=C("accent"),
            font=(FONT, FS["base"]), relief="flat", bd=0,
            highlightthickness=2,
            highlightbackground=C("border"),
            highlightcolor=C("input_hl"),
            width=20,
        )
        search_entry.pack(side="left", fill="x", expand=True, ipady=4)

        # Clear search button
        def _clear_search():
            self._plant_search_var.set("")
            search_entry.focus_set()
        HBtn(search_row, C("surface2"), C("surface3"),
             text="✕", font=(FONT, FS["sm"]),
             fg=C("text3"), relief="flat", bd=0,
             padx=6, pady=2, cursor="hand2",
             command=_clear_search).pack(side="left", padx=(4, 0))

        tk.Frame(self._plant_popup, bg=C("border"), height=1).pack(fill="x")

        # Scrollable grid area
        grid_canvas = tk.Canvas(
            self._plant_popup, bg=C("surface2"),
            highlightthickness=0, height=160
        )
        grid_vsb = tk.Scrollbar(self._plant_popup, orient="vertical",
                                command=grid_canvas.yview, width=6)
        grid_canvas.configure(yscrollcommand=grid_vsb.set)
        grid_vsb.pack(side="right", fill="y")
        grid_canvas.pack(fill="both", expand=True, padx=0)

        self._plant_grid = tk.Frame(grid_canvas, bg=C("surface2"))
        _plant_grid_win = grid_canvas.create_window(
            (0, 0), window=self._plant_grid, anchor="nw"
        )

        def _on_grid_configure(event):
            grid_canvas.configure(scrollregion=grid_canvas.bbox("all"))

        def _on_grid_canvas_configure(event):
            grid_canvas.itemconfig(_plant_grid_win, width=event.width)

        self._plant_grid.bind("<Configure>", _on_grid_configure)
        grid_canvas.bind("<Configure>", _on_grid_canvas_configure)

        # Mousewheel scroll on plant list
        def _grid_scroll(event):
            grid_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        grid_canvas.bind("<MouseWheel>", _grid_scroll)
        self._plant_grid.bind("<MouseWheel>", _grid_scroll)

        def _rebuild_plant_grid(filter_text=""):
            """Rebuild checkboxes filtered by search text."""
            for w in self._plant_grid.winfo_children():
                w.destroy()
            ft = filter_text.strip().lower()
            visible = [p2 for p2 in self._plant_vars.keys()
                       if ft == "" or ft in p2.lower()]
            for i, plant2 in enumerate(visible):
                cb = tk.Checkbutton(
                    self._plant_grid, text=plant2,
                    variable=self._plant_vars[plant2],
                    font=(FONT, FS["base"]),
                    fg=C("text"), bg=C("surface2"),
                    selectcolor=C("surface3"),
                    activebackground=C("surface2"),
                    activeforeground=C("text"),
                    relief="flat", bd=0,
                    command=self._update_plant_summary,
                )
                cb.grid(row=i // 3, column=i % 3, sticky="w", padx=6, pady=2)
            # Update count label
            sel = sum(1 for v in self._plant_vars.values() if v.get())
            total = len(self._plant_vars)
            shown = len(visible)
            if ft:
                self._plant_count_lbl.config(
                    text=f"{shown} cocok dari {total} plant  ({sel} dipilih)"
                )
            else:
                self._plant_count_lbl.config(
                    text=f"{total} plant tersedia  ({sel} dipilih)"
                )

        # Store rebuild fn for external use (e.g. after loading new plant map)
        self._rebuild_plant_grid = _rebuild_plant_grid

        def _on_search_change(*args):
            _rebuild_plant_grid(self._plant_search_var.get())

        self._plant_search_var.trace_add("write", _on_search_change)

        # Initial render
        _rebuild_plant_grid()

        tk.Label(c1, text="", bg=C("surface")).pack(pady=2)

        # Plant mapping file
        pf = tk.Frame(c1, bg=C("surface"))
        pf.pack(fill="x", pady=(0, 6))
        tk.Label(pf, text="Plant Map", fg=C("text2"), bg=C("surface"),
                 font=(FONT, FS["base"]), width=14, anchor="w").pack(side="left")
        tk.Entry(pf, textvariable=self.plant_map_var,
                 bg=C("input_bg"), fg=C("text"),
                 insertbackground=C("accent"),
                 font=(FONT, FS["sm"]), relief="flat", bd=0,
                 highlightthickness=2,
                 highlightbackground=C("border"),
                 highlightcolor=C("input_hl"),
                 width=16).pack(side="left", padx=(0, 6), ipady=3)
        HBtn(pf, C("surface2"), C("surface3"),
             text="📁", font=(FONT, FS["md"]),
             fg=C("text2"), relief="flat", bd=0,
             padx=6, pady=3, cursor="hand2",
             command=self._browse_plant_map).pack(side="left")

        _status_text  = "  ✔ Plant dibaca dari Excel" if self._plant_map_loaded_ok \
                        else "  ⚠ File tidak ditemukan — pilih via 📁"
        _status_color = C("success") if self._plant_map_loaded_ok else C("warning")
        self._plant_map_status_lbl = tk.Label(
            c1, text=_status_text,
            fg=_status_color, bg=C("surface"),
            font=(FONT, FS["sm"]), anchor="w",
        )
        self._plant_map_status_lbl.pack(anchor="w", pady=(0, 6))

        # ── U2C FILE PATH ───────────────────────────────────
        _make_section_label(p, "U2C File Path")
        c_u2c = _make_card(p)

        try:
            from u2c_upload import get_u2c_filepath
            _default_u2c = get_u2c_filepath()
        except Exception:
            _default_u2c = r"C:\Users\User\Documents\PGD\EOD\U2C.txt"

        self.u2c_path_var = tk.StringVar(value=_default_u2c)

        uf = tk.Frame(c_u2c, bg=C("surface"))
        uf.pack(fill="x", pady=(0, 6))
        tk.Label(uf, text="File Path", fg=C("text2"), bg=C("surface"),
                 font=(FONT, FS["base"]), width=10, anchor="w").pack(side="left")
        tk.Entry(uf, textvariable=self.u2c_path_var,
                 bg=C("input_bg"), fg=C("text"),
                 insertbackground=C("accent"),
                 font=(FONT, FS["sm"]), relief="flat", bd=0,
                 highlightthickness=2,
                 highlightbackground=C("border"),
                 highlightcolor=C("input_hl"),
                 width=18).pack(side="left", padx=(4, 6), ipady=3)
        HBtn(uf, C("surface2"), C("surface3"),
             text="📁", font=(FONT, FS["md"]),
             fg=C("text2"), relief="flat", bd=0,
             padx=6, pady=3, cursor="hand2",
             command=self._browse_u2c_path).pack(side="left")

        sf = tk.Frame(c_u2c, bg=C("surface"))
        sf.pack(fill="x")
        HBtn(sf, C("accent2"), C("accent"),
             text="💾  Simpan Path U2C",
             font=(FONT, FS["base"], "bold"), fg=C("accent_txt"),
             activeforeground=C("accent_txt"), relief="flat", bd=0,
             padx=14, pady=6, cursor="hand2",
             command=self._save_u2c_path).pack(side="left")
        self._u2c_save_lbl = tk.Label(sf, text="", fg=C("success"),
                                       bg=C("surface"), font=(FONT, FS["sm"]))
        self._u2c_save_lbl.pack(side="left", padx=10)

        # ── RUN MODE ───────────────────────────────────────
        _make_section_label(p, "Run Mode")
        c2 = _make_card(p)
        self.mode_var = tk.StringVar(value="list_plants")

        for val, label_txt, sub, color in [
            ("list_plants", "Proses dari daftar Plants",
             "Hanya plant yang tercantum di atas", C("text")),
            ("auto_scan", "Auto-scan Not Completed",
             "Scan semua Not Completed di ListEod", C("warning")),
        ]:
            rf = tk.Frame(c2, bg=C("surface"))
            rf.pack(fill="x", pady=(4, 0))
            rb = tk.Radiobutton(
                rf, text=label_txt, variable=self.mode_var, value=val,
                fg=color, bg=C("surface"), selectcolor=C("rb_sel"),
                activebackground=C("surface"), activeforeground=color,
                font=(FONT, FS["base"], "bold"), cursor="hand2"
            )
            rb.pack(anchor="w")
            tk.Label(rf, text=f"  {sub}", fg=C("text3"), bg=C("surface"),
                     font=(FONT, FS["sm"])).pack(anchor="w")
        tk.Frame(c2, bg=C("surface"), height=4).pack()

        # ── INFO ───────────────────────────────────────────
        _make_section_label(p, "Info & Settings")
        c3 = _make_card(p)

        info_items = [
            (C("accent"),   "SMTP direct — tanpa Thunderbird"),
            (C("success"),  "Chrome dibuka otomatis via CDP"),
            (C("success"),  "Login portal otomatis"),
            (C("warning"),  "Pastikan SAP sudah terbuka"),
        ]
        for dot_c, msg in info_items:
            rf = tk.Frame(c3, bg=C("surface"))
            rf.pack(fill="x", pady=2)
            tk.Label(rf, text="●", fg=dot_c, bg=C("surface"),
                     font=(FONT, FS["sm"])).pack(side="left", padx=(0, 8))
            tk.Label(rf, text=msg, fg=C("text2"), bg=C("surface"),
                     font=(FONT, FS["base"])).pack(side="left")

        tk.Frame(c3, bg=C("sep"), height=1).pack(fill="x", pady=8)

        HBtn(c3, C("surface2"), C("surface3"),
             text="⚙  Konfigurasi Email SMTP",
             font=(FONT, FS["base"]), fg=C("text2"),
             relief="flat", bd=0, pady=8, cursor="hand2",
             command=self._open_email_config).pack(fill="x")

        tk.Frame(p, height=20, bg=C("bg")).pack()

    # ─────────────────────────────────────────
    # RIGHT PANEL
    # ─────────────────────────────────────────

    def _build_right(self, p):
        # Top padding
        tk.Frame(p, bg=C("bg"), height=14).pack()

        # ACTIVITY LOG
        top = tk.Frame(p, bg=C("bg"))
        top.pack(fill="both", expand=True, padx=16)

        hdr = tk.Frame(top, bg=C("bg"))
        hdr.pack(fill="x", pady=(0, 6))
        tk.Label(hdr, text="ACTIVITY LOG", fg=C("text3"), bg=C("bg"),
                 font=(FONT, FS["xs"], "bold")).pack(side="left")
        HBtn(hdr, C("bg"), C("surface2"),
             text="Bersihkan", font=(FONT, FS["sm"]), fg=C("text3"),
             relief="flat", bd=0, padx=10, pady=3, cursor="hand2",
             command=self._clear_log).pack(side="right")

        log_outer = tk.Frame(top, bg=C("border"), bd=0)
        log_outer.pack(fill="both", expand=True)
        log_inner = tk.Frame(log_outer, bg=C("surface"))
        log_inner.pack(fill="both", padx=1, pady=1)

        self.log = scrolledtext.ScrolledText(
            log_inner, bg=C("surface"), fg=C("text2"),
            font=(FONT, FS["base"]), relief="flat", bd=0,
            state="disabled", wrap="word",
            padx=16, pady=14,
            selectbackground=C("log_sel"),
        )
        self.log.pack(fill="both", expand=True)
        self.log.tag_config("INFO",  foreground=C("text2"))
        self.log.tag_config("OK",    foreground=C("success"))
        self.log.tag_config("WARN",  foreground=C("warning"))
        self.log.tag_config("ERROR", foreground=C("danger"))

        # Progress bar
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("RPA.Horizontal.TProgressbar",
                        troughcolor=C("surface2"),
                        background=C("accent"),
                        bordercolor=C("surface2"),
                        lightcolor=C("accent"),
                        darkcolor=C("accent2"),
                        thickness=4)
        self.prog = ttk.Progressbar(top, mode="indeterminate",
                                    style="RPA.Horizontal.TProgressbar")
        self.prog.pack(fill="x", pady=(6, 0))

        # Separator
        tk.Frame(p, bg=C("sep"), height=1).pack(fill="x", padx=16, pady=(10, 0))

        # LIMIT ALERT
        bot = tk.Frame(p, bg=C("bg"))
        bot.pack(fill="x", padx=16, pady=(8, 0))

        lhdr = tk.Frame(bot, bg=C("bg"))
        lhdr.pack(fill="x", pady=(0, 6))
        self._alert_dot = tk.Label(lhdr, text="●", fg=C("text3"), bg=C("bg"),
                                    font=(FONT, FS["sm"]))
        self._alert_dot.pack(side="left", padx=(0, 6))
        tk.Label(lhdr, text="LIMIT ALERT", fg=C("text3"), bg=C("bg"),
                 font=(FONT, FS["xs"], "bold")).pack(side="left")
        self._alert_count = tk.Label(lhdr, text="", fg=C("warning"), bg=C("bg"),
                                      font=(FONT, FS["sm"], "bold"))
        self._alert_count.pack(side="left", padx=(8, 0))
        HBtn(lhdr, C("bg"), C("surface2"),
             text="Bersihkan", font=(FONT, FS["sm"]), fg=C("text3"),
             relief="flat", bd=0, padx=10, pady=3, cursor="hand2",
             command=self._clear_alert_log).pack(side="right")

        alert_outer = tk.Frame(bot, bg=C("border"))
        alert_outer.pack(fill="x")
        alert_inner = tk.Frame(alert_outer, bg=C("surface"))
        alert_inner.pack(fill="both", padx=1, pady=1)

        self.alert_log = scrolledtext.ScrolledText(
            alert_inner, bg=C("surface"), fg=C("text2"),
            font=(FONT, FS["base"]), relief="flat", bd=0,
            state="disabled", wrap="word",
            padx=16, pady=10, height=7,
        )
        self.alert_log.pack(fill="x")
        self.alert_log.tag_config("HEAD",  foreground=C("warning"),
                                   font=(FONT, FS["base"], "bold"))
        self.alert_log.tag_config("ITEM",  foreground=C("text2"))
        self.alert_log.tag_config("OVER",  foreground=C("danger"))
        self.alert_log.tag_config("PLANT", foreground=C("accent"),
                                   font=(FONT, FS["sm"], "bold"))

        tk.Frame(p, bg=C("bg"), height=14).pack()

    # ─────────────────────────────────────────
    # BOTTOM BAR
    # ─────────────────────────────────────────

    def _build_bottom(self):
        bar = tk.Frame(self.root, bg=C("surface"), height=64)
        bar.pack(fill="x", side="bottom")
        bar.pack_propagate(False)

        # Thin top border
        tk.Frame(bar, bg=C("sep"), height=1).pack(fill="x")

        inner = tk.Frame(bar, bg=C("surface"))
        inner.pack(fill="both", expand=True, padx=20)

        self._lastrun = tk.Label(inner, text="Belum pernah dijalankan",
                                 fg=C("text3"), bg=C("surface"),
                                 font=(FONT, FS["sm"]))
        self._lastrun.pack(side="left", anchor="center", pady=14)

        btn_area = tk.Frame(inner, bg=C("surface"))
        btn_area.pack(side="right", pady=10)

        self.stop_btn = HBtn(
            btn_area, C("surface2"), C("stop_hov"),
            text="  ■  Stop  ",
            font=(FONT, FS["md"], "bold"), fg=C("text2"),
            activeforeground="#FFFFFF", relief="flat", bd=0,
            padx=16, pady=10, cursor="hand2", state="disabled",
            command=self._on_stop
        )
        self.stop_btn.pack(side="right", padx=(10, 0))

        self.run_btn = HBtn(
            btn_area, C("run_bg"), C("run_hov"),
            text="  ▶   Run RPA  ",
            font=(FONT, FS["lg"], "bold"), fg="#FFFFFF",
            activeforeground="#FFFFFF", relief="flat", bd=0,
            padx=24, pady=10, cursor="hand2",
            command=self._on_run
        )
        self.run_btn.pack(side="right")

    # ── HELPERS ──────────────────────────────────────────────

    def _load_email_defaults(self):
        try:
            from email_config_ui import load_credentials
            cred = load_credentials()

            def _load_into(list_attr, chip_attr, raw_val):
                lst = getattr(self, list_attr, None)
                chips = getattr(self, chip_attr, None)
                if lst is None or chips is None:
                    return
                lst.clear()
                for e in raw_val.replace("\n", ",").split(","):
                    e = e.strip()
                    if e and e not in lst:
                        lst.append(e)
                # chips = (el, entry, commit_fn, refresh_fn)
                if chips[3]:
                    chips[3]()  # call refresh_fn

            _load_into("_email_to_list", "_email_to_chips",
                       cred.get("email_to", ""))
            _load_into("_email_cc_list", "_email_cc_chips",
                       cred.get("email_cc", ""))
        except Exception:
            pass

    def _open_email_config(self):
        try:
            import subprocess, sys
            subprocess.Popen([sys.executable, "email_config_ui.py"])
            self.root.after(3000, self._load_email_defaults)
        except Exception as e:
            messagebox.showerror("Error", f"Gagal buka konfigurasi:\n{e}")

    def _set_running(self, on: bool):
        self._running = on
        if on:
            self.run_btn.config(state="disabled", text="  ⏳  Berjalan...")
            self.run_btn.recolor(C("surface2"), C("surface2"))
            self.stop_btn.config(state="normal")
            self.stop_btn.recolor(C("stop_bg"), C("stop_hov"))
            self._dot.config(fg=C("warning"))
            self._stat.config(text="Running", fg=C("warning"))
            self.prog.start(12)
        else:
            self.run_btn.config(state="normal", text="  ▶   Run RPA  ")
            self.run_btn.recolor(C("run_bg"), C("run_hov"))
            self.stop_btn.config(state="disabled")
            self.stop_btn.recolor(C("surface2"), C("stop_hov"))
            self._dot.config(fg=C("success"))
            self._stat.config(text="Ready", fg=C("text2"))
            self.prog.stop()
            self._lastrun.config(
                text=f"Terakhir dijalankan: {datetime.now().strftime('%d %b %Y  %H:%M')}")

    def _clear_log(self):
        self.log.config(state="normal")
        self.log.delete("1.0", "end")
        self.log.config(state="disabled")

    def _clear_alert_log(self):
        self.alert_log.config(state="normal")
        self.alert_log.delete("1.0", "end")
        self.alert_log.config(state="disabled")
        self._alert_total = 0
        self._alert_dot.config(fg=C("text3"))
        self._alert_count.config(text="")

    def _append_alert(self, plant: str, items_skip: list, limits: dict):
        if not items_skip:
            return
        self._alert_total += len(items_skip)
        self._alert_dot.config(fg=C("danger"))
        self._alert_count.config(text=f"{self._alert_total} item lewat limit")
        self.alert_log.config(state="normal")
        ts = datetime.now().strftime("%H:%M:%S")
        self.alert_log.insert("end",
            f"[{ts}]  Plant {plant}  —  {len(items_skip)} item lewat batas\n", "HEAD")
        for item in items_skip:
            lim   = limits.get(item["material"], {})
            lim_p = lim.get("limit_plus", "N/A")
            lim_m = lim.get("limit_minus", "N/A")
            diff  = item["diff"]
            mat   = item["material"]
            sloc  = item["sloc"]
            if diff > 0:
                batas = f"limit+ {lim_p}"
                arah  = f"{diff:+.3f} > {lim_p}"
            else:
                batas = f"limit- {lim_m}"
                arah  = f"{diff:+.3f} < {lim_m}"
            self.alert_log.insert("end",
                f"  ⚠  {mat}  SLoc={sloc}  diff={diff:+.6f}  [{batas}]  {arah}\n", "OVER")
        self.alert_log.insert("end", "\n", "ITEM")
        self.alert_log.see("end")
        self.alert_log.config(state="disabled")

    def _write_log(self, msg: str, level: str = "INFO"):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n", level)
        self.log.see("end")
        self.log.config(state="disabled")

    # ── PLANT POPUP HANDLERS ──────────────────────────────────

    def _toggle_plant_popup(self, anchor_btn=None):
        if self._plant_popup_visible:
            self._plant_popup.pack_forget()
            self._plant_popup_visible = False
        else:
            self._plant_popup.pack(fill="x", pady=(0, 4))
            self._plant_popup_visible = True

    def _update_plant_summary(self):
        selected = [p2 for p2, v in self._plant_vars.items() if v.get()]
        total    = len(self._plant_vars)
        if len(selected) == 0:
            self._plant_summary.set("— tidak ada —")
        elif len(selected) == total:
            self._plant_summary.set(f"Semua ({total} plant)")
        elif len(selected) <= 3:
            self._plant_summary.set(", ".join(selected))
        else:
            self._plant_summary.set(f"{', '.join(selected[:3])} +{len(selected) - 3} lagi")

    def _plant_select_all(self):
        for v in self._plant_vars.values():
            v.set(True)
        self._update_plant_summary()
        if hasattr(self, "_rebuild_plant_grid"):
            q = self._plant_search_var.get() if hasattr(self, "_plant_search_var") else ""
            self._rebuild_plant_grid(q)

    def _plant_clear_all(self):
        for v in self._plant_vars.values():
            v.set(False)
        self._update_plant_summary()
        if hasattr(self, "_rebuild_plant_grid"):
            q = self._plant_search_var.get() if hasattr(self, "_plant_search_var") else ""
            self._rebuild_plant_grid(q)

    def get_selected_plants(self) -> list:
        return [p2 for p2, v in self._plant_vars.items() if v.get()]

    def _show_tooltip(self, widget, text: str):
        self._hide_tooltip()
        x = widget.winfo_rootx() + 10
        y = widget.winfo_rooty() + widget.winfo_height() + 4
        self._tooltip = tk.Toplevel(self.root)
        self._tooltip.wm_overrideredirect(True)
        self._tooltip.wm_geometry(f"+{x}+{y}")
        tk.Label(self._tooltip, text=text,
                 bg=C("surface2"), fg=C("text"),
                 font=(FONT, FS["sm"]), padx=8, pady=4,
                 relief="flat").pack()

    def _hide_tooltip(self):
        if hasattr(self, "_tooltip") and self._tooltip:
            try:
                self._tooltip.destroy()
            except Exception:
                pass
            self._tooltip = None

    def _get_email_to(self) -> str:
        # Commit any pending text in entry first
        if hasattr(self, "_email_to_chips"):
            el, entry, commit_fn, refresh_fn = self._email_to_chips
            if entry and commit_fn:
                commit_fn()
        return ", ".join(getattr(self, "_email_to_list", []))

    def _get_email_cc(self) -> str:
        if hasattr(self, "_email_cc_chips"):
            el, entry, commit_fn, refresh_fn = self._email_cc_chips
            if entry and commit_fn:
                commit_fn()
        return ", ".join(getattr(self, "_email_cc_list", []))

    def _browse_plant_map(self):
        from tkinter import filedialog
        current  = self.plant_map_var.get()
        init_dir = os.path.dirname(current) if current and os.path.exists(os.path.dirname(current)) else os.path.expanduser("~")
        path = filedialog.askopenfilename(
            title="Pilih file Excel Plant Mapping",
            initialdir=init_dir,
            filetypes=[("Excel File", "*.xlsx *.xls"), ("All Files", "*.*")],
        )
        if path:
            self.plant_map_var.set(path)
            from config import Config
            Config.PLANT_MAPPING_FILE = path
            plants = self._load_plants_from_excel(path)
            self._rebuild_plant_checklist(self._plant_grid, plants)
            if hasattr(self, '_plant_count_lbl'):
                self._plant_count_lbl.config(text=f"{len(self._plant_vars)} plant tersedia")
            n = len(plants)
            is_fallback = (sorted(plants) == sorted(list(Config.PLANTS)) and n == len(Config.PLANTS))
            if not is_fallback:
                status_text  = f"  ✔ {n} plant dibaca dari Excel"
                status_color = C("success")
            else:
                status_text  = "  ⚠ Gagal baca Excel — cek path & sheet 'Plant_CostCenter'"
                status_color = C("warning")
            if hasattr(self, "_plant_map_status_lbl"):
                self._plant_map_status_lbl.config(text=status_text, fg=status_color)
            self._write_log(
                f"Plant Map: {path}  →  {n} plant dimuat" if not is_fallback
                else f"⚠ Plant Map gagal dibaca — menggunakan fallback hardcode",
                "OK" if not is_fallback else "WARN"
            )

    def _browse_u2c_path(self):
        from tkinter import filedialog
        current   = self.u2c_path_var.get()
        init_dir  = os.path.dirname(current) if current else r"C:\Users\User\Documents\PGD\EOD"
        init_file = os.path.basename(current) if current else "U2C.txt"
        path = filedialog.asksaveasfilename(
            title="Pilih lokasi & nama file U2C",
            initialdir=init_dir,
            initialfile=init_file,
            defaultextension=".txt",
            filetypes=[("Text File", "*.txt"), ("All Files", "*.*")],
        )
        if path:
            self.u2c_path_var.set(path)
            self._u2c_save_lbl.config(text="")

    def _save_u2c_path(self):
        path = self.u2c_path_var.get().strip()
        if not path:
            messagebox.showwarning("Peringatan", "Path file U2C tidak boleh kosong!")
            return
        try:
            from u2c_upload import save_u2c_config
            save_u2c_config(path)
            self._u2c_save_lbl.config(text="✔ Tersimpan", fg=C("success"))
            self.root.after(3000, lambda: self._u2c_save_lbl.config(text=""))
        except Exception as e:
            messagebox.showerror("Error", f"Gagal simpan config U2C:\n{e}")

    # ── RUN / STOP ────────────────────────────────────────────

    def _on_run(self):
        self._on_portal_change()
        if self._running:
            return

        plants       = self.get_selected_plants()
        posting_date = self.date_var.get().strip()
        email_to     = self._get_email_to().strip()
        email_cc     = self._get_email_cc().strip()
        mode         = self.mode_var.get()

        if mode == "list_plants" and not plants:
            messagebox.showwarning("Peringatan", "Masukkan minimal 1 plant!")
            return
        if not self._get_email_to():
            messagebox.showwarning("Peringatan", "Email To tidak boleh kosong!")
            return

        try:
            from email_config_ui import load_credentials
            load_credentials()
        except FileNotFoundError:
            if messagebox.askyesno("Setup Email",
                                   "Kredensial email belum diset.\n"
                                   "Buka konfigurasi email sekarang?"):
                self._open_email_config()
            return
        except Exception as e:
            messagebox.showerror("Error", f"Gagal baca kredensial:\n{e}")
            return

        stop_event.clear()
        self._set_running(True)
        self._write_log(f"Run RPA  ·  {posting_date}  ·  mode={mode}", "OK")
        self._write_log(f"Email → {email_to}" + (f"  CC: {email_cc}" if email_cc else ""), "INFO")

        from config import Config
        self._write_log(
            f"SAP T-code: /{Config.ACTIVE_TCODE_SAPSTK}  +  /{Config.ACTIVE_TCODE_U2C}  "
            f"(portal: {self._portal_var.get()})",
            "INFO"
        )

        threading.Thread(
            target=run_robot,
            args=(plants, posting_date, email_to, email_cc, mode),
            daemon=True
        ).start()

    def _on_stop(self):
        stop_event.set()
        self._write_log("=" * 40, "WARN")
        self._write_log("Stop diklik -- menghentikan robot...", "WARN")
        self._write_log("Tunggu proses saat ini selesai.", "WARN")
        self._write_log("=" * 40, "WARN")
        self.stop_btn.config(state="disabled")

    # ── POLL ─────────────────────────────────────────────────

    def _poll(self):
        try:
            while True:
                level, msg = log_queue.get_nowait()
                if level == "DONE":
                    self._set_running(False)
                    self._write_log("━" * 50, "INFO")
                    self._write_log("RPA selesai  —  cek inbox email.", "OK")
                elif level == "WAIT_LOGIN":
                    self.root.after(0, self._ask_login_confirm)
                elif level == "SAP_WARNING":
                    self.root.after(0, self._show_sap_warning)
                elif level == "SAP_WAIT":
                    self.root.after(0, self._ask_sap_confirm)
                else:
                    self._write_log(msg, level)
        except queue.Empty:
            pass
        self.root.after(80, self._poll)

    # ── DIALOGS ──────────────────────────────────────────────

    def _make_dialog(self, title: str, w=460, h=230) -> tk.Toplevel:
        popup = tk.Toplevel(self.root)
        popup.title(title)
        popup.geometry(f"{w}x{h}")
        popup.configure(bg=C("surface"))
        popup.resizable(False, False)
        popup.grab_set()
        popup.lift()
        popup.focus_force()
        # Center on parent
        self.root.update_idletasks()
        rx = self.root.winfo_x() + (self.root.winfo_width() - w) // 2
        ry = self.root.winfo_y() + (self.root.winfo_height() - h) // 2
        popup.geometry(f"{w}x{h}+{rx}+{ry}")
        return popup

    def _ask_sap_confirm(self):
        self._write_log("━" * 50, "WARN")
        self._write_log("⚠  SAP belum terdeteksi!", "WARN")
        self._write_log("   Buka SAP GUI dan login terlebih dahulu,", "WARN")
        self._write_log("   lalu klik OK untuk mulai download SAPSTK.", "WARN")
        self._write_log("━" * 50, "WARN")

        popup = self._make_dialog("Buka SAP", w=460, h=220)

        tk.Label(popup, text="🖥  Buka SAP GUI",
                 font=(FONT, FS["xl"], "bold"),
                 fg=C("text"), bg=C("surface")).pack(pady=(24, 8))
        tk.Label(popup,
                 text="Silakan buka SAP GUI dan login.\n\nKlik OK setelah SAP terbuka dan siap digunakan.",
                 font=(FONT, FS["base"]),
                 fg=C("text2"), bg=C("surface"), justify="center").pack(pady=(0, 18))

        def on_ok():
            popup.destroy()
            sap_event.set()

        HBtn(popup, C("run_bg"), C("run_hov"),
             text="✓  SAP Sudah Terbuka — Lanjutkan",
             font=(FONT, FS["base"], "bold"), fg="#FFFFFF",
             activeforeground="#FFFFFF", relief="flat", bd=0,
             padx=24, pady=10, cursor="hand2",
             command=on_ok).pack(pady=(0, 18))

    def _show_sap_warning(self):
        self._write_log("━" * 50, "WARN")
        self._write_log("⚠  SAP sedang dikontrol robot!", "WARN")
        self._write_log("   Jangan sentuh keyboard / mouse", "WARN")
        self._write_log("   sampai log menampilkan 'Download SAP selesai'", "WARN")
        self._write_log("━" * 50, "WARN")
        self._dot.config(fg=C("warning"))
        self._stat.config(text="SAP Running", fg=C("warning"))

    def _ask_login_confirm(self):
        self._write_log("━" * 50, "WARN")
        self._write_log("⚠  Chrome terbuka — silakan LOGIN di browser!", "WARN")
        self._write_log("   Masukkan username & password di Chrome,", "WARN")
        self._write_log("   lalu klik OK di sini untuk melanjutkan RPA.", "WARN")
        self._write_log("━" * 50, "WARN")

        popup = self._make_dialog("Konfirmasi Login", w=460, h=230)

        tk.Label(popup, text="🔐  Login Portal Mayora",
                 font=(FONT, FS["xl"], "bold"),
                 fg=C("text"), bg=C("surface")).pack(pady=(24, 8))
        tk.Label(popup,
                 text="Silakan masukkan username & password\ndi Chrome yang sudah terbuka.\n\nKlik OK setelah berhasil login.",
                 font=(FONT, FS["base"]),
                 fg=C("text2"), bg=C("surface"), justify="center").pack(pady=(0, 18))

        def on_ok():
            popup.destroy()
            login_event.set()

        HBtn(popup, C("run_bg"), C("run_hov"),
             text="✓  Sudah Login — Lanjutkan RPA",
             font=(FONT, FS["base"], "bold"), fg="#FFFFFF",
             activeforeground="#FFFFFF", relief="flat", bd=0,
             padx=24, pady=10, cursor="hand2",
             command=on_ok).pack(pady=(0, 18))


# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    try:
        root.iconbitmap("icon.ico")
    except Exception:
        pass
    RpaGui(root)
    root.mainloop()