"""
maintain_material.py
Maintain Material Module — RPA PT Mayora Indah Tbk

Launched from launcher.py — receives back_callback to return to menu.

Flow:
  1. User browses material codes Excel file (Kode Barang column)
  2. User browses List_Plant.xlsx (Plant + Prch.Grp)
  3. User picks plants from checklist
  4. Click Run → SAP MM01 automation starts
     Outer loop : each selected plant
     Inner loop : each material code
       Step 1 — MM01 Initial Screen  (pyautogui — type & navigate)
       Step 2 — Select Views popup   (pyautogui — just Enter)
       Step 3 — Org Levels popup     (pyautogui — fill fields)
       Step 4 — Purchasing tab       (pyautogui — tab to Prch.Grp)
       Step 5 — Accounting 1 tab     (pyautogui — tab to Moving Price)

No SAP COM / SAP GUI scripting used.
Robot types exactly as a human would — SAP sees only keyboard input.
"""

import os
import time
import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext, filedialog
import threading
import queue
import pyautogui
try:
    import keyboard   # pip install keyboard
    _KEYBOARD_OK = True
except ImportError:
    _KEYBOARD_OK = False
import win32gui
import win32con
import openpyxl
from datetime import datetime

# ─────────────────────────────────────────────
# ENABLE DPI AWARENESS - MUST BE FIRST
# ─────────────────────────────────────────────

try:
    from ctypes import windll
    windll.shcore.SetProcessDpiAwareness(2)
except Exception:
    pass

try:
    from theme_manager import get_theme_manager
    theme = get_theme_manager()
except ImportError:
    theme = None

# ─────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────

FONT       = "Segoe UI"
FONT_DISP  = "Segoe UI"
FONT_TITLE = ("Segoe UI", 12, "bold")
FONT_SUB   = ("Segoe UI", 10)
FONT_SMALL = ("Segoe UI", 9)

# ─────────────────────────────────────────────
# SAP CONSTANTS
# ─────────────────────────────────────────────

INDUSTRY      = "F"
MATERIAL_TYPE = "FERT"
SLOC          = "WH01"
SALES_ORG     = "CS00"
MOVING_PRICE  = "1"
DC_GT         = "41"
DC_MT         = "21"
COPY_PLANT_GT = "B100"
COPY_PLANT_MT = "B242"

# Keywords to identify SAP window
SAP_KEYWORDS = [
    "SAP Easy Access", "SAP R/3", "SAP NetWeaver",
    "Create Material", "MM01",
]
SKIP_TITLES = [
    "SAP Logon", "Firefox", "Chrome", "Edge",
    "Visual Studio", "Code", "Notepad", "Claude",
    "WPS", "Excel", "Word", "PowerPoint", "LibreOffice",
    "Spreadsheet", "Writer", "Calc",
]

# SAP window class names — more reliable than title matching
SAP_CLASS_NAMES = ["SAP_FRONTEND_SESSION", "SAP GUI"]

pyautogui.FAILSAFE = True
pyautogui.PAUSE    = 0.25

SMART_WAIT = False   # toggled by UI checkbox — use screenshot diffing instead of fixed waits


# ─────────────────────────────────────────────
# HARD THREAD KILL
# ─────────────────────────────────────────────

import ctypes as _ctypes

def _raise_in_thread(thread_id: int):
    """
    Inject a SystemExit exception into a running thread.
    This causes the thread to stop immediately — no waiting for loop checks.
    """
    res = _ctypes.pythonapi.PyThreadState_SetAsyncExc(
        _ctypes.c_ulong(thread_id),
        _ctypes.py_object(SystemExit),
    )
    return res


# ─────────────────────────────────────────────
# DATA LOADERS  (unchanged from original)
# ─────────────────────────────────────────────

def load_plant_data(filepath: str) -> dict:
    """
    Read Plant_Master_MM01.xlsx — reads ALL sheets.
    Return: { plant_code: { name, prch_grp, type, dc, copy_plant, copy_dc } }
    """
    import re as _re
    wb     = openpyxl.load_workbook(filepath, data_only=True)
    result = {}

    for sheet_name in wb.sheetnames:
        ws           = wb[sheet_name]
        header_found = False
        col = {"plant": 0, "name": 1, "prch_grp": 2, "type": 3,
               "dc": 4, "copy_plant": 5, "copy_dc": 6}

        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            cell0 = str(row[0]).strip()

            if cell0 == "Plant":
                header_found = True
                for ci, cell in enumerate(row):
                    if cell is None:
                        continue
                    h = str(cell).strip().lower()
                    if h == "plant":
                        col["plant"] = ci
                    elif "name" in h:
                        col["name"] = ci
                    elif "prch" in h or "purch" in h or "grp" in h:
                        col["prch_grp"] = ci
                    elif h == "type":
                        col["type"] = ci
                    elif "dc" in h and "dest" in h:
                        col["dc"] = ci
                    elif "copy" in h and "plant" in h:
                        col["copy_plant"] = ci
                    elif "copy" in h and "dc" in h:
                        col["copy_dc"] = ci
                continue

            if not header_found:
                continue
            if not any(c.isdigit() for c in cell0):
                continue

            def _safe(idx, default=""):
                v = row[idx] if idx < len(row) else None
                if v is None:
                    return default
                s = str(v).strip()
                s = _re.sub(r"[★*]", "", s).strip()
                return s

            plant_code = cell0
            plant_name = _safe(col["name"])

            prch_raw = row[col["prch_grp"]] if col["prch_grp"] < len(row) else None
            prch_grp = ""
            if prch_raw is not None:
                try:
                    prch_grp = str(int(float(str(prch_raw)))).strip()
                except Exception:
                    prch_grp = str(prch_raw).strip()

            type_raw = _safe(col["type"]).upper()
            if "MT" in type_raw:
                ptype = "MT"
            elif "GT" in type_raw:
                ptype = "GT"
            else:
                ptype = "MT" if plant_name.upper().strip().endswith(" MT") else "GT"

            dc         = _safe(col["dc"])         or (DC_MT if ptype == "MT" else DC_GT)
            copy_plant = _safe(col["copy_plant"]) or (COPY_PLANT_MT if ptype == "MT" else COPY_PLANT_GT)
            copy_dc    = _safe(col["copy_dc"])    or (DC_MT if ptype == "MT" else DC_GT)

            for _v in [dc, copy_dc]:
                try:
                    _v = str(int(float(_v)))
                except Exception:
                    pass

            result[plant_code] = {
                "name":       plant_name,
                "prch_grp":   prch_grp,
                "type":       ptype,
                "dc":         dc,
                "copy_plant": copy_plant,
                "copy_dc":    copy_dc,
            }

    return result


def load_material_codes(filepath: str) -> list:
    """
    Read material list Excel.
    Returns: [{'code': '315332', 'dc': '41'}, ...]
    Blank DC = applies to all plants.
    """
    wb      = openpyxl.load_workbook(filepath, data_only=True)
    entries = []

    for sheet_name in wb.sheetnames:
        ws         = wb[sheet_name]
        kode_col   = None
        dc_col     = None
        header_row = None

        for ri, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if not row:
                continue
            for ci, cell in enumerate(row):
                if not cell:
                    continue
                h = str(cell).strip()
                if h == "Kode Barang":
                    kode_col   = ci
                    header_row = ri
                if h == "DC":
                    dc_col = ci
            if kode_col is not None:
                break

        if kode_col is None:
            continue

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or kode_col >= len(row):
                continue
            val = row[kode_col]
            if val is None:
                continue
            try:
                code = str(int(float(str(val)))).strip()
            except Exception:
                code = str(val).strip()
            if not code:
                continue

            dc = ""
            if dc_col is not None and dc_col < len(row) and row[dc_col] is not None:
                try:
                    dc = str(int(float(str(row[dc_col])))).strip()
                except Exception:
                    dc = str(row[dc_col]).strip()

            entries.append({"code": code, "dc": dc})

        if entries:
            break

    return entries


# ─────────────────────────────────────────────
# GLOBAL KILLSWITCH — press Q to stop the robot
# ─────────────────────────────────────────────

# Module-level stop flag shared between GUI and robot thread
_KILL_FLAG = threading.Event()


# Module-level reference to the GUI instance so Q key can hard-kill the thread
_gui_instance    = None
_watchdog_active = threading.Event()


def _start_corner_watchdog():
    """
    Background thread that polls mouse position 20x/sec.
    The moment cursor hits top-left corner (within 5px),
    it hard-kills the RPA thread instantly — same as Q key.
    No need to wait for the next pyautogui action.
    """
    def _watch():
        while _watchdog_active.is_set():
            try:
                x, y = pyautogui.position()
                if x <= 5 and y <= 5:
                    _KILL_FLAG.set()
                    if (_gui_instance
                            and _gui_instance._rpa_thread
                            and _gui_instance._rpa_thread.is_alive()):
                        _raise_in_thread(_gui_instance._rpa_thread.ident)
                    break
            except Exception:
                break
            time.sleep(0.05)   # check 20x per second

    t = threading.Thread(target=_watch, daemon=True)
    t.start()


def _stop_corner_watchdog():
    _watchdog_active.clear()


def _register_killswitch():
    """
    Register Q key as an emergency stop hotkey.
    Works even when SAP is the foreground window.
    Requires the keyboard package (pip install keyboard).
    """
    if not _KEYBOARD_OK:
        return
    try:
        def _q_kill():
            _KILL_FLAG.set()
            # Hard-kill the RPA thread immediately
            if _gui_instance and _gui_instance._rpa_thread and _gui_instance._rpa_thread.is_alive():
                _raise_in_thread(_gui_instance._rpa_thread.ident)
        keyboard.add_hotkey("q", _q_kill, suppress=False)
    except Exception:
        pass


def _unregister_killswitch():
    if not _KEYBOARD_OK:
        return
    try:
        keyboard.remove_hotkey("q")
    except Exception:
        pass


# ─────────────────────────────────────────────
# SAP WINDOW FOCUS  (no COM — just win32gui)
# ─────────────────────────────────────────────

def _focus_sap() -> bool:
    """
    Find and bring SAP window to foreground using only win32gui.
    Checks BOTH window class name (most reliable) and title keywords.
    Returns True if found, False if SAP window not found.
    """
    found = []

    def cb(hwnd, _):
        if not win32gui.IsWindowVisible(hwnd):
            return
        title = win32gui.GetWindowText(hwnd)
        if not title:
            return
        # Skip known non-SAP apps by title
        if any(s in title for s in SKIP_TITLES):
            return
        # Check window class name first — most reliable SAP identifier
        try:
            cls = win32gui.GetClassName(hwnd)
            if any(sc in cls for sc in SAP_CLASS_NAMES):
                found.append((0, hwnd))   # priority 0 = class match
                return
        except Exception:
            pass
        # Fallback: title keyword match
        if any(kw in title for kw in SAP_KEYWORDS):
            found.append((1, hwnd))       # priority 1 = title match

    win32gui.EnumWindows(cb, None)

    # Sort by priority — class matches first
    found.sort(key=lambda x: x[0])
    found = [hwnd for _, hwnd in found]

    if not found:
        return False

    hwnd = found[0]
    win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
    win32gui.SetForegroundWindow(hwnd)
    time.sleep(0.6)
    return True


# ─────────────────────────────────────────────
# PURE PYAUTOGUI HELPERS  (no COM at all)
# ─────────────────────────────────────────────

def _wait(seconds: float):
    """
    Wait for SAP to render the next screen.
    If SMART_WAIT is enabled, uses screenshot-diff detection instead
    of the fixed `seconds` value (seconds becomes the max timeout).
    """
    if SMART_WAIT:
        _wait_for_screen_change(timeout=max(seconds, 3.0))
    else:
        time.sleep(seconds)


def _wait_for_sap_ready(timeout: float = 8.0, poll: float = 0.15):
    """
    Smart wait — polls SAP window responsiveness every 150ms.
    SAP marks its window as "Not Responding" (HWND unresponsive) while
    processing a screen transition. We wait until it becomes responsive
    again, then add a small settle buffer before proceeding.

    - Fast network: returns as soon as SAP is responsive (~0.3s)
    - Slow network: keeps waiting up to timeout seconds
    - Never proceeds while SAP is still loading
    """
    import ctypes as _ct

    SMTO_ABORTIFHUNG = 0x0002
    deadline = time.time() + timeout

    # First wait until SAP becomes unresponsive (started processing)
    # then wait until it becomes responsive again (done processing).
    # If SAP never becomes unresponsive (fast response), just wait min 0.3s.
    _min_wait = 0.3
    time.sleep(_min_wait)

    while time.time() < deadline:
        try:
            hwnd = _get_sap_hwnd()
            if not hwnd:
                break
            # SendMessageTimeout returns 0 if window is hung/not responding
            result = _ct.c_ulong(0)
            ret = _ct.windll.user32.SendMessageTimeoutW(
                hwnd, 0x0000, 0, 0,   # WM_NULL
                SMTO_ABORTIFHUNG, 300,  # 300ms timeout
                _ct.byref(result)
            )
            if ret != 0:
                # Window is responsive — SAP finished processing
                time.sleep(0.15)  # small settle buffer
                return
        except Exception:
            break
        time.sleep(poll)


def _wait_for_screen_change(timeout: float = 8.0, poll: float = 0.15,
                             stable_for: float = 0.3) -> bool:
    """
    Detect SAP screen changes WITHOUT SAP scripting/COM by comparing
    screenshots of the SAP window region over time.

    How it works:
    1. Take a screenshot of the SAP window right now (the "before" state)
    2. Keep taking screenshots in a loop
    3. The screen has "changed" once a new screenshot differs from the
       previous one — meaning SAP rendered something new (popup, error,
       next screen, etc.)
    4. Once the screenshot stops changing for `stable_for` seconds, the
       screen is considered stable/settled and we return.

    This catches BOTH fast SAP (returns almost instantly) and slow SAP
    (keeps waiting as long as needed, up to timeout).

    Returns True if a change was detected, False if timed out without
    any change (SAP may be frozen, or the action had no visual effect).
    """
    hwnd = _get_sap_hwnd()
    if not hwnd:
        time.sleep(0.5)
        return False

    try:
        rect = win32gui.GetWindowRect(hwnd)
    except Exception:
        time.sleep(0.5)
        return False

    deadline      = time.time() + timeout
    prev_img      = None
    changed_once  = False
    stable_timer  = 0.0

    while time.time() < deadline:
        try:
            shot = pyautogui.screenshot(region=(
                rect[0], rect[1], rect[2]-rect[0], rect[3]-rect[1]))
            # Downscale for fast comparison — exact pixels don't matter,
            # just whether something visibly changed
            shot_small = shot.resize((120, 90))
            cur_bytes  = shot_small.tobytes()

            if prev_img is not None:
                if cur_bytes != prev_img:
                    changed_once = True
                    stable_timer = 0.0
                else:
                    stable_timer += poll
                    if changed_once and stable_timer >= stable_for:
                        return True
                    elif not changed_once and stable_timer >= 1.0:
                        # Nothing changed for 1s straight — give up waiting,
                        # action may not have produced a visual change
                        return False

            prev_img = cur_bytes
        except Exception:
            break
        time.sleep(poll)

    return changed_once


def _get_sap_hwnd():
    """Return the hwnd of the SAP window, or None."""
    found = []
    def cb(hwnd, _):
        if not win32gui.IsWindowVisible(hwnd):
            return
        title = win32gui.GetWindowText(hwnd)
        if not title:
            return
        if any(s in title for s in SKIP_TITLES):
            return
        try:
            cls = win32gui.GetClassName(hwnd)
            if any(sc in cls for sc in SAP_CLASS_NAMES):
                found.append((0, hwnd))
                return
        except Exception:
            pass
        if any(kw in title for kw in SAP_KEYWORDS):
            found.append((1, hwnd))
    win32gui.EnumWindows(cb, None)
    found.sort(key=lambda x: x[0])
    return found[0][1] if found else None


def _tab(n: int = 1, delay: float = 0.18):
    for _ in range(n):
        pyautogui.press("tab")
        time.sleep(delay)


def _enter(n: int = 1, delay: float = 0.4):
    """Press Enter n times, waiting delay seconds between each."""
    for _ in range(n):
        pyautogui.press("enter")
        time.sleep(delay)


def _type(value: str, interval: float = 0.07):
    """Select-all then type value into current field."""
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.08)
    pyautogui.typewrite(str(value), interval=interval)
    time.sleep(0.08)


def _delete_clear():
    """
    Clear a SAP field by selecting all text with Ctrl+Shift+Right
    then pressing Delete. Fast and reliable — no timed loops.
    """
    pyautogui.hotkey("ctrl", "shift", "right")
    time.sleep(0.15)
    pyautogui.press("delete")
    time.sleep(0.2)


def _tcode(code: str):
    """
    Navigate to a transaction code via the command field.
    Works like a human: Ctrl+/ to jump to command field, type /nXX01, Enter.
    """
    pyautogui.hotkey("ctrl", "slash")   # focus command field
    time.sleep(0.3)
    pyautogui.hotkey("ctrl", "a")
    time.sleep(0.1)
    pyautogui.typewrite(f"/n{code}", interval=0.07)
    time.sleep(0.15)
    pyautogui.press("enter")


# ─────────────────────────────────────────────
# MM01 — ONE MATERIAL × ONE PLANT  (pure pyautogui)
# ─────────────────────────────────────────────

# Timing constants — adjust if SAP is slow on your network
T_TCODE    = 1.2   # wait after navigating to MM01
T_POPUP    = 1.0   # wait for a popup to appear
T_TAB      = 0.18  # between tab presses
T_FIELD    = 0.15  # after typing into a field
T_ENTER    = 0.45  # after pressing Enter
T_SAVE     = 1.2   # after final save Enter sequence


def _mm01_one(material: str, plant_info: dict,
              plant_code: str, log_fn=None,
              prev_material: str = "",
              org_levels_open: bool = False,
              auto_skip_maintained: bool = False,
              prev_type: str = "") -> bool:
    """
    Run MM01 for one material x one plant.
    Pure pyautogui -- no COM, no SAP scripting API.
    Robot types exactly as a human would.

    SAP is already on MM01 initial screen — either from _tcode(MM01) called
    once before the loop starts, or from returning after a save.

    Returns True on success, False on error.
    """
    def _l(msg, level="INFO"):
        if log_fn:
            log_fn(msg, level)

    prch_grp   = plant_info["prch_grp"]
    dc         = plant_info["dc"]
    copy_plant = plant_info["copy_plant"]
    copy_dc    = plant_info["copy_dc"]
    ptype      = plant_info["type"]

    _l(f"  [{ptype}] {plant_code} | {material} | "
       f"PrchGrp={prch_grp} | DC={dc} | CopyFrom={copy_plant}")

    try:
        if not org_levels_open:
            # Normal flow — focus SAP and fill MM01 initial screen
            if not _focus_sap():
                raise Exception("SAP window not found. Make sure SAP is open and logged in.")

            pyautogui.hotkey("ctrl", "Home")
            time.sleep(0.3)

        # ── STEP 1 & 2: MM01 Initial Screen + Select Views ─
        if org_levels_open:
            # Org Levels popup already open from previous "already maintained" skip
            # Jump straight to Step 3 — just fill the new plant info
            pass
        elif prev_material == "":
            # Very first material — fields may not be clean, full sequence.
            # 1. Delete 7s on Material field
            _delete_clear()

            # 2. Type material code
            _type(material);         _wait(T_FIELD)

            # 3. Press Down → Industry Sector
            pyautogui.press("down"); _wait(0.18)

            # 4. Type F
            _type(INDUSTRY);         _wait(T_FIELD)

            # 5. Press Down → Material Type
            pyautogui.press("down"); _wait(0.18)

            # 6. Type FERT
            _type(MATERIAL_TYPE);    _wait(T_FIELD)

            # 7. Press Down → Change Number
            pyautogui.press("down"); _wait(0.18)

            # 8. Press Down → Copy from Material
            pyautogui.press("down"); _wait(0.18)

            # 9. Delete 7s then type same material code
            _delete_clear()
            _type(material);         _wait(T_FIELD)

        elif material == prev_material:
            # Same material code, next plant — Copy from is already correct.
            # Just fill the Material field and go.
            _type(material);         _wait(T_FIELD)

        else:
            # New material code — fill Material field, then navigate to
            # Copy from and delete old code first before filling new one.
            _type(material);         _wait(T_FIELD)

            # Navigate down to Copy from field
            pyautogui.press("down"); _wait(0.18)   # Industry Sector
            pyautogui.press("down"); _wait(0.18)   # Material Type
            pyautogui.press("down"); _wait(0.18)   # Change Number
            pyautogui.press("down"); _wait(0.18)   # Copy from Material

            # Delete old code first then fill new one
            _delete_clear()
            _type(material);         _wait(T_FIELD)

        # Enter → SAP shows Select Views popup
        if not org_levels_open:
            _enter()                 # Enter → SAP shows Select Views popup
            _wait(T_POPUP)

            # ── STEP 2: Select Views popup ─────────────────
            _enter()
            _wait(1.5)   # wait until Org Levels popup is stable

        # ── STEP 3: Organizational Levels popup ────────────
        # DO NOT call _focus_sap() here — the popup is a child dialog inside SAP.
        # _focus_sap() would grab the main SAP window behind it and steal focus
        # away from the popup. The popup already has focus when it opens — just
        # wait for it and start typing immediately.
        _wait(T_POPUP)               # extra wait to make sure popup is fully rendered

        if prev_type == ptype:
            # Same plant type as previous (GT→GT or MT→MT) — this applies
            # whether coming from a normal plant or from an "already maintained"
            # skip, since the Org Levels popup retains DC, Copy From, and
            # Copy DC from the last fill. Only the Plant field needs updating.
            _delete_clear();         _wait(T_FIELD)
            _type(plant_code);       _wait(T_FIELD)

        else:
            # Type changed (GT→MT or MT→GT) OR first plant OR coming from skip.
            # Fill all fields fresh.

            # 1. Fill in the plant we chose
            _delete_clear();         _wait(T_FIELD)
            _type(plant_code);       _wait(T_FIELD)

            # 2. Press Down x4
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)

            # 3. Fill in our plant DC (41 GT / 21 MT)
            _delete_clear();         _wait(T_FIELD)
            _type(dc);               _wait(T_FIELD)

            # 4. Press Tab
            _tab(1)

            # 5. Fill in the plant we are copying from (B100 GT / B242 MT)
            _delete_clear();         _wait(T_FIELD)
            _type(copy_plant);       _wait(T_FIELD)

            # 6. Press Down x4
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)
            pyautogui.press("down"); _wait(0.18)

            # 7. Fill in the copy DC (41 GT / 21 MT)
            _delete_clear();         _wait(T_FIELD)
            _type(copy_dc);          _wait(T_FIELD)

        # Enter → popup closes → arrive on page 1
        # BUT if material already maintained, SAP shows an error popup.
        # We wait a bit longer then take a screenshot pixel-check isn't available,
        # so we use a simple approach: check window titles via win32gui after Enter.
        _enter()
        _wait(T_ENTER)  # wait for page 1 or error popup — adjustable in Timing settings

        # ── CHECK: "Material already maintained" error popup ──
        # If SAP shows the error, a second small window appears.
        # Detect it by looking for a second visible SAP-style window.
        _already_maintained = False
        try:
            import win32gui as _wg
            _wins = []
            def _cb(hwnd, _):
                if not _wg.IsWindowVisible(hwnd):
                    return
                t = _wg.GetWindowText(hwnd)
                if "Error" in t or "already" in t.lower() or "maintained" in t.lower():
                    _wins.append(hwnd)
            _wg.EnumWindows(_cb, None)
            if _wins:
                _already_maintained = True
        except Exception:
            pass

        if _already_maintained:
            if log_fn:
                log_fn(f"  ⚠ {material} already maintained for {plant_code}", "WARN")

            if auto_skip_maintained:
                # Remember Me was set — skip without asking
                if log_fn:
                    log_fn(f"  → Auto-skipping {plant_code} (Remember Me active)", "WARN")
                _focus_sap()
                _enter()
                _wait(0.8)
                return "already_maintained"

            # Show popup on RPA UI with a Remember Me checkbox
            _answer  = [None]   # True=Yes, False=No
            _remember = [False]
            import threading as _th
            _ev = _th.Event()

            def _ask():
                import tkinter as _tk
                root = _tk._default_root

                popup = _tk.Toplevel(root)
                popup.title("Material Already Maintained")
                popup.attributes("-topmost", True)
                popup.resizable(False, False)
                popup.grab_set()
                pw, ph = 420, 210
                sw = popup.winfo_screenwidth()
                sh = popup.winfo_screenheight()
                popup.geometry(f"{pw}x{ph}+{(sw-pw)//2}+{(sh-ph)//2}")
                popup.configure(bg="#1A1A1A")

                _tk.Label(
                    popup,
                    text="⚠  Material Already Maintained",
                    font=("Segoe UI", 10, "bold"), fg="#FFA726", bg="#1A1A1A"
                ).pack(pady=(16, 4))

                _tk.Label(
                    popup,
                    text=f"{material}  already maintained for  {plant_code}\n"
                         f"Skip this plant and continue to the next one?",
                    font=("Segoe UI", 9), fg="#E8E8E8", bg="#1A1A1A", justify="center"
                ).pack(pady=(0, 10))

                # Remember Me checkbox
                remember_var = _tk.BooleanVar(value=False)
                _tk.Checkbutton(
                    popup,
                    text="Remember me — auto-skip if this happens again",
                    variable=remember_var,
                    font=("Segoe UI", 8), fg="#A8A8A8", bg="#1A1A1A",
                    selectcolor="#2A2A2A", activebackground="#1A1A1A",
                    activeforeground="#A8A8A8", relief="flat",
                ).pack(pady=(0, 12))

                btn_row = _tk.Frame(popup, bg="#1A1A1A")
                btn_row.pack()

                def _yes():
                    _answer[0]   = True
                    _remember[0] = remember_var.get()
                    popup.destroy()
                    _ev.set()

                def _no():
                    _answer[0] = False
                    popup.destroy()
                    _ev.set()

                _tk.Button(
                    btn_row, text="Yes — Skip & Continue",
                    font=("Segoe UI", 9, "bold"),
                    fg="#1A1A1A", bg="#66BB6A",
                    relief="flat", padx=14, pady=6,
                    cursor="hand2", command=_yes
                ).pack(side="left", padx=(0, 8))

                _tk.Button(
                    btn_row, text="No — Stop Robot",
                    font=("Segoe UI", 9, "bold"),
                    fg="#FFFFFF", bg="#EF5350",
                    relief="flat", padx=14, pady=6,
                    cursor="hand2", command=_no
                ).pack(side="left")

            try:
                import tkinter as _tk
                _tk._default_root.after(0, _ask)
            except Exception:
                _ask()
            _ev.wait(timeout=60)

            if not _answer[0]:
                if log_fn:
                    log_fn("  → User chose to stop.", "WARN")
                raise SystemExit("User stopped after already-maintained error.")

            # User said Yes → refocus SAP, dismiss error
            _focus_sap()
            _enter()
            _wait(0.8)

            if log_fn:
                msg = f"  → Skipping {plant_code}"
                if _remember[0]:
                    msg += " (Remember Me enabled — will auto-skip next time)"
                log_fn(msg, "WARN")

            # Return sentinel — include remember flag so caller can set auto_skip
            return ("already_maintained", _remember[0])

        # ── STEP 4 ─────────────────────────────────────────
        # 2. Press Enter 3 times (navigate through page 1) — adjustable gap
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_ENTER)

        # 3. Press Tab x3 → Purchasing Group field, then fill it
        _tab(3)
        _type(prch_grp);             _wait(T_FIELD)

        # 4. Press Enter x3 (navigate to Accounting 1 tab) — adjustable gap
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_ENTER)

        # 5. Press Tab x8 → Moving Price field, then type 1
        _tab(8)
        _type(MOVING_PRICE);         _wait(T_FIELD)

        # 6. Press Enter x3 → save — adjustable gap
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_ENTER)
        _enter(); _wait(T_SAVE)

        _l(f"  ✓ {material} → {plant_code}", "OK")
        return True

    except Exception as e:
        _l(f"  ✗ {material} → {plant_code} : {e}", "ERROR")
        # Try to escape back to a clean state
        try:
            for _ in range(5):
                pyautogui.press("escape")
                time.sleep(0.3)
            # Always navigate back to MM01 cleanly after an error
            _tcode("MM01")
            _wait(1.5)
        except Exception:
            pass
        return False


# ─────────────────────────────────────────────
# MAINTAIN MATERIAL GUI CLASS
# ─────────────────────────────────────────────

# ─────────────────────────────────────────────
# FLOATING PROGRESS OVERLAY
# ─────────────────────────────────────────────

class ProgressOverlay:
    """
    Small always-on-top floating bar shown while SAP automation runs.
    Sits at the bottom of the screen so it is visible over SAP.
    Shows: current material, current plant, progress bar, ok/fail counts.
    """

    BAR_H   = 54
    BAR_W   = 480
    PAD     = 12
    BG      = "#1A1A1A"
    FG      = "#E8E8E8"
    FG2     = "#888888"
    GREEN   = "#66BB6A"
    RED     = "#EF5350"
    FILL    = "#4CAF50"
    EMPTY   = "#333333"

    def __init__(self, parent_root: tk.Tk, total: int):
        self._total   = max(total, 1)
        self._done    = 0
        self._ok      = 0
        self._fail    = 0
        self._current = "Starting..."

        # Get screen dimensions
        sw = parent_root.winfo_screenwidth()
        sh = parent_root.winfo_screenheight()

        x = (sw - self.BAR_W) // 2
        y = sh - self.BAR_H - 100  # above taskbar with comfortable margin

        self.win = tk.Toplevel(parent_root)
        self.win.overrideredirect(True)          # no title bar
        self.win.attributes("-topmost", True)    # always on top of SAP
        self.win.attributes("-alpha", 0.92)
        self.win.geometry(f"{self.BAR_W}x{self.BAR_H}+{x}+{y}")
        self.win.configure(bg=self.BG)

        # Drag support so user can reposition it
        self._drag_x = 0
        self._drag_y = 0
        self.win.bind("<ButtonPress-1>",   self._drag_start)
        self.win.bind("<B1-Motion>",        self._drag_move)

        # ── Layout ────────────────────────────────────────
        outer = tk.Frame(self.win, bg=self.BG, padx=self.PAD, pady=6)
        outer.pack(fill="both", expand=True)

        # Top row: current action + counts
        top = tk.Frame(outer, bg=self.BG)
        top.pack(fill="x")

        self._lbl_current = tk.Label(
            top, text="Starting...",
            font=("Segoe UI", 8), fg=self.FG, bg=self.BG, anchor="w")
        self._lbl_current.pack(side="left", fill="x", expand=True)

        self._lbl_counts = tk.Label(
            top, text="✓ 0  ✗ 0",
            font=("Segoe UI", 8), fg=self.FG2, bg=self.BG, anchor="e")
        self._lbl_counts.pack(side="right")

        # Bottom row: progress bar + percentage
        bot = tk.Frame(outer, bg=self.BG)
        bot.pack(fill="x", pady=(4, 0))

        self._bar_frame = tk.Frame(bot, bg=self.EMPTY, height=8)
        self._bar_frame.pack(side="left", fill="x", expand=True)
        self._bar_frame.pack_propagate(False)

        self._bar_fill = tk.Frame(self._bar_frame, bg=self.FILL, height=8, width=0)
        self._bar_fill.place(x=0, y=0, relheight=1.0, width=0)

        self._lbl_pct = tk.Label(
            bot, text="0%",
            font=("Segoe UI", 8, "bold"), fg=self.FG, bg=self.BG, width=5, anchor="e")
        self._lbl_pct.pack(side="right", padx=(6, 0))

    def update(self, material: str, plant: str, ok: int, fail: int, skipped: int = 0):
        """Call from background thread via after() to update the overlay."""
        self._done = ok + fail + skipped   # skipped counts as complete for progress
        self._ok      = ok
        self._fail    = fail
        self._skipped = skipped
        pct = int(self._done / self._total * 100)

        self._lbl_current.config(text=f"{material}  →  {plant}")

        counts = f"✓ {ok}"
        if skipped:
            counts += f"  ↷ {skipped}"
        if fail:
            counts += f"  ✗ {fail}"
        self._lbl_counts.config(
            text=counts,
            fg=self.RED if fail > 0 else self.GREEN)
        self._lbl_pct.config(text=f"{pct}%")

        # Update bar width
        try:
            bar_w = self._bar_frame.winfo_width()
            fill_w = int(bar_w * self._done / self._total)
            self._bar_fill.place(x=0, y=0, relheight=1.0, width=max(fill_w, 0))
        except Exception:
            pass

    def finish(self, ok: int, fail: int, skipped: int = 0):
        """Call when run completes."""
        parts = [f"✓ {ok} saved"]
        if skipped:
            parts.append(f"↷ {skipped} already maintained")
        if fail:
            parts.append(f"✗ {fail} failed")
        self._lbl_current.config(text="Done — " + "  ".join(parts))
        self._lbl_counts.config(text="")
        self._lbl_pct.config(text="100%")
        self._bar_fill.place(x=0, y=0, relheight=1.0, relwidth=1.0)
        self._bar_frame.config(bg=self.FILL if fail == 0 else self.RED)
        # Auto-close after 4 seconds
        self.win.after(4000, self.close)

    def close(self):
        try:
            self.win.destroy()
        except Exception:
            pass

    def _drag_start(self, e):
        self._drag_x = e.x
        self._drag_y = e.y

    def _drag_move(self, e):
        x = self.win.winfo_x() + e.x - self._drag_x
        y = self.win.winfo_y() + e.y - self._drag_y
        self.win.geometry(f"+{x}+{y}")


class MaintainMaterialGui:
    """
    Maintain Material module — launched from launcher.py.
    parent         : tk.Frame provided by launcher
    back_callback  : callable to return to main menu
    theme_manager  : ThemeManager instance (optional)
    """

    def __init__(self, parent: tk.Frame, back_callback=None, theme_manager=None):
        self.parent        = parent
        self.back_callback = back_callback
        self.theme         = theme_manager or (get_theme_manager() if theme else None)
        self._log_queue    = queue.Queue()
        self._running      = False
        self._stop_flag    = threading.Event()
        self._plant_vars    = {}
        self._plant_data    = {}
        self._plant_checked = {}   # persistent source of truth for checkbox states
        self._rpa_thread    = None

        self.parent.tk.call('tk', 'scaling', 2.0)

        global _gui_instance
        _gui_instance = self
        self._refresh_colors()
        self._build_ui()
        self._poll_log()

    def _refresh_colors(self):
        if self.theme:
            self.BG        = self.theme.get_color("bg")
            self.BG_CARD   = self.theme.get_color("bg_card")
            self.BG_INPUT  = self.theme.get_color("input_bg")
            self.TEXT      = self.theme.get_color("text_dark")
            self.TEXT2     = self.theme.get_color("text_mid")
            self.TEXT3     = self.theme.get_color("text_light")
            self.BORDER    = self.theme.get_color("border")
            self.SUCCESS   = self.theme.get_color("success")
            self.WARNING   = self.theme.get_color("warning")
            self.DANGER    = self.theme.get_color("danger")
            self.BG_DARK   = self.theme.get_color("bg_active")
        else:
            self.BG        = "#F5F5F5"
            self.BG_CARD   = "#FFFFFF"
            self.BG_INPUT  = "#FFFFFF"
            self.TEXT      = "#1A1A1A"
            self.TEXT2     = "#555555"
            self.TEXT3     = "#999999"
            self.BORDER    = "#E0E0E0"
            self.SUCCESS   = "#2E7D32"
            self.WARNING   = "#E65100"
            self.DANGER    = "#C62828"
            self.BG_DARK   = "#1A1A1A"

    # ── BUILD UI ─────────────────────────────────────────────

    def _build_ui(self):
        self.parent.configure(bg=self.BG)

        main = tk.Frame(self.parent, bg=self.BG)
        main.pack(fill="both", expand=True)

        # ── LEFT PANEL (Scrollable) ───────────────────────────
        left_outer = tk.Frame(main, bg=self.BG, width=340)
        left_outer.pack(side="left", fill="both", padx=(32, 0), pady=24)
        left_outer.pack_propagate(False)

        left_canvas = tk.Canvas(left_outer, bg=self.BG,
                                highlightthickness=0, width=320)
        left_vsb = tk.Scrollbar(left_outer, orient="vertical",
                                command=left_canvas.yview)
        left_canvas.configure(yscrollcommand=left_vsb.set)

        left_vsb.pack(side="right", fill="y")
        left_canvas.pack(side="left", fill="both", expand=True)

        left = tk.Frame(left_canvas, bg=self.BG)
        _left_win = left_canvas.create_window((0, 0), window=left, anchor="nw")

        def _on_left_cfg(e):
            left_canvas.configure(scrollregion=left_canvas.bbox("all"))

        def _on_canvas_cfg(e):
            left_canvas.itemconfig(_left_win, width=e.width)

        left.bind("<Configure>", _on_left_cfg)
        left_canvas.bind("<Configure>", _on_canvas_cfg)

        def _scroll_left(e):
            left_canvas.yview_scroll(int(-1*(e.delta/120)), "units")

        def _on_mousewheel_global(event):
            try:
                x = event.x_root
                y = event.y_root
                x1 = left_outer.winfo_rootx()
                y1 = left_outer.winfo_rooty()
                x2 = x1 + left_outer.winfo_width()
                y2 = y1 + left_outer.winfo_height()
                if x1 <= x <= x2 and y1 <= y <= y2:
                    _scroll_left(event)
            except Exception:
                pass

        left_canvas.bind("<MouseWheel>", _scroll_left)
        self.parent.bind_all("<MouseWheel>", _on_mousewheel_global)

        tk.Label(left, text="Maintain Material",
                 font=(FONT_DISP, 15, "bold"), fg=self.TEXT, bg=self.BG,
                 ).pack(anchor="w", pady=(0, 2))
        tk.Label(left, text="SAP MM01 — Create Material per Plant",
                 font=(FONT, 8), fg=self.TEXT3, bg=self.BG,
                 ).pack(anchor="w", pady=(0, 4))

        # ── SAP notice (no COM needed) ────────────────────────
        notice = tk.Frame(left, bg="#FFF8E1", highlightbackground="#FFB300",
                          highlightthickness=1)
        notice.pack(fill="x", pady=(0, 12))
        tk.Label(notice,
                 text="⚠  Open SAP GUI and log in before clicking Run.\n"
                      "   Robot types like a human — no SAP scripting needed.",
                 font=(FONT, 8), fg="#7A5400", bg="#FFF8E1",
                 justify="left", padx=8, pady=6).pack(anchor="w")

        # ── SECTION: Files ────────────────────────────────────
        self._section_label(left, "Files")
        f_card = self._card(left)

        self._field_row(f_card, "Material Codes File")
        mf_row = tk.Frame(f_card, bg=self.BG_CARD)
        mf_row.pack(fill="x", pady=(2, 6))
        self.mat_file_var = tk.StringVar()
        tk.Entry(mf_row, textvariable=self.mat_file_var,
                 font=(FONT, 8), bg=self.BG_INPUT, fg=self.TEXT,
                 insertbackground=self.TEXT, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=self.BORDER,
                 highlightcolor=self.BG_DARK, width=24,
                 ).pack(side="left", ipady=4)
        tk.Button(mf_row, text="📁", font=(FONT, 9),
                  bg=self.BG_CARD, fg=self.TEXT2, relief="flat", bd=0,
                  cursor="hand2", padx=6,
                  command=self._browse_material_file,
                  ).pack(side="left", padx=(4, 0))
        self._mat_file_lbl = tk.Label(
            f_card, text="  No file selected",
            font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD, anchor="w")
        self._mat_file_lbl.pack(anchor="w", pady=(0, 4))

        self._field_row(f_card, "Plant List  (List_Plant.xlsx)")
        pl_row = tk.Frame(f_card, bg=self.BG_CARD)
        pl_row.pack(fill="x", pady=(2, 6))
        self.plant_file_var = tk.StringVar()
        tk.Entry(pl_row, textvariable=self.plant_file_var,
                 font=(FONT, 8), bg=self.BG_INPUT, fg=self.TEXT,
                 insertbackground=self.TEXT, relief="flat", bd=0,
                 highlightthickness=1, highlightbackground=self.BORDER,
                 highlightcolor=self.BG_DARK, width=24,
                 ).pack(side="left", ipady=4)
        tk.Button(pl_row, text="📁", font=(FONT, 9),
                  bg=self.BG_CARD, fg=self.TEXT2, relief="flat", bd=0,
                  cursor="hand2", padx=6,
                  command=self._browse_plant_file,
                  ).pack(side="left", padx=(4, 0))
        self._plant_file_lbl = tk.Label(
            f_card, text="  No file selected",
            font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD, anchor="w")
        self._plant_file_lbl.pack(anchor="w", pady=(0, 4))

        # ── SECTION: Timing ───────────────────────────────────
        self._section_label(left, "Timing (seconds)")
        t_card = self._card(left)
        tk.Label(t_card,
                 text="Increase if SAP is slow on your network.\n"
                      "• MM01 load: wait after /nMM01 navigation\n"
                      "• Popup: wait for Select Views / Org Levels to open\n"
                      "• Save: wait for SAP to finish saving\n"
                      "• Between Enters: gap after Org Levels — used for\n"
                      "  the repeated Enter presses on the Basic Data /\n"
                      "  Purchasing / Accounting screens (most affected by lag)",
                 font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD,
                 justify="left").pack(anchor="w", pady=(0, 6))

        self._timing_vars = {}
        # Smart wait toggle — uses screenshot diffing instead of fixed seconds
        self._smart_wait_var = tk.BooleanVar(value=False)
        smart_row = tk.Frame(t_card, bg=self.BG_CARD)
        smart_row.pack(fill="x", pady=(2, 8))
        tk.Checkbutton(
            smart_row, text="Smart wait (detect screen change instead of fixed seconds)",
            variable=self._smart_wait_var,
            font=(FONT, 8), fg=self.TEXT2, bg=self.BG_CARD,
            selectcolor=self.BG_CARD, activebackground=self.BG_CARD,
            activeforeground=self.TEXT2, relief="flat", wraplength=260, justify="left",
        ).pack(anchor="w")
        tk.Label(t_card,
                 text="When enabled, ignores the seconds below and instead\n"
                      "watches SAP for visual changes — adapts automatically\n"
                      "to fast or slow days. Slightly slower per-step overhead.",
                 font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD,
                 justify="left").pack(anchor="w", pady=(0, 8))

        timing_rows = [
            ("After MM01 load",   "t_tcode",  T_TCODE),
            ("After popup",       "t_popup",  T_POPUP),
            ("After save",        "t_save",   T_SAVE),
            ("Between Enters",    "t_enter",  1.5),
        ]
        for label, key, default in timing_rows:
            row = tk.Frame(t_card, bg=self.BG_CARD)
            row.pack(fill="x", pady=2)
            tk.Label(row, text=label, font=(FONT, 8), fg=self.TEXT2,
                     bg=self.BG_CARD, width=18, anchor="w").pack(side="left")
            var = tk.DoubleVar(value=default)
            self._timing_vars[key] = var
            tk.Spinbox(row, textvariable=var, from_=0.3, to=10.0,
                       increment=0.1, width=5, format="%.1f",
                       font=(FONT, 8), bg=self.BG_INPUT, fg=self.TEXT,
                       relief="flat", bd=0,
                       highlightthickness=1,
                       highlightbackground=self.BORDER).pack(side="left")

        # ── SECTION: Plants ───────────────────────────────────
        self._section_label(left, "Select Plants")
        p_card = self._card(left)

        ph = tk.Frame(p_card, bg=self.BG_CARD)
        ph.pack(fill="x", pady=(0, 4))
        self._plant_count_lbl = tk.Label(
            ph, text="Load plant list file first",
            font=(FONT, 7), fg=self.TEXT3, bg=self.BG_CARD)
        self._plant_count_lbl.pack(side="left")
        tk.Button(ph, text="All", font=(FONT, 7),
                  fg=self.TEXT2, bg=self.BG_CARD, relief="flat", bd=0,
                  cursor="hand2", command=self._plant_select_all,
                  ).pack(side="right", padx=(4, 0))
        tk.Button(ph, text="Clear", font=(FONT, 7),
                  fg=self.TEXT3, bg=self.BG_CARD, relief="flat", bd=0,
                  cursor="hand2", command=self._plant_clear_all,
                  ).pack(side="right")

        search_row = tk.Frame(p_card, bg=self.BG_CARD)
        search_row.pack(fill="x", pady=(2, 4))
        tk.Label(search_row, text="🔍", font=(FONT, 9),
                 fg=self.TEXT3, bg=self.BG_CARD).pack(side="left", padx=(0, 4))
        self._plant_search_var = tk.StringVar()
        self._plant_search_var.trace_add(
            "write", lambda *a: self._rebuild_plant_checklist())
        tk.Entry(
            search_row,
            textvariable=self._plant_search_var,
            font=(FONT, 9), bg=self.BG_INPUT, fg=self.TEXT,
            insertbackground=self.TEXT, relief="flat", bd=0,
            highlightthickness=1,
            highlightbackground=self.BORDER,
            highlightcolor=self.BG_DARK,
            width=20,
        ).pack(side="left", fill="x", expand=True, ipady=4)

        tk.Frame(p_card, bg=self.BORDER, height=1).pack(fill="x", pady=(2, 4))

        self._grid_canvas = tk.Canvas(p_card, bg=self.BG_CARD,
                                       highlightthickness=0, height=160)
        grid_canvas = self._grid_canvas
        grid_vsb = tk.Scrollbar(p_card, orient="vertical",
                                command=grid_canvas.yview)
        grid_canvas.configure(yscrollcommand=grid_vsb.set)
        grid_vsb.pack(side="right", fill="y")
        grid_canvas.pack(fill="both", expand=True)

        self._plant_grid = tk.Frame(grid_canvas, bg=self.BG_CARD)
        _win = grid_canvas.create_window(
            (0, 0), window=self._plant_grid, anchor="nw")

        def _on_cfg(e):
            grid_canvas.configure(scrollregion=grid_canvas.bbox("all"))

        def _on_canvas_cfg2(e):
            grid_canvas.itemconfig(_win, width=e.width)

        self._plant_grid.bind("<Configure>", _on_cfg)
        grid_canvas.bind("<Configure>", _on_canvas_cfg2)

        def _scroll_grid(e):
            self._grid_canvas.yview_scroll(int(-1*(e.delta/120)), "units")
            return "break"

        self._grid_canvas.bind("<MouseWheel>", _scroll_grid)
        self._plant_grid.bind("<MouseWheel>", _scroll_grid)

        # ── SECTION: Selected Plants display ─────────────────
        self._section_label(left, "Selected Plants")
        sel_card = self._card(left)

        self._selected_lbl = tk.Label(
            sel_card,
            text="None selected",
            font=(FONT, 8), fg=self.TEXT3, bg=self.BG_CARD,
            anchor="w", justify="left", wraplength=280,
        )
        self._selected_lbl.pack(anchor="w")

        # ── SECTION: Info ─────────────────────────────────────
        self._section_label(left, "Info")
        i_card = self._card(left)
        for dot_c, msg in [
            (self.SUCCESS, "GT → DC=41, Copy from B100"),
            (self.WARNING, "MT → DC=21, Copy from B242"),
            (self.TEXT2,   "Outer loop = plants"),
            (self.TEXT2,   "Inner loop = material codes"),
            (self.WARNING, "SAP must be open & logged in"),
            (self.WARNING, "Do NOT touch mouse/keyboard while running"),
            (self.DANGER,   "Press Q anytime to emergency stop the robot"),
        ]:
            rf = tk.Frame(i_card, bg=self.BG_CARD)
            rf.pack(fill="x", pady=1)
            tk.Label(rf, text="●", fg=dot_c, bg=self.BG_CARD,
                     font=(FONT, 8)).pack(side="left", padx=(0, 6))
            tk.Label(rf, text=msg, fg=self.TEXT2, bg=self.BG_CARD,
                     font=(FONT, 8)).pack(side="left")

        # ── RUN / STOP buttons ────────────────────────────────
        btn_frame = tk.Frame(left, bg=self.BG)
        btn_frame.pack(fill="x", pady=(16, 0))

        self.stop_btn = tk.Button(
            btn_frame,
            text="■  Stop",
            font=(FONT_DISP, 10, "bold"),
            fg=self.BG_CARD, bg="#8B0000",
            activebackground="#C62828",
            activeforeground=self.BG_CARD,
            relief="flat", bd=0,
            padx=14, pady=10,
            cursor="hand2", state="disabled",
            command=self._on_stop,
        )
        self.stop_btn.pack(side="right", padx=(8, 0))

        self.run_btn = tk.Button(
            btn_frame,
            text="▶   Run MM01",
            font=(FONT_DISP, 11, "bold"),
            fg=self.BG_CARD, bg=self.BG_DARK,
            activebackground="#333333",
            activeforeground=self.BG_CARD,
            relief="flat", bd=0,
            padx=20, pady=10,
            cursor="hand2",
            command=self._on_run,
        )
        self.run_btn.pack(side="left", fill="x", expand=True)

        # ── RIGHT PANEL: log ──────────────────────────────────
        right = tk.Frame(main, bg=self.BG)
        right.pack(side="left", fill="both", expand=True,
                   padx=24, pady=24)

        log_hdr = tk.Frame(right, bg=self.BG)
        log_hdr.pack(fill="x", pady=(0, 6))
        tk.Label(log_hdr, text="ACTIVITY LOG",
                 font=(FONT, 8, "bold"), fg=self.TEXT3, bg=self.BG,
                 ).pack(side="left")
        tk.Button(log_hdr, text="Clear", font=(FONT, 7),
                  fg=self.TEXT3, bg=self.BG, relief="flat", bd=0,
                  cursor="hand2", command=self._clear_log,
                  ).pack(side="right")

        log_outer = tk.Frame(right, bg=self.BORDER, bd=0)
        log_outer.pack(fill="both", expand=True)
        log_inner = tk.Frame(log_outer, bg=self.BG_CARD)
        log_inner.pack(fill="both", padx=1, pady=1)

        self.log_box = scrolledtext.ScrolledText(
            log_inner,
            bg=self.BG_CARD, fg=self.TEXT2,
            font=(FONT, 9),
            relief="flat", bd=0,
            state="disabled", wrap="word",
            padx=12, pady=10,
        )
        self.log_box.pack(fill="both", expand=True)
        self.log_box.tag_config("OK",    foreground=self.SUCCESS)
        self.log_box.tag_config("ERROR", foreground=self.DANGER)
        self.log_box.tag_config("WARN",  foreground=self.WARNING)
        self.log_box.tag_config("INFO",  foreground=self.TEXT2)

        self._write_log("Maintain Material ready — pure keyboard mode.", "INFO")
        self._write_log("Killswitch: press Q at any time to stop the robot.", "WARN")
        self._write_log("1. Browse material codes file", "INFO")
        self._write_log("2. Browse plant list file", "INFO")
        self._write_log("3. Open SAP GUI and log in", "INFO")
        self._write_log("4. Select plants → Run MM01", "INFO")

    # ── UI HELPERS ───────────────────────────────────────────

    def _section_label(self, parent, text):
        f = tk.Frame(parent, bg=self.BG)
        f.pack(fill="x", pady=(12, 4))
        tk.Frame(f, bg=self.BG_DARK, width=3, height=14).pack(
            side="left", padx=(0, 6))
        tk.Label(f, text=text.upper(),
                 font=(FONT, 8, "bold"), fg=self.TEXT, bg=self.BG,
                 ).pack(side="left")

    def _card(self, parent):
        outer = tk.Frame(parent, bg=self.BORDER)
        outer.pack(fill="x", pady=(0, 4))
        inner = tk.Frame(outer, bg=self.BG_CARD, padx=14, pady=10)
        inner.pack(fill="x", padx=1, pady=1)
        return inner

    def _field_row(self, parent, label):
        tk.Label(parent, text=label,
                 font=(FONT, 9), fg=self.TEXT2, bg=self.BG_CARD, anchor="w",
                 ).pack(anchor="w", pady=(6, 0))

    # ── FILE BROWSERS ─────────────────────────────────────────

    def _browse_material_file(self):
        path = filedialog.askopenfilename(
            title="Select Material Codes Excel File",
            filetypes=[("Excel File", "*.xlsx *.xls"),
                       ("All Files", "*.*")],
        )
        if not path:
            return
        self.mat_file_var.set(path)
        try:
            codes = load_material_codes(path)
            self._mat_file_lbl.config(
                text=f"  ✔ {len(codes)} codes — {os.path.basename(path)}",
                fg=self.SUCCESS)
            self._write_log(
                f"Material file: {os.path.basename(path)} "
                f"→ {len(codes)} codes", "OK")
        except Exception as e:
            self._mat_file_lbl.config(text=f"  ⚠ {e}", fg=self.WARNING)
            self._write_log(f"Material file error: {e}", "ERROR")

    def _browse_plant_file(self):
        path = filedialog.askopenfilename(
            title="Select Plant List Excel File",
            filetypes=[("Excel File", "*.xlsx *.xls"),
                       ("All Files", "*.*")],
        )
        if not path:
            return
        self.plant_file_var.set(path)
        try:
            self._plant_data = load_plant_data(path)
            self._plant_file_lbl.config(
                text=f"  ✔ {len(self._plant_data)} plants — "
                     f"{os.path.basename(path)}",
                fg=self.SUCCESS)
            self._write_log(
                f"Plant list: {os.path.basename(path)} "
                f"→ {len(self._plant_data)} plants", "OK")
            self._rebuild_plant_checklist()
        except Exception as e:
            self._plant_file_lbl.config(text=f"  ⚠ {e}", fg=self.WARNING)
            self._write_log(f"Plant file error: {e}", "ERROR")

    # ── PLANT CHECKLIST ───────────────────────────────────────

    def _rebuild_plant_checklist(self):
        # Flush current BooleanVar values into the persistent checked dict FIRST
        # so that plants not in the current filter keep their last known state.
        for p, v in self._plant_vars.items():
            self._plant_checked[p] = v.get()

        for w in self._plant_grid.winfo_children():
            w.destroy()

        self._plant_vars = {}

        search = ""
        try:
            search = self._plant_search_var.get().strip().lower()
        except Exception:
            pass

        all_plants = sorted(self._plant_data.keys())

        filtered = [
            p for p in all_plants
            if search in p.lower()
            or search in self._plant_data[p]["name"].lower()
        ] if search else all_plants

        for i, plant in enumerate(filtered):
            info  = self._plant_data[plant]
            is_mt = info["type"] == "MT"
            prch  = info.get("prch_grp", "")
            grp_txt = f"({prch})" if prch else ""
            label = f"{plant}  {info['name']}  {grp_txt}".strip()
            color = self.WARNING if is_mt else self.TEXT2

            # Use persistent checked dict — default True for new plants
            var = tk.BooleanVar(value=self._plant_checked.get(plant, True))
            self._plant_vars[plant] = var

            cb = tk.Checkbutton(
                self._plant_grid, text=label,
                variable=var,
                font=(FONT, 8),
                fg=color, bg=self.BG_CARD,
                selectcolor=self.BG_CARD,
                activebackground=self.BG_CARD,
                activeforeground=color,
                relief="flat", bd=0,
            )
            cb.grid(row=i, column=0, sticky="w", padx=4, pady=1)
            # Bind scroll directly on each checkbutton — fixes scroll only working on white bar
            cb.bind("<MouseWheel>", lambda e: self._grid_canvas.yview_scroll(
                int(-1*(e.delta/120)), "units") or "break")
            # Update selected label whenever a checkbox is toggled
            var.trace_add("write", lambda *a: self._update_selected_label())

        total = len(all_plants)
        shown = len(filtered)
        gt    = sum(1 for p in self._plant_data.values() if p["type"] == "GT")
        mt    = sum(1 for p in self._plant_data.values() if p["type"] == "MT")

        if search:
            self._plant_count_lbl.config(
                text=f"{shown} of {total} plants  (GT:{gt}  MT:{mt})",
                fg=self.TEXT2)
        else:
            self._plant_count_lbl.config(
                text=f"{total} plants  (GT:{gt}  MT:{mt})",
                fg=self.TEXT2)

        self._update_selected_label()

    def _update_selected_label(self):
        """Refresh the selected plants display label."""
        try:
            # Flush current vars into checked dict first
            for p, v in self._plant_vars.items():
                self._plant_checked[p] = v.get()
            selected = [p for p in sorted(self._plant_data.keys())
                        if self._plant_checked.get(p, False)]
            if not selected:
                self._selected_lbl.config(text="None selected", fg=self.TEXT3)
            else:
                self._selected_lbl.config(
                    text="  ".join(selected),
                    fg=self.TEXT2)
        except Exception:
            pass

    def _plant_select_all(self):
        for p in self._plant_data:
            self._plant_checked[p] = True
        for v in self._plant_vars.values():
            v.set(True)
        self._update_selected_label()

    def _plant_clear_all(self):
        for p in self._plant_data:
            self._plant_checked[p] = False
        for v in self._plant_vars.values():
            v.set(False)
        self._update_selected_label()

    def _get_selected_plants(self) -> list:
        # Flush current var states first so hidden (filtered-out) plants are included
        for p, v in self._plant_vars.items():
            self._plant_checked[p] = v.get()
        return [p for p in sorted(self._plant_data.keys())
                if self._plant_checked.get(p, False)]

    # ── CLEAR LOG ─────────────────────────────────────────────

    def _clear_log(self):
        self.log_box.configure(state="normal")
        self.log_box.delete("1.0", "end")
        self.log_box.configure(state="disabled")

    # ── RUN / STOP ────────────────────────────────────────────

    def _on_run(self):
        if self._running:
            return

        mat_file   = self.mat_file_var.get().strip()
        plant_file = self.plant_file_var.get().strip()
        plants     = self._get_selected_plants()

        if not mat_file or not os.path.exists(mat_file):
            messagebox.showwarning("Missing Input",
                                   "Please select a material codes file.")
            return
        if not plant_file or not os.path.exists(plant_file):
            messagebox.showwarning("Missing Input",
                                   "Please select a plant list file.")
            return
        if not plants:
            messagebox.showwarning("Missing Input",
                                   "Please select at least one plant.")
            return

        # Confirm SAP is open — just like rpa_gui.py does
        if not messagebox.askyesno(
            "SAP Ready?",
            "Before running:\n\n"
            "  1. SAP GUI is open and you are logged in\n"
            "  2. You are on the SAP Easy Access main screen\n"
            "  3. You will NOT touch the mouse or keyboard\n\n"
            "Is SAP ready to go?\n\n"
        ):
            return

        self._stop_flag.clear()
        self._running = True
        self.run_btn.configure(state="disabled",
                               text="⏳  Running...", bg="#555555")
        self.stop_btn.configure(state="normal")

        self._write_log("━" * 45, "INFO")
        self._write_log("Maintain Material started  [keyboard-only mode]", "OK")
        self._write_log(f"Plants   : {', '.join(plants)}", "INFO")
        self._write_log(f"Material : {os.path.basename(mat_file)}", "INFO")
        self._write_log("━" * 45, "INFO")

        # Snapshot timing values before thread starts
        timing = {k: v.get() for k, v in self._timing_vars.items()}
        timing["smart_wait"] = self._smart_wait_var.get()

        def _run():
            self._execute_mm01(mat_file, plant_file, plants, timing)

        self._rpa_thread = threading.Thread(target=_run, daemon=True)
        self._rpa_thread.start()

    def _on_stop(self):
        self._stop_flag.set()
        _KILL_FLAG.set()
        self._write_log("─" * 45, "WARN")
        self._write_log("KILL — stopping robot immediately...", "WARN")
        self._write_log("─" * 45, "WARN")
        self.stop_btn.configure(state="disabled")
        # Hard-kill: inject SystemExit into the RPA thread right now
        if self._rpa_thread and self._rpa_thread.is_alive():
            _raise_in_thread(self._rpa_thread.ident)

    def _show_overlay(self, total_ops: int):
        try:
            self._overlay = ProgressOverlay(self.parent.winfo_toplevel(), total_ops)
        except Exception as e:
            self._overlay = None

    def _reset_buttons(self):
        self._running = False
        self._rpa_thread = None
        # Close overlay if still showing
        try:
            if self._overlay:
                self._overlay.close()
                self._overlay = None
        except Exception:
            pass
        self.run_btn.configure(
            state="normal", text="▶   Run MM01", bg=self.BG_DARK)
        self.stop_btn.configure(state="disabled")
        # Force UI update so it doesn't appear frozen
        self.parent.update_idletasks()

    # ── SAP EXECUTION ─────────────────────────────────────────

    def _execute_mm01(self, mat_file: str, plant_file: str,
                      plants: list, timing: dict):
        """
        Background thread: run MM01 for all plants × all materials.
        Pure pyautogui — no COM whatsoever.
        """
        # Apply user timing overrides to module-level constants
        global T_TCODE, T_POPUP, T_SAVE, T_ENTER, SMART_WAIT
        T_TCODE    = timing.get("t_tcode", T_TCODE)
        T_POPUP    = timing.get("t_popup", T_POPUP)
        T_SAVE     = timing.get("t_save",  T_SAVE)
        T_ENTER    = timing.get("t_enter", T_ENTER)
        SMART_WAIT = timing.get("smart_wait", False)

        def _l(msg, level="INFO"):
            self._log(msg, level)

        _KILL_FLAG.clear()
        _register_killswitch()
        _watchdog_active.set()
        _start_corner_watchdog()
        _l("Killswitch active — press Q or move cursor to top-left corner to stop instantly.", "WARN")

        # Initialize counters BEFORE try block so finally can always reference them
        total_ok      = 0
        total_fail    = 0
        total_skipped = 0

        try:
            _l("Loading material codes...")
            materials = load_material_codes(mat_file)
            if not materials:
                _l("No material codes found in file!", "ERROR")
                return
            _l(f"{len(materials)} material codes loaded", "OK")

            _l("Loading plant data...")
            plant_data = load_plant_data(plant_file)
            _l(f"{len(plant_data)} plants loaded", "OK")

            # Verify SAP is visible before starting
            if not _focus_sap():
                _l("SAP window not found! Open SAP and log in first.", "ERROR")
                return

            _l("SAP window found — starting robot...", "OK")
            _l("Do NOT touch keyboard or mouse until done!", "WARN")

            # Count total operations for progress bar
            total_ops = sum(
                len(plants) if not e["dc"] else
                sum(1 for p in plants if p in plant_data and plant_data[p]["dc"] == e["dc"])
                for e in materials
            )

            # Show floating overlay
            self._overlay = None
            self.parent.after(0, lambda: self._show_overlay(total_ops))

            # Navigate to MM01 once — SAP returns here automatically after every save
            _tcode("MM01")
            _wait(T_TCODE)

            prev_mat   = ""     # tracks previous material code across all plants
            prev_type  = ""     # tracks previous plant type (GT/MT) for smart org fill
            org_open   = False  # True when Org Levels popup is already open (after skip)
            auto_skip  = False  # True when user checked Remember Me

            # ── OUTER LOOP: materials ────────────────────────
            for mat_idx, entry in enumerate(materials, start=1):
                if self._stop_flag.is_set() or _KILL_FLAG.is_set():
                    _l("Stopped by user.", "WARN")
                    break

                material = entry["code"]
                mat_dc   = entry["dc"]   # blank = all plants

                _l("─" * 40, "INFO")
                _l(f"Material [{mat_idx}/{len(materials)}] {material}", "INFO")
                prev_type = ""   # reset type tracking for each new material

                # ── INNER LOOP: plants ─────────────────────
                for plant_code in plants:
                    if self._stop_flag.is_set() or _KILL_FLAG.is_set():
                        _l("Stopped by user.", "WARN")
                        break

                    if plant_code not in plant_data:
                        _l(f"  Plant {plant_code} not in plant list — skip", "WARN")
                        continue

                    info     = plant_data[plant_code]
                    ptype    = info["type"]
                    plant_dc = info["dc"]

                    # Skip if material DC doesn't match this plant's DC
                    if mat_dc and mat_dc != plant_dc:
                        _l(f"  Skip {plant_code} (DC={plant_dc}) — material DC={mat_dc}", "INFO")
                        continue

                    _l(f"  → {plant_code} [{ptype}] DC={plant_dc} | PrchGrp={info['prch_grp']} | CopyFrom={info['copy_plant']}", "INFO")

                    result = _mm01_one(
                        material              = material,
                        plant_info            = info,
                        plant_code            = plant_code,
                        log_fn                = _l,
                        prev_material         = prev_mat,
                        org_levels_open       = org_open,
                        auto_skip_maintained  = auto_skip,
                        prev_type             = prev_type,
                    )
                    prev_mat  = material
                    prev_type = info["type"]
                    org_open  = False   # reset for next iteration

                    # Handle tuple return (already_maintained, remember_flag)
                    if isinstance(result, tuple) and result[0] == "already_maintained":
                        org_open   = True
                        prev_mat   = ""
                        prev_type  = info["type"]   # type is still same after skip
                        total_skipped += 1
                        if result[1]:   # remember me checked
                            auto_skip = True
                            _l("  Remember Me set — will auto-skip maintained plants from now on", "WARN")
                    elif result == "already_maintained":
                        # auto_skip path returns plain string
                        org_open  = True
                        prev_mat  = ""
                        prev_type = info["type"]   # type is still same after skip
                        total_skipped += 1
                    elif result:
                        total_ok  += 1
                    else:
                        total_fail += 1

                    # Update overlay from main thread
                    _ok, _fail, _skip = total_ok, total_fail, total_skipped
                    _mat, _plt = material, plant_code
                    self.parent.after(0, lambda m=_mat, p=_plt, o=_ok, f=_fail, s=_skip:
                        self._overlay.update(m, p, o, f, s) if self._overlay else None)

                    time.sleep(0.3)

                # ── End of inner loop (all plants done for this material) ──
                if org_open:
                    # The last plant was "already maintained" — its error was
                    # already dismissed once inside _mm01_one, leaving the
                    # Org Levels popup open with that plant's data.
                    # Press Escape once → cancels Org Levels, back to MM01
                    # initial screen, ready for the next material.
                    _focus_sap()
                    pyautogui.press("escape"); time.sleep(1.0)    # cancel → back to MM01
                    org_open  = False
                    prev_mat  = ""
                    prev_type = ""
                    _l("  All selected plants already maintained for this material — "
                       "returning to MM01 for next material", "WARN")

            _l("━" * 45, "INFO")
            _l("SUMMARY", "INFO")
            _l(f"Total: ✓ {total_ok} created | ✗ {total_fail} failed | ↷ {total_skipped} already maintained",
               "OK" if total_fail == 0 else "WARN")
            _l("━" * 45, "INFO")

        except (Exception, SystemExit) as e:
            if isinstance(e, SystemExit):
                _l("Robot killed immediately.", "WARN")
            elif "fail-safe" in str(e).lower() or "failsafe" in str(e).lower():
                _l("Fail-safe triggered (top-left corner) — robot killed.", "WARN")
            else:
                _l(f"FATAL: {e}", "ERROR")

        finally:
            # Release any keys pyautogui might be holding mid-action
            try:
                pyautogui.keyUp("delete")
                pyautogui.keyUp("ctrl")
                pyautogui.keyUp("shift")
            except Exception:
                pass
            _unregister_killswitch()
            _stop_corner_watchdog()
            if _KILL_FLAG.is_set():
                _l("Robot stopped.", "WARN")
            _KILL_FLAG.clear()
            _ok, _fail, _skip = total_ok, total_fail, total_skipped
            self.parent.after(0, lambda o=_ok, f=_fail, s=_skip:
                self._overlay.finish(o, f, s) if self._overlay else None)
            self.parent.after(0, self._reset_buttons)

    # ── LOGGING ──────────────────────────────────────────────

    def _log(self, msg: str, level: str = "INFO"):
        ts = datetime.now().strftime("%H:%M:%S")
        self._log_queue.put((level, f"{ts}  {msg}"))

    def _write_log(self, msg: str, level: str = "INFO"):
        self.log_box.configure(state="normal")
        self.log_box.insert("end", msg + "\n", level)
        self.log_box.see("end")
        self.log_box.configure(state="disabled")

    def _poll_log(self):
        try:
            while True:
                level, msg = self._log_queue.get_nowait()
                self._write_log(msg, level)
        except queue.Empty:
            pass
        self.parent.after(80, self._poll_log)


# ─────────────────────────────────────────────
# STANDALONE ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    root.title("Maintain Material — MM01")
    root.geometry("1100x680")
    root.configure(bg="#F5F5F5")
    frame = tk.Frame(root, bg="#F5F5F5")
    frame.pack(fill="both", expand=True)
    MaintainMaterialGui(frame, back_callback=root.destroy)
    root.mainloop()